# -*- coding: utf-8 -*-
"""
reporting/tooltip_formatter.py
────────────────────────────────────────────────────────────────────────────────
Centralised tooltip system for the NSE/BSE Analyser Excel dashboard.

Session 16 (final push): replaces the legacy "💡 short\\n\\nfull" comment style
with a polished, structured hover + visible ⓘ header cue + rich Tooltip
Reference sheet.

Public API:
    TIPS                            — dict[header_name] -> (short, full)
    format_tooltip(header, short, full) -> str
    apply_tooltips(ws, header_row, col_headers,
                   add_cue=True, ref_anchors=None) -> None
    build_reference_sheet(wb, after_sheet=None) -> dict[header -> row]

Design notes:
- Openpyxl writes only the legacy (VML) comment format, so we cannot style
  the hover box itself (yellow background, Tahoma font, callout arrow are
  drawn by Excel). Everything we CAN control — text layout, icons, box
  dimensions — is polished here.
- The visible ⓘ cue on the header tells users a tooltip exists without
  requiring them to hover blindly.
- The Tooltip Reference sheet is the main "modern" visual upgrade — full
  colour, borders, typography — linked from every header via hyperlink.
"""

from __future__ import annotations
from typing import Dict, Iterable, Optional, Tuple

from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════════
# TIP DATA — single source of truth for every column explanation
# ════════════════════════════════════════════════════════════════════════════
TIPS: Dict[str, Tuple[str, str]] = {
    # ── Identity & classification ───────────────────────────────────────────
    "Symbol": ("Stock ticker / trading symbol", "NSE or BSE code used for placing orders."),
    "Company Name": ("Full legal entity name", "Registered company name on the exchange."),
    "Company": ("Full company name", "Legal registered company name."),
    "Sector": ("Industry classification",
               "Used for rotation analysis and peer comparison.\n"
               "Stage 2 sectors offer the best entry opportunities."),
    "Exchange": ("DUAL_LISTED = best liquidity",
                 "DUAL_LISTED: Both NSE + BSE (best liquidity).\n"
                 "NSE_ONLY: Good liquidity. BSE_ONLY: lower, check volume.\n"
                 "BSE_SME: very low liquidity — high impact cost.\n"
                 "When BSE bhavcopy unavailable, an allowlist of dual-listed\n"
                 "Nifty-100 + popular mid-caps tags DUAL_LISTED."),
    "Cap Category": ("LARGE=safest · MICRO=speculative",
                     "BUY cutoffs: LARGE≥60, MID≥63, SMALL≥66, MICRO≥70.\n"
                     "LARGE: >₹20,000Cr. MID: ₹5,000–20,000Cr.\n"
                     "SMALL: ₹500–5,000Cr. MICRO: <₹500Cr."),
    "BSE Code": ("6-digit BSE scrip code",
                 "Used on BSE terminal. NSE uses Symbol; BSE uses this numeric code.\n"
                 "Helpful when the same company has different tickers across exchanges."),

    # ── Verdicts & composite scores ─────────────────────────────────────────
    "Verdict": ("BUY · OVERVALUED · WATCHLIST · NEUTRAL · AVOID",
                "BUY: clears cap-tier threshold + MoS>−10%.\n"
                "OVERVALUED: clears threshold but MoS blocks (wait for pullback).\n"
                "WATCHLIST: in watch band below BUY threshold.\n"
                "NEUTRAL: above AVOID floor but below WATCHLIST min.\n"
                "AVOID: Score<38. Cutoffs: LARGE≥60, MID≥63, SMALL≥66, MICRO≥70.\n"
                "Tech Override: MoS relaxes to −20% if Score≥70+ST=BUY+Stage 2.\n"
                "Confidence dots (Session 24): ●●●HIGH ●●○MED ●○○LOW.\n"
                "v10.9: Score also carries a forensic quality adjustment\n"
                "from Altman Z / Earn Quality / ND-EBITDA / Int Coverage\n"
                "(max +8 / min −10) — see Score /100 tooltip for detail.\n"
                "v10.17: BUY requires ≥3 of 5 sub-score dimensions to be\n"
                "informed (real data fired). Otherwise demoted to\n"
                "WATCHLIST ●●● (thin data).\n"
                "v12.0: NEUTRAL stocks now appear in the Full Dashboard\n"
                "regardless of fundamentals. Stage 3 already enforces the\n"
                "quality gate; the prior 'exceptional NEUTRAL only' filter\n"
                "was removed because it silently shrank the dashboard\n"
                "below 100 rows whenever Gemini quota was exhausted.\n"
                "v12.1: reconciler hotfix — Exchange tag is now derived\n"
                "purely from ISIN match (not symbol-match) for stocks with\n"
                "no ISIN, so non-equity tickers (indices, ETFs) no longer\n"
                "get falsely tagged DUAL_LISTED."),
    "Score /100": ("≥70 strong · ≥60 watch · <38 avoid",
                   "Weighted composite (0-100):\n"
                   "Fundamental 35% + Technical 30% + EarlyEntry 15%\n"
                   "+ Sentiment 10% + Safety 10% + MoS adj (−10 to +12)\n"
                   "+ Spike bonus (max +10) + Early Mover +5 − Risk −10.\n"
                   "v10.9 Forensic Quality Adjustment (max +8, min −10):\n"
                   "  Altman Z ≥3: +3 | <1.8: −5\n"
                   "  Earn Quality HIGH: +2 | LOW: −3\n"
                   "  ND/EBITDA <1: +1 | >5: −2\n"
                   "  Int Coverage >5x: +2 | <1.5x: −3\n"
                   "BUY cutoffs by cap: LARGE≥60, MID≥63, SMALL≥66, MICRO≥70.\n"
                   "Session 24: sentiment redistributes when no paid signals;\n"
                   "spike bonus capped at +3 if fundamental<55."),
    "Early Entry /100": ("≥50 Early Mover · ≥35 Ahead of Consensus",
                         "Detects stocks 4–12 weeks BEFORE institutional coverage.\n"
                         "Key triggers: Vol+RSI, Trend Confluence, Momentum,\n"
                         "52W Breakout, Deep Value+BUY, Inst Footprint,\n"
                         "Score Convergence, FII/Promoter Accum, Dual-Listed.\n"
                         "≥50 EARLY MOVER, ≥35 Ahead of Consensus, <35 Emerging.\n"
                         "Low EE on Gold is OK — two archetypes: MOMENTUM and VALUE.\n"
                         "See Tooltip Reference for all 12 signals."),
    "Spike Score /6": ("≥2 notable · ≥4 strong · 6 very rare",
                       "Six momentum triggers — how many fire simultaneously.\n"
                       "T1: CMP near 52W High + vol>2×. T2: MACD+ST=BUY + vol>1.5×.\n"
                       "T3: ADX>25 + delivery>60%. T4: RSI 45–65 + vol>2×.\n"
                       "T5: vol>3× + delivery>60%. T6: 2w_chg>3% + 2w>4w + vol>1.5×.\n"
                       "Suppressed to 0 if pledge>20% or Altman/Beneish flags fire.\n"
                       "Low Spike on a Gold stock is fine (VALUE archetype)."),
    "Spike /6": ("≥2 notable | ≥4 strong | 6 very rare",
                 "Six momentum triggers — how many fire simultaneously.\n"
                 "Suppressed to 0 if pledge>20% or Altman/Beneish flags active.\n"
                 "Low Spike on Gold is OK — pure-value candidates qualify\n"
                 "on fundamentals + MoS + safety without momentum signals."),
    "Storm Score /10": ("≥8 Storm Safe | ≥5 Moderate | <5 High Risk",
                        "Defensive quality — how safe in a market crash?\n"
                        "Beta<0.8 +2, D/E<0.3 +2, FCF positive +2, Div yield>2% +1,\n"
                        "Rev growth>10% +1, Margin Expansion +1,\n"
                        "Promoter QoQ up +1, FII buying 3Q +1."),
    "Storm /10": ("≥8 Storm Safe | ≥5 Moderate | <5 High Risk",
                  "Defensive quality — higher score = more resilient in downturns."),
    # v12.5: F-Score column renamed to Piotroski F /9 (matches Full
    # Dashboard); the tooltip entry that used to live here was removed.
    # The Piotroski F /9 tooltip below now serves both sheets.

    # ── Price / momentum ────────────────────────────────────────────────────
    "CMP (₹)": ("Current market price",
                "Compare with CFV. CMP<CFV=undervalued (buying opportunity)."),
    "Day Chg %": ("Today's move | >+3% strong | <−3% weak",
                  ">5% may trigger circuit rules on some counters."),
    "52W High (₹)": ("52-week peak — CMP near here = breakout",
                    "CMP near 52W High = breakout territory (T1 spike fires)."),
    "52W Low (₹)": ("52-week floor — value or continued decline",
                   "Verify fundamentals before buying near 52W Low."),
    "Vol Spike (×50D)": ("≥2× unusual | ≥3× institutional",
                         "1×: Normal | 1.5–2×: Above avg | 2–3×: Unusual | >3×: Major event"),
    "Delivery %": ("≥60% institutional · <40% speculative",
                   "Share of traded volume actually delivered (not intraday).\n"
                   "≥70% strong institutional conviction. 40–70% mixed.\n"
                   "<40% mostly speculative — caution.\n"
                   "Sentiment impact: >70%: +4, >60%: +2, <30%: −3."),
    "Beta": ("<0.8 defensive | >1.2 volatile",
            "Beta<0.8: Less volatile — good in downturns (+2 Storm)."),
    "Chg% [2-Weekly]": ("2-week return | >3% strong momentum",
                        ">2%: EE Momentum signal fires. 2W>4W = accelerating (bullish)."),
    "Chg% [4-Weekly]": ("4-week return | 2W>4W = acceleration",
                        "2W>4W = accelerating | 2W<4W = decelerating"),
    "Chg% [6-Weekly]": ("6-week trend",
                        "Positive + rising across 2W/4W/6W/8W = sustained uptrend."),
    "Chg% [8-Weekly]": ("8-week return | best for trend direction",
                        "Consistently positive = confirmed uptrend."),
    "Chg% [2-Wk]": ("2-week return | >3% strong",
                    ">2%: EE Momentum signal fires."),
    "Chg% [4-Wk]": ("4-week return | 2W>4W=acceleration",
                    "2W>4W=accelerating | 2W<4W=decelerating."),
    "Chg% [6-Wk]": ("6-week trend",
                    "Positive + rising across 2W/4W/6W/8W = sustained uptrend."),
    "Chg% [8-Wk]": ("8-week return | best for trend direction",
                    "Consistently positive = confirmed uptrend."),

    # ── Fair value & valuation ──────────────────────────────────────────────
    "CFV (₹)": ("Composite Fair Value (7 models)",
                "Weights: M1 DCF 30%, M2 Graham 15%, M3 PE 20%, M4 PB 15%,\n"
                "M5 EV/EBITDA 10%, M6 DDM 5%, M7 PEG 5%.\n"
                "CMP<CFV = undervalued. CMP>CFV = overvalued.\n"
                "Session 19 safety caps: M1 DCF capped at 4× CMP,\n"
                "composite CFV capped at 3× CMP.\n"
                "Prevents low-beta stocks with tiny terminal-value denominators\n"
                "from producing implausible 5×–10× fair values.\n"
                "v12.2: M3/M4/M5 sector resolution rewritten — production sector\n"
                "strings (Basic Materials, Industrials, Communication Services,\n"
                "Consumer Cyclical/Defensive, Financial Services, Real Estate)\n"
                "now canonicalize via SECTOR_ALIASES before substring matching\n"
                "against benchmark multipliers, so 31 of 100 stocks no longer\n"
                "silently fall through to defaults.\n"
                "v12.3 Round 2: M5 now uses proper EV math (annual_ebitda ×\n"
                "sector_mult − net_debt) when full data available; banks/NBFCs/\n"
                "insurance correctly skip M5 entirely. M7 PEG_BENCHMARK = 1.0\n"
                "made explicit constant (was implicit)."),
    "FV Low (₹)": ("Conservative FV = CFV × 0.85",
                  "CMP below FV Low = very deeply undervalued."),
    "FV High (₹)": ("Optimistic FV = CFV × 1.15",
                   "CMP above FV High = significantly overvalued."),
    "MoS %": (">25% strong buy · <−15% overvalued",
              "Margin of Safety = (CFV − CMP) / CMP × 100.\n"
              ">40% Exceptional (+12), >25% Strong (+8), 10–25% Adequate (+4).\n"
              "−15 to −30% Overvalued (−5), <−30% Significant premium (−10).\n"
              "Effectively capped near 200% because CFV is capped at 3× CMP\n"
              "(Session 19 safety net). If MoS ≈200%, verify inputs rather\n"
              "than treating it as a guaranteed bargain."),
    "MoS Label": ("Valuation summary (* = capped CFV, † = thin-FV evidence)",
                  "EXCEPTIONAL >40% | STRONG >25% | ADEQUATE >10% | THIN 0-10%\n"
                  "SLIGHT PREMIUM −10% to 0% | SIGNIFICANT PREMIUM <−10%\n"
                  "v12.5: trailing `*` (e.g., 'EXCEPTIONAL*') means CFV was\n"
                  "clipped to 3× CMP — the underlying models projected even\n"
                  "higher upside but the safety cap fired. Treat with extra\n"
                  "scrutiny: the model average is unusually optimistic.\n"
                  "v12.6: trailing `†` (e.g., 'EXCEPTIONAL†') means CFV was\n"
                  "based on fewer than 3 valuation models (M1–M7) firing —\n"
                  "the FV evidence is thin. The CFV value is still shown so\n"
                  "you can decide for yourself, but the automatic +score bonus\n"
                  "(+4 to +12) is suppressed in composite_score to prevent\n"
                  "thin-evidence false BUYs. Markers can stack: '*†' means\n"
                  "BOTH conditions fired — treat with extreme caution."),
    # v10.8: 'Upside to FV %' and 'Upside %' removed entirely — duplicated MoS %
    # (same formula, same number). MoS % is the single source of truth now.
    "M1: DCF FV (₹)": ("Discounted Cash Flow fair value — 30% weight in CFV",
                       "Projects 10 years of free cash flow and discounts at WACC.\n"
                       "Best for mature, cash-generating businesses.\n\n"
                       "Session 19 cap: M1 is limited to 4× CMP to prevent unrealistic\n"
                       "valuations when low-beta inputs (e.g., SBIN with β=0.2) inflate\n"
                       "the DCF output to absurd levels. Capping here also protects the\n"
                       "composite CFV (which is separately capped at 3× CMP)."),
    "M2: Graham FV (₹)": ("Benjamin Graham formula — 15% weight in CFV",
                          "Classic value formula: FV = √(22.5 × EPS × BVPS)\n"
                          "Conservative — rewards stable earnings + book value."),
    "M3: PE FV (₹)": ("Sector-relative P/E fair value — 20% weight in CFV",
                      "FV = EPS × sector median P/E.\n"
                      "Shows 0 when EPS negative.\n"
                      "v12.2: 28 sector benchmarks (Banks 18, Tech 30, FMCG 45,\n"
                      "Steel 10, Metals 12, Realty 25, Telecom 22 etc.).\n"
                      "Production sector strings canonicalize through\n"
                      "SECTOR_ALIASES — e.g., 'Basic Materials' → Metals (PE 12),\n"
                      "'Industrials' → Infra (PE 22), 'Communication Services' →\n"
                      "Telecom (PE 22). 'General' falls to default 25 by design."),
    "M4: PB FV (₹)": ("Price/Book fair value — 15% weight in CFV",
                      "FV = Book Value × sector median P/B.\n"
                      "Most useful for banks, NBFCs, and asset-heavy businesses.\n"
                      "v12.2: 28 sector benchmarks. Same SECTOR_ALIASES path as\n"
                      "M3 — Basic Materials gets PB 1.5 (Metals), Industrials\n"
                      "gets PB 2.5 (Infra), Communication Services gets PB 2.5\n"
                      "(Telecom). BVPS fallback derives from close/PB if missing."),
    "M5: EV FV (₹)": ("EV/EBITDA fair value — 10% weight in CFV",
                      "Capital-structure neutral — useful when debt levels vary widely.\n"
                      "v12.3 Round 2: proper EV-based formula primary —\n"
                      "  fair_per_share = CMP × ((annual_ebitda × sector_mult\n"
                      "                            − net_debt) / mcap_cr)\n"
                      "Three-tier dispatch:\n"
                      "  Tier 1 (proper): when q_ebitda_cr + total_debt_cr +\n"
                      "    cash_cr + mcap_cr all available\n"
                      "  Tier 2 (shortcut): legacy v12.2 formula as fallback\n"
                      "  Tier 3 (skip): Banks/NBFCs/Insurance always (EV/EBITDA\n"
                      "    isn't meaningful for financials)\n"
                      "v12.2: 28-sector benchmarks (Tech 22, FMCG 30, Steel 5,\n"
                      "Banks 12, Realty 12) + SECTOR_ALIASES canonicalization."),
    "M6: DDM FV (₹)": ("Dividend Discount Model — 5% weight in CFV",
                       "Only meaningful for consistent dividend payers.\n"
                       "Gordon Growth: FV = D1 / (r - g).\n"
                       "v12.2 fix: removed the 2% growth floor that previously\n"
                       "inflated FV for stocks with declining earnings. A stock\n"
                       "with pat_yoy=-20 now correctly gets 0% div growth (was\n"
                       "getting 2% before). Yield range gate: 0.1% < yield < 15%."),
    "M7: PEG FV (₹)": ("Growth-adjusted P/E fair value — 5% weight in CFV",
                       "Rewards genuine growth companies at reasonable multiples.\n"
                       "Formula: EPS × growth_pct × PEG_BENCHMARK.\n"
                       "v12.3 Round 2: PEG_BENCHMARK = 1.0 made an explicit named\n"
                       "constant (Lynch's rule: stock fair when PEG = 1.0).\n"
                       "Mathematically identical to v12.2 but constant is now\n"
                       "tunable — value-tilted setups could use 0.8, growth\n"
                       "setups 1.2.\n"
                       "Shows 0 when earnings growth negative.\n"
                       "v12.2 unit guard: skips when growth < 1.0 (catches\n"
                       "decimal-fraction unit error)."),

    # ── Ratios ──────────────────────────────────────────────────────────────
    "P/E TTM": ("<20 cheap | 20–40 fair | >40 expensive",
                "Score: ≤20 = +12 | ≤40 = +7 | 41–60 = 0 | >60 = −8 | ≥500 = 0 (neutral).\n"
                "Source: yfinance .info['trailingPE'].\n"
                "Display: v10.16 shows '—' when raw value ≥ 500 — stocks\n"
                "with near-zero earnings produce mathematical PE ratios in\n"
                "the thousands (AMAGI hit 1,981 pre-fix). These aren't real\n"
                "'expensive' signals — they're arithmetic noise from tiny-EPS\n"
                "denominators. Real quality businesses never exceed PE 500,\n"
                "so the threshold preserves every plausible premium-growth\n"
                "case while honestly flagging 'valuation not meaningful here'.\n"
                "Scoring: v10.16 treats clamped noise (pe_num ≥ 500) as\n"
                "neutral, not penalised — 'unknown' not 'expensive'."),
    "P/E": ("<20 cheap | 20–40 fair | >40 expensive",
            "Price to Earnings (TTM). Score: ≤20 = +12 | ≤40 = +7 | >60 = −8.\n"
            "v10.16: display '—' when raw ≥ 500; scoring treats as neutral.\n"
            "(Same rationale as P/E TTM.)"),
    "Earn Yield %": (">6% undervalued vs bonds",
                     "EPS/CMP × 100. >6%: Cheap. Compare to 10Y bond yield."),
    "P/CF": ("<15 value | >25 expensive",
            "More reliable than P/E (cash harder to fake)."),
    "PEG Ratio": ("<1 undervalued | >2 expensive",
                  "P/E / Growth. <1: Undervalued (Peter Lynch favourite).\n"
                  "Source: yfinance .info['pegRatio'] primary; 4-tier\n"
                  "fallback computes PE / PAT-growth, PE / Rev-growth,\n"
                  "or PE / sustainable-growth (ROE × retention).\n"
                  "Display: v10.16 shows '—' when any tier yields ≥ 50.\n"
                  "PEG beyond 50 means P/E divided by near-zero growth —\n"
                  "pure arithmetic noise, not a 'very expensive' signal.\n"
                  "Even extreme glamour stocks rarely exceed PEG of 10."),
    "PEG": ("<1 undervalued | >2 expensive",
            "P/E / Growth. <1: Undervalued.\n"
            "v10.16: display '—' when value ≥ 50."),
    "P/B": ("<2 value | >5 expensive",
            "<1: Below asset value | >5: Only justified by very high ROE.\n"
            "Source: yfinance .info['priceToBook'].\n"
            "Display: v10.16 shows '—' when raw value ≥ 500 — tiny book-value\n"
            "denominators produce mathematical P/B in the thousands that\n"
            "carry no real 'overvalued' signal."),
    "P/S": ("<3 cheap | >10 expensive",
            "Useful when P/E unavailable.\n"
            "Source: yfinance .info['priceToSalesTrailing12Months'].\n"
            "Display: v10.16 shows '—' when raw value ≥ 500."),
    "EV/EBITDA": ("<12 value | >20 expensive",
                  "IT: 20 | Pharma: 18 | FMCG: 30 | Banks: 12 | Metals: 8.\n"
                  "Source: yfinance .info['enterpriseToEbitda'].\n"
                  "Display: v10.16 shows '—' when raw value ≥ 500 — near-zero\n"
                  "EBITDA produces mathematical EV/EBITDA in the thousands\n"
                  "(RHETAN hit 1,352 pre-fix). These aren't real 'expensive'\n"
                  "signals — they mean EBITDA ≈ 0, valuation undefined.\n"
                  "Even richly-valued growth names rarely exceed 50x."),
    "ROE %": (">20% excellent | <10% weak",
             "Score: >20% = +12 | >10% = +6 | <5% = −5.\n"
             "Source: yfinance .info['returnOnEquity'] × 100 when\n"
             "available; else derived as Earnings Yield × P/B when\n"
             "both are positive (ROE ≈ EPS/BVPS).\n"
             "Display: v10.15 FIX #1 now stores as a FLOAT (was a\n"
             "quoted string pre-v10.15, which broke Excel sorting,\n"
             "filtering, and conditional formatting on this column).\n"
             "Shows '—' when neither direct nor derivable."),
    "ROCE %": (">15% good capital allocation",
               "Return on Capital Employed. ROCE > cost of capital = value-creating.\n"
               "Not available from yfinance; derived from ROE + leverage."),
    "ROA %": (">10% efficient | <5% poor",
             "Low for Banks/Utilities is normal.\n"
             "Source: yfinance .info['returnOnAssets'] × 100 when\n"
             "available; else derived as ROE / (1 + D/E).\n"
             "Display: v10.15 FIX #1 stores as FLOAT (was string).\n"
             "v12.4: clamped to ±100 % — yfinance occasionally returns\n"
             "absurd values for finance/NBFC stocks (M&MFIN had 189 %)."),
    "Gross Mgn %": (">40% strong moat | >20% decent",
                    "Score: >40% = +8 | >20% = +4\n"
                    "v12.4: clamped to [0, 100] %."),
    "EBITDA Mgn %": (">25% excellent | >15% good",
                     ">30%: Excellent | >20%: Good | <10%: Tight\n"
                     "v12.4: clamped to ±100 %."),
    "NPM %": (">15% excellent | <5% thin",
             "Score: >15% = +8 | >5% = +4 | <0% = −8\n"
             "v12.4: clamped to ±100 %. yfinance occasionally feeds\n"
             "values >100 % on thin-revenue / one-time-gain rows\n"
             "(DGCONTENT 126 %, AMAGI 189 % pre-clamp)."),
    "NPM Q (latest) %": ("Most recent quarter margin vs TTM",
                 "Q(latest) > NPM(TTM): margins accelerating.\n"
                 "Source: (Quarterly PAT / Quarterly Revenue) × 100\n"
                 "from yfinance quarterly_income_stmt.\n"
                 "v12.6 (#11): renamed from 'NPM Q1 %' for chronological\n"
                 "clarity — old labels read L→R as Q1 Q2 Q3 suggesting\n"
                 "chronological, but Q1 was actually the LATEST quarter.\n"
                 "Display: v10.15 caps at ±500% — tiny-revenue denominator\n"
                 "(₹0.13 Cr quarterly rev for micro-caps like EMAMIREAL)\n"
                 "produced −762% NPM in prior runs. Same clamp pattern as\n"
                 "v10.14 CAGR fix."),
    "NPM Q-1 %": ("Previous quarter margin",
                 "Track Q-2 → Q-1 → Q(latest) trend. Source:\n"
                 "quarterly_income_stmt 2nd-most-recent quarter.\n"
                 "v12.6 (#11): renamed from 'NPM Q2 %'.\n"
                 "Display: v10.15 caps at ±500%\n"
                 "(EMAMIREAL Q-1 hit −387% pre-clamp)."),
    "NPM Q-2 %": ("Two quarters ago — rising = Margin Expansion",
                 "Rising Q-2 → Q-1 → Q(latest) triggers Margin Expansion = YES.\n"
                 "Source: quarterly_income_stmt 3rd-most-recent quarter.\n"
                 "v12.6 (#11): renamed from 'NPM Q3 %'.\n"
                 "Display: v10.15 caps at ±500% (EMAMIREAL Q-2 hit\n"
                 "−845% pre-clamp)."),
    "Margin Expansion": ("YES = 3 consecutive qtrs of rising NPM",
                         "Score: Fundamental +5 | Safety +3 | Storm +1."),

    # ── Growth ──────────────────────────────────────────────────────────────
    "Rev CAGR 1Y %": (">20% high growth | >10% good",
                      "1Y > 3Y CAGR = growth accelerating.\n"
                      "Source: latest FY revenue ÷ prior FY revenue − 1\n"
                      "(discrete annual fiscal-year growth, from yfinance's\n"
                      "income_stmt table). Differs from 'Rev YoY %' which uses\n"
                      "rolling TTM growth from .info['revenueGrowth']. A large\n"
                      "divergence between the two usually means the company\n"
                      "is mid-year with a strong/weak recent quarter.\n"
                      "Display: v10.14 caps at ±500% to prevent tiny-base\n"
                      "CAGR distortions (e.g., old base of ₹1 Cr)."),
    "Rev CAGR 3Y %": (">15% strong | >8% decent",
                      "Score: >15% = +5 | >8% = +3.\n"
                      "Source: latest FY revenue ÷ FY-3 revenue,\n"
                      "then ^(1/3) − 1. Requires ≥4 annual columns in\n"
                      "yfinance income_stmt; shows '—' if fewer (newer\n"
                      "IPOs). 3Y is more reliable than 1Y because it\n"
                      "smooths single-quarter distortions.\n"
                      "Display: v10.14 caps at ±500%."),
    "PAT CAGR 1Y %": (">20% strong earnings momentum",
                      "1Y > 3Y = accelerating profitability.\n"
                      "Source: latest FY PAT ÷ prior FY PAT − 1.\n"
                      "Shows '—' if any FY had loss (CAGR undefined).\n"
                      "Compare with 'PAT YoY %' which uses rolling TTM.\n"
                      "Display: v10.14 caps at ±500%."),
    "PAT CAGR 3Y %": (">20% compounder | >10% good",
                      "Score: >20% = +8 | >10% = +4.\n"
                      "Source: latest FY PAT ÷ FY-3 PAT, ^(1/3) − 1.\n"
                      "Shows '—' if any FY had loss OR if income_stmt\n"
                      "has <4 annual columns (common for new listings).\n"
                      "Best long-horizon compounding signal.\n"
                      "Display: v10.14 caps at ±500%."),
    "EBITDA CAGR 1Y %": (">15% strong operating growth",
                         "Score: >15% = +4 | >8% = +2.\n"
                         "Source: latest FY EBITDA ÷ prior FY EBITDA − 1.\n"
                         "EBITDA recovery from near-zero base can produce\n"
                         "outsized percentages — v10.14 caps at ±500% to\n"
                         "prevent tiny-base noise. Real sustained EBITDA\n"
                         "CAGR rarely exceeds 100%."),
    "Rev YoY %": (">10% growing | <0% declining",
                  "Score: >15% = +5 | >8% = +3 | <−5% = −4.\n"
                  "Source: yfinance .info['revenueGrowth'] × 100 — this\n"
                  "is TRAILING TWELVE MONTH (TTM) growth, a rolling\n"
                  "4-quarter comparison. NOT the same as 'Rev CAGR 1Y %'\n"
                  "which uses discrete fiscal years. TTM and CAGR diverge\n"
                  "most in insurance/NBFC stocks (premium accounting) and\n"
                  "companies with recent restructuring.\n"
                  "Display: v10.14 caps at ±500% to filter yfinance junk\n"
                  "signals on micro-cap stocks."),
    "PAT YoY %": (">20% strong | >10% good",
                  "Score: >20% = +8 | >10% = +4 | >0% = +2 | <−10% = −7.\n"
                  "Source: yfinance .info['earningsGrowth'] × 100 — rolling\n"
                  "TTM earnings growth. Same TTM vs fiscal-year distinction\n"
                  "as Rev YoY %.\n"
                  "Display: v10.14 caps at ±500%."),
    "Q3 Rev (₹Cr)": ("Latest quarter revenue",
                     "Rising QoQ = business growing.\n"
                     "Source: yfinance quarterly_income_stmt 3rd-most-recent\n"
                     "column (i.e., quarter two-before-latest). Divided by\n"
                     "₹1 Cr = ₹10 million conversion. Shows '—' if quarterly\n"
                     "data unavailable (rare for NSE/BSE main-board)."),
    "Q3 PAT (₹Cr)": ("Latest quarter net profit",
                     "Positive and growing = healthy earnings.\n"
                     "Source: quarterly_income_stmt PAT row, 3rd column.\n"
                     "Shows '—' if loss-making that quarter (NULL in DB)."),
    "Q3 EBITDA (₹Cr)": ("Latest quarter operating profit",
                        "Q3 EBITDA Margin = Q3 EBITDA / Q3 Rev — compare vs TTM.\n"
                        "Source: quarterly_income_stmt EBITDA row (plain, not\n"
                        "normalized) 3rd column. Falls back to 'operating\n"
                        "income' row if EBITDA not reported.\n"
                        "Shows '—' if neither available."),

    # ── Balance sheet ───────────────────────────────────────────────────────
    "D/E Ratio": ("<0.3 excellent | >2 risky | >3 danger",
                  "<0.3: +8 Fundamental, +2 Storm | 0.3–1: +4 | >2: −10 | >3: BS ALERT"),
    "D/E": ("<0.3 excellent | >2 risky",
            "<0.3: +8 Fundamental, +2 Storm | >2: −10 | >3: BS ALERT"),
    "ND/EBITDA": ("<2 safe | >4 risky",
                  "Net Debt / EBITDA = years to repay debt from operating cash.\n"
                  "Safety score impact:\n"
                  "  <1 (nearly debt-free): +5\n"
                  "  <0: Net cash position (separate +6 via Cash vs Debt check)"),
    "Int Coverage": (">5 safe | <2 danger",
                     "EBIT / Interest Expense. Higher = easier to service debt.\n"
                     "Safety score impact:\n"
                     "  >10: Very safe → +5\n"
                     "  5–10: Safe → +2\n"
                     "  <2: Danger | <1: Critical"),
    "Current Ratio": (">2 healthy | 1–2 adequate | <1 risky",
                      "Score: >2 = +6 | >1.5 = +3 | <1 = −7"),
    "Quick Ratio": (">1 safe | <0.5 risky",
                    "(Current Assets − Inventory) / Current Liabilities"),
    "Cash (₹Cr)": ("Higher = stronger safety net",
                   "Safety score impact:\n"
                   "  Cash > Total Debt (NET CASH COMPANY) → +6\n"
                   "Cash reserves are the first line of defence in a downturn."),
    "Total Debt (₹Cr)": ("Lower = better | 0 = ideal",
                          "Compare with Cash and EBITDA."),
    "FCF (₹Cr)": (">0 cash generator | <0 cash consuming",
                  "Score: >0 = +2 Storm, +3 Safety | <0: cash consuming."),
    "FCF Yield %": (">6% undervalued | >3% fair",
                    "Score: >6% = +6 | >3% = +3 | <0% = −5"),
    "CCC Days": ("Lower / negative = more efficient ('—' for finance sector)",
                 "Cash Conversion Cycle = DIO + DSO − DPO.\n"
                 "Negative CCC = collects cash before paying suppliers.\n"
                 "Source: (inventory + receivables − payables) / revenue × 365\n"
                 "from yfinance balance_sheet + income_stmt (COGS used\n"
                 "when available, else revenue as proxy).\n"
                 "Display: v10.15 caps at ±500 days. Tiny-revenue denominators\n"
                 "previously produced 16,821 days (EMAMIREAL = 46 years —\n"
                 "arithmetic noise, not signal). If totalRevenue < 1000 the\n"
                 "computation is skipped entirely and shown as 0 or '—'.\n"
                 "v12.5: skipped entirely for Banks/NBFCs/HFCs/Insurance —\n"
                 "the metric is meaningless for finance-sector stocks (no\n"
                 "inventory; loans aren't 'receivables' in the same sense).\n"
                 "TATACAP showed 7,739 days, FUSION 3,216 in prior runs."),
    "Div Yield %": (">2% good income | >4% check sustainability",
                    ">2%: +1 Storm Score.\n"
                    "Display: '—' means company pays no dividend (v10.9).\n"
                    "Zero vs '—' distinction: '—' = no dividend policy;\n"
                    "0 would mean the stock pays but at 0% yield (rare)."),
    "Payout Ratio %": ("40–60% balanced | >80% unsustainable",
                       "30–60%: Balanced | >80%: Check FCF coverage."),
    "Capex / Rev %": ("<5% asset-light | >15% capital-heavy",
                      "<3%: Asset-light = high FCF."),

    # ── Ownership ───────────────────────────────────────────────────────────
    "Promoter %": (">50% aligned | <20% concern",
                   "Score: >50% = +5 | >35% = +2 | <20% = −3"),
    "Pro QoQ Δ": (">0.3% buying signal | negative = selling",
                  "Promoter shareholding change vs previous quarter.\n"
                  "Sentiment score impact:\n"
                  "  >+0.5%: Promoters buying → +5\n"
                  "  <−0.5%: Promoters selling → −5\n"
                  "Storm score also +1 if >0.3%.\n"
                  "Display: '—' when no real delta can be computed.\n"
                  "v10.15 FIX #5: three states clearly distinguished —\n"
                  "  Real number (incl. 0.0) = measured delta\n"
                  "  '—' = no ≥90-day history OR prior value was the\n"
                  "       backfill literal 0.0 (yfinance cannot supply\n"
                  "       real QoQ; populates after ~3 months of runs).\n"
                  "Pre-v10.15 showed 0 for 83/86 stocks indistinguishably —\n"
                  "now honestly '—' unless backed by computed data."),
    "Pledge %": ("0% ideal | >10% watch | >20% RED FLAG",
                 "Safety score impact:\n"
                 "  0%: Clean cap structure → +4 (rewarded, not just neutral)\n"
                 "  10–20%: Watch → −7\n"
                 "  >20%: RED FLAG → −15, plus suppresses ALL spike signals.\n"
                 "Display: v10.15 FIX #6 now shows '—' when value is 0,\n"
                 "because pledge data is only in BSE corporate filings which\n"
                 "have no free API. Zero pledge is structurally indistinguishable\n"
                 "from 'unknown pledge' on free-tier, so honest display is '—'.\n"
                 "If a paid BSE feed is added, real 0 and real pledge will\n"
                 "display as numbers. Score gates still fire on numeric values\n"
                 "only — '—' treated as 0 for guard purposes (safe default)."),
    "Pledge Direction": ("FALLING = positive | RISING = risk",
                          "Trend in promoter share pledge over time.\n"
                          "IMPROVING: Pledge dropped — promoters repaying loans (positive).\n"
                          "DETERIORATING: Pledge rose — more shares pledged (risk).\n"
                          "STABLE: Pledge unchanged at a non-zero level.\n"
                          "— : No pledge data (yfinance has no free source; needs BSE filings)."),
    "FII %": (">15% institutional backed",
              ">25%: High global interest | Rising FII = strong signal."),
    "FII QoQ Δ": (">1% accumulation | <−1% selling",
                  "+8 EE if >1% | +1 Storm if >0.3%.\n"
                  "Display: '—' when no ≥90-day history in shareholding\n"
                  "table (v10.4/v10.9). Populates after ~3 months of runs."),
    "DII %": (">10% domestic confidence",
              "Rising DII + FII = dual institutional accumulation = bullish.\n"
              "Source: NSE corporate-info JSON API (heldPercentInstitutions\n"
              "in yfinance is FII+DII combined; DII alone needs NSE).\n"
              "Display: v10.15 FIX #6 shows '—' when value is 0, because\n"
              "NSE corp-info API is blocked on cloud IPs (common for GH\n"
              "Actions runs). Zero DII is indistinguishable from 'API\n"
              "blocked' on free-tier, so honest display is '—'. Real DII\n"
              "values display as numbers when API responds."),
    "DII QoQ Δ": (">0.5% domestic accumulation",
                  "Domestic institutional (MF/insurance) holding change.\n"
                  "Sentiment score impact:\n"
                  "  >+0.5%: Strong DII accumulation → +6\n"
                  "  +0.3–0.5%: Moderate accumulation → +4\n"
                  "  <−0.3%: DII distribution → −3.\n"
                  "Display: '—' when DII % source unavailable (NSE corp-info\n"
                  "API often blocked on cloud IPs) OR no ≥90-day history."),
    "Public Float %": (">50% good liquidity | <20% manipulation risk",
                       "<20%: Volatile, easier to manipulate."),

    # ── Forensics ───────────────────────────────────────────────────────────
    "Piotroski F /9": ("≥7 strong · ≤3 weak",
                       "9 criteria across Profitability (4), Leverage/Liquidity (3),\n"
                       "Efficiency (2). 8–9 Excellent, 6–7 Good, ≤3 Avoid.\n"
                       "Safety impact: F≥7 → +6, F=5–6 → +3.\n"
                       "Computed from free yfinance data (Session 14+20);\n"
                       "typical run distribution: 4–8, quality stocks 6–8."),
    "Altman Z": (">2.99 safe | <1.81 distress zone (capped at 10)",
                 "<1.81: Triggers anti-trigger guard (Spike suppressed).\n"
                 "Requires balance-sheet inputs (working capital, retained\n"
                 "earnings, EBIT, total assets, total liabilities) which\n"
                 "yfinance free data doesn't provide. Column displays '—'\n"
                 "(em-dash) for stocks without the required BS feed.\n"
                 "v12.5: clamped at 10 — values >10 (ALIVUS 14.69, GOPAL\n"
                 "17.27, CPEDU 26.70 in prior runs) are typically\n"
                 "unit-mismatch artefacts in the X4 component (mcap /\n"
                 "total_liab) where one figure is in raw rupees and the\n"
                 "other in Cr. Z>7 already signals exceptional safety."),
    "Beneish M": ("<−2.22 honest | >−2.22 possible manipulation",
                  ">−2.22: Triggers anti-trigger guard (Spike suppressed).\n"
                  "Requires net income, cash flow from operations, and total\n"
                  "assets from the balance sheet — yfinance free data doesn't\n"
                  "provide these. Column displays '—' (em-dash) for stocks\n"
                  "without the required BS feed."),
    "Earn Quality": ("HIGH = cash-backed earnings",
                     "CFO / PAT ratio bucketed into HIGH / MODERATE / LOW.\n"
                     "HIGH (≥0.8): Cash flow matches profits — healthy earnings.\n"
                     "MODERATE (0.5-0.8): Some divergence — worth monitoring.\n"
                     "LOW (<0.5): Accounting concern — profits aren't backed by cash.\n"
                     "— : PAT is zero/negative (ratio undefined)."),

    # ── Catalysts ───────────────────────────────────────────────────────────
    "OB/Bill Ratio": (">1 strong pipeline | >3 excellent visibility",
                      ">3: 3+ year revenue visibility. For infra/defence/engineering."),
    "Pipeline Vis": ("HIGH = strong revenue visibility",
                     "HIGH: Strong order book or recurring revenue."),
    "L1 Wins 90D": ("Recent govt contract wins", ">3: Active and winning bidder."),
    "L1 Est (₹Cr)": ("Estimated govt contract value",
                     "Higher = more near-term revenue locked in."),
    "New Mkt Entry": ("YES = new revenue stream potential",
                      "YES: New geography or product launch."),

    # ── Signals ─────────────────────────────────────────────────────────────
    "Early Signals": ("Signals fired today — more = higher conviction",
                      "EE + spike signals: VOL SURGE + RSI | TREND CONFLUENCE\n"
                      "TECHNICAL BREAKOUT | INSTITUTIONAL FOOTPRINT | 52W BREAKOUT"),
    "Sector Stage": ("Stage 2 = best entry | Stage 4 = avoid / exit",
                     "STAGE 1 EARLY ACCUM: Smart money entering\n"
                     "STAGE 2 CONFIRMED UPTREND: All signals aligned (BEST ENTRY)\n"
                     "STAGE 3 MOMENTUM PEAK: Overbought (caution)\n"
                     "STAGE 4 DISTRIBUTION: Smart money exiting (avoid)\n"
                     "NEUTRAL: No clear stage yet (insufficient signals)."),
    "Smart Money": ("ACCUMULATION = institutional buying",
                    "Possible values:\n"
                    "HIGH DELIVERY BUYING: delivery >70% + good volume.\n"
                    "RSI ACCUMULATION ZONE: RSI 50–65 + sideways action.\n"
                    "INST ACCUMULATION: block deals + delivery uptick.\n"
                    "INSIDER BUYING: promoter/director buys on record.\n"
                    "FII/PROMOTER BUYING: QoQ holding up (paid data).\n"
                    "NEUTRAL: none of the above."),

    # ── Technical ───────────────────────────────────────────────────────────
    "SMA 200": ("CMP > SMA200 = bull trend confirmed",
                "200-Day SMA — the long-term trend anchor.\n"
                "Technical score impact (new):\n"
                "  CMP > SMA200 × 1.02 (clearly above) → +3\n"
                "  CMP < SMA200 × 0.98 (clearly below) → −3\n"
                "Golden Cross (50>200) = major buy | Death Cross = major sell."),
    "Supertrend": ("BUY=uptrend | SELL=downtrend | NEUTRAL=sideways",
                   "BUY: Price > SMA20 + 0.5×ATR14 | SELL: Price < SMA20 − 0.5×ATR14"),
    "ADX": (">25 strong trend | <20 weak / sideways",
            "Measures TREND STRENGTH (not direction).\n"
            "Technical score impact:\n"
            "  >30: Established trend → +7\n"
            "  25–30: Strong → +5\n"
            "  20–25: Moderate → +2\n"
            "  <20: Weak / sideways → 0"),
    "RSI (14)": ("45–65 sweet spot · >70 overbought · <30 oversold",
                 "Technical score impact:\n"
                 "60–70 SWEET SPOT → +10 (strongest reward).\n"
                 ">70 Overbought → +8 (rewarded but capped).\n"
                 "50–60 Mildly bullish → +4.\n"
                 "40–50 Mildly bearish → −4. <40 Bearish → −8.\n"
                 "45–65 is also the entry zone for spike trigger T4."),
    "MACD Signal": ("BUY = bullish crossover | SELL = bearish",
                    "BUY: +6 Technical | SELL: −6 Technical"),
    "Stoch %K": ("20–40 accumulation zone | >80 overbought",
                 "20–40: Accum zone (+5) | >80: Overbought (−3)"),
    "MFI": (">60 money inflow | <30 outflow",
            ">60: +4 Technical | <30: −3 Technical"),
    "OBV Signal": ("RISING = accumulation | FALLING = distribution",
                   "RISING: +4 Technical | FALLING: −4 Technical"),
    "Above VWAP": ("YES = institutional support | NO = weak",
                   "YES: +4 Technical | NO: −2 Technical"),
    "Chart Pattern": ("Today's candle pattern from OHLC",
                      "Detected from OHLC + previous close. Possible values:\n"
                      "BULLISH / BEARISH CANDLE (close vs open vs prev close).\n"
                      "DOJI (indecision), HAMMER (bullish reversal signal).\n"
                      "HANGING MAN / SHOOTING STAR (bearish reversal).\n"
                      "UPPER / LOWER CIRCUIT (hit price band).\n"
                      "NEUTRAL (sideways). '—' (OHLC incomplete)."),
    "Pattern": ("Today's candle pattern from OHLC",
                "Same as Chart Pattern, abbreviated for Gold sheet.\n"
                "Values: BULLISH / BEARISH CANDLE, DOJI, HAMMER,\n"
                "HANGING MAN, SHOOTING STAR, UPPER / LOWER CIRCUIT,\n"
                "NEUTRAL, '—'."),

    # ── Support / resistance ────────────────────────────────────────────────
    "Support 1 (₹)": ("Nearest support = buy zone floor",
                      "20-day rolling low. Breach = bearish. Used for Stop Loss."),
    "Support 2 (₹)": ("Major floor — prior 52-week low",
                      "Lowest price in the 252 trading days BEFORE the\n"
                      "most recent 20 — i.e., the prior major floor.\n"
                      "v12.4: excludes the last 20 days so a fresh\n"
                      "breakdown to a new low doesn't make S1 == S2.\n"
                      "Falls back to '—' for stocks with < 80 days\n"
                      "of price history."),
    "Resist 1 (₹)": ("First resistance = Target 1",
                     "20-day rolling high. Breakout with volume = bullish."),
    "Resist 2 (₹)": ("Major supply ceiling — prior 52-week high",
                     "Highest price in the 252 trading days BEFORE the\n"
                     "most recent 20 — i.e., the prior major ceiling.\n"
                     "v12.4: excludes the last 20 days so a fresh\n"
                     "breakout to a new high doesn't make R1 == R2\n"
                     "(the v10.9 logic collapsed for 87.9 % of rows\n"
                     "when the 52-week max landed inside the last 20 d).\n"
                     "Falls back to '—' for stocks with < 80 days\n"
                     "of price history."),

    # ── Balance sheet health ────────────────────────────────────────────────
    "BS Health Flag": ("HEALTHY = safe | WATCH = monitor | ALERT = danger",
                       "HEALTHY: No red flags\n"
                       "WATCH: One concern (D/E>2 or low liq or neg FCF)\n"
                       "ALERT: Serious (pledge>20% or D/E>3 or leveraged + neg FCF)"),
    "BS Health Note": ("Explains the health flag",
                       "Examples: NET CASH COMPANY | HIGH D/E 2.5× | NEGATIVE FCF | HIGH PLEDGE"),

    # ── Trade plan ──────────────────────────────────────────────────────────
    "Entry Range (₹)": ("Ideal buy zone = CMP ± 0.5 × ATR",
                         "Avoid chasing if CMP moves significantly above upper bound."),
    "Stop Loss (₹)": ("Exit if CMP closes below this level",
                       "Never risk >2–3% of portfolio per trade."),
    "Target 1 (₹)": ("First target = Resistance 1",
                     "Book 30–50% of position here. R:R should be >1:2."),
    "Target 2 (₹)": ("Second target = Resistance 2",
                     "Hold remainder after Target 1."),
    "Target 3 (₹)": ("Final target = Fair Value (CFV)",
                     "High MoS stocks can give 20–50% upside."),
    "Time Horizon": ("How long to hold",
                     "SHORT TERM: 2–4 weeks (BUY + Spike≥2)\n"
                     "POSITIONAL: 1–3 months (Score≥68 + ST=BUY)\n"
                     "LONG TERM: 3–12 months (Score≥72 + no spike)"),
    "Horizon": ("How long to hold",
                "SHORT TERM: 2–4 weeks | POSITIONAL: 1–3 mo | LONG TERM: 3–12 mo"),
    "Risk Level": ("LOW=safest | VERY HIGH=speculative only",
                   "LOW: High score + low beta + low D/E + no pledge\n"
                   "MEDIUM: Acceptable | HIGH: Small/micro | VERY HIGH: Speculative"),
    "R:R Ratio": ("Aim for >1:2 · Higher = better risk/reward",
                  "R:R = (T1 − Entry mid) / (Entry mid − SL).\n"
                  ">2 Excellent, 1–2 Acceptable, <1 Avoid.\n"
                  "Session 22: T1 auto-derived to ensure R:R ≥ 2.0:\n"
                  "T1 = max(Entry + 2×risk_distance, CFV-weighted target).\n"
                  "High-CFV stocks get value-anchored targets; others get\n"
                  "pure risk-symmetric targets."),

    # ── Narrative / AI ──────────────────────────────────────────────────────
    "Key Catalyst": ("Primary near-term growth driver",
                     "Product launch, order win, policy tailwind, expansion."),
    "News Sentiment": ("POSITIVE = tailwind | NEGATIVE = headwind",
                       "AI-analysed sentiment from recent company news.\n"
                       "Sentiment score impact:\n"
                       "  POSITIVE: Favourable → +4\n"
                       "  NEUTRAL: No news movement → 0\n"
                       "  NEGATIVE: Headwinds → −5"),
    "Primary Risk": ("Biggest downside risk", "Always read before investing."),
    "SEBI Flags": ("NONE = clean | Any flag = investigate first",
                   "Any flag = investigate before buying."),
    "View Analysis Summary": ("Gemini AI investor narrative (150–250 words)",
                              "Business quality, ratios, risks, catalysts, verdict rationale.\n"
                              "Generated fresh each trading day."),

    # ── Alert Log specific ──────────────────────────────────────────────────
    "Date": ("Date alert was generated", "Trading day of the alert."),
    "Time (IST)": ("Time pipeline ran",
                   "Typically 05:00–05:30 IST on trading days."),
    "Alert Type": ("Type of signal that fired",
                   "🔔 SPIKE FIRED: momentum trigger(s) lit up today.\n"
                   "⭐ EARLY MOVER DETECTED: early-entry score ≥ 70.\n"
                   "⬇ SCORE DEGRADED: composite dropped ≥ 3 vs yesterday."),
    "Trigger Detail": ("Full details of the alert",
                       "Shows all signals fired and current score context."),
    "Prev Score": ("Yesterday's composite score",
                   "Compare with New Score to see improvement / decline."),
    "New Score": ("Today's composite score",
                  "Rising = improving | Falling = deteriorating."),
    "Score Δ": ("Score change vs yesterday | +ve = improving",
                "+ve: Getting stronger | −ve: Getting weaker."),
    "Action Required": ("What to do with this stock today",
                        "CONSIDER ENTRY: BUY + MoS>10% + Score≥65.\n"
                        "MONITOR FOR ENTRY: BUY, weaker conviction.\n"
                        "STRONG STOCK — WAIT FOR PULLBACK: OVERVALUED verdict.\n"
                        "VOLUME ALERT — INVESTIGATE: vol spike ≥ 3× avg.\n"
                        "EARLY MOVER — ACCUMULATE: early_entry ≥ 70.\n"
                        "SCORE IMPROVING — WATCH (Δ ≥ +3) /\n"
                        "SCORE DECLINING — CAUTION (Δ ≤ −3).\n"
                        "REVIEW FOR EXIT: Score < 30.\n"
                        "MONITOR CLOSELY: default."),

    # Session 19: reference-only entries (not actual Excel column headers).
    # These don't get hover tooltips attached to any cell, but they DO show up
    # as reference cards on the 📖 Tooltip Reference sheet, giving the user
    # a quick-reference card for concepts that the Glossary also documents.
    # Keeping Glossary + Tooltip Reference sheet in sync matters — both are
    # discovery surfaces for the same knowledge.
    "Gold-Tier Filter": (
        "11-condition filter for Gold – Early Movers sheet (v10.11)",
        "ALL 11 conditions must be true for Gold qualification:\n"
        "1. Verdict = BUY (not WATCHLIST).\n"
        "2. Composite Score ≥ 70.\n"
        "3. 15% ≤ MoS ≤ 100% (real upside, not phantom).\n"
        "4. Storm Score ≥ 5 (defensively sound).\n"
        "5. RSI ≤ 70 (not already overbought).\n"
        "6. BS Health Flag ≠ ALERT.\n"
        "7. Pledge % ≤ 10 (clean capital structure).\n"
        "8. Not spike-suppressed (anti-trigger guard clear).\n"
        "v10.11 forensic quality gates (3 new):\n"
        "9. Altman Z ≥ 1.8 or missing (not in distress zone).\n"
        "10. Earn Quality ≠ LOW (no accounting concern).\n"
        "11. Int Coverage ≥ 1.5× or missing (can service interest).\n"
        "Missing forensic data passes the v10.11 gates — small caps without\n"
        "forensic feeds aren't unfairly excluded. Daily count varies 0-10+."),
    "CFV Safety Cap": (
        "Composite Fair Value capped at 3× CMP",
        "CFV is capped at 3× Current Market Price as a safety net, which\n"
        "means MoS never exceeds approximately 200%. Prevents any single\n"
        "model misbehaving (e.g., DCF on a low-beta stock with tiny\n"
        "terminal-value denominator) from distorting the composite.\n\n"
        "If you see MoS near 200%, treat it as 'deeply undervalued by\n"
        "model — verify inputs' rather than a guaranteed bargain."),
    # Session 27: Removed duplicate "M1 DCF Safety Cap" entry. Its content
    # (4× CMP cap + WACC 10% floor) is already documented in the main
    # "M1: DCF FV (₹)" entry above (added in Session 26). A separate entry
    # created a confusing duplicate row in the Tooltip Reference sheet
    # (rows 24 & 25 both showed M1-related info).
}


# ════════════════════════════════════════════════════════════════════════════
# GROUP-HEADER TOOLTIPS (section-level, above data-column tips)
# ════════════════════════════════════════════════════════════════════════════
# Hover over the merged section headers in Full Dashboard row 3 and Gold
# sheet row 4 (IDENTITY / SCORES / FAIR VALUE / PROFITABILITY / etc.) and
# explain what the whole section is about. Complementary to per-column tips.

GROUP_TIPS: Dict[str, Tuple[str, str]] = {
    "IDENTITY": ("Who the stock is",
                 "Symbol, company name, sector, exchange, cap category.\n"
                 "Identifies the stock and its segment — not scored."),
    "SCORES": ("Core conviction scores — read this first",
               "Verdict • Composite /100 • Early Entry /100 • Spike /6 • Storm /10.\n"
               "Summarises the whole analysis in one strip.\n"
               "v10.9: Composite /100 includes forensic quality adjustment\n"
               "(Altman Z, Earn Quality, ND/EBITDA, Int Coverage) capped at\n"
               "+8 bonus / −10 penalty. See 'Score /100' cell tooltip for\n"
               "specific thresholds."),
    "PRICE & MARKET": ("Current price, volume, liquidity, volatility",
                       "CMP, Day Chg, 52W range, volume spike, delivery %, beta.\n"
                       "What the market is doing with this stock TODAY."),
    "PRICE": ("Current market price",
              "CMP — most recent close. Compare against FAIR VALUE columns."),
    "WEEKLY CHANGE %": ("Multi-window returns — momentum direction",
                        "2/4/6/8-week returns. 2W>4W = accelerating (bullish).\n"
                        "All positive + rising = sustained uptrend."),
    "FAIR VALUE": ("What the stock SHOULD be worth",
                   "Composite Fair Value (CFV) from 7 models + Margin of Safety (MoS %).\n"
                   "Individual models: M1 DCF (30%), M2 Graham (15%), M3 PE (20%),\n"
                   "M4 PB (15%), M5 EV/EBITDA (10%), M6 DDM (5%), M7 PEG (5%).\n"
                   "CMP < CFV → undervalued (buying opportunity)."),
    "VALUATION": ("Traditional valuation ratios",
                  "P/E, Earnings Yield, P/CF, PEG, P/B, P/S, EV/EBITDA.\n"
                  "Lower = cheaper. Compare against sector medians.\n"
                  "v10.16 (Option B): valuation ratios show '—' when raw\n"
                  "value ≥ 500 (PEG ≥ 50) — these values come from near-zero\n"
                  "denominators (EPS, book value, EBITDA) where the ratio\n"
                  "is arithmetic noise, not a real 'expensive' signal.\n"
                  "Pre-v10.16 showed a capped 1000/100 which users could\n"
                  "misread as 'valued at 1000× earnings' (AMAGI raw PE was\n"
                  "1,981 — meaningless). Scoring: clamped values (pe_num ≥\n"
                  "500) are now NEUTRAL in fundamental_score derivation —\n"
                  "no penalty for 'unknown'. Real expensive stocks (PE 60-\n"
                  "499) still get the −8 penalty. DB still persists clamped\n"
                  "numeric (500 max) for hygiene; display layer converts to\n"
                  "'—'. Every plausible premium-growth valuation fits inside\n"
                  "the thresholds."),
    "PROFITABILITY": ("How well the business makes money",
                      "ROE, ROCE, ROA, Gross/EBITDA/Net margins, quarterly NPM trend.\n"
                      "Higher + improving = higher fundamental score.\n"
                      "v10.15: NPM quarterly columns capped at ±500% (tiny-revenue\n"
                      "distortion — EMAMIREAL hit −845% on a tiny-rev quarter).\n"
                      "v12.6: NPM quarterly columns relabelled to NPM Q (latest)\n"
                      "/ Q-1 / Q-2 (chronological clarity; was Q1/Q2/Q3).\n"
                      "ROE/ROA stored as floats (were strings pre-v10.15 —\n"
                      "broke Excel sort/filter on those columns)."),
    "GROWTH": ("Revenue & earnings trajectory",
               "1-year & 3-year CAGRs, YoY growth, last quarter absolute numbers.\n"
               "PAT growing faster than revenue = operating leverage.\n"
               "v10.14 important caveat: Rev YoY %/PAT YoY % use TRAILING\n"
               "TTM growth (rolling 4-quarter), while Rev/PAT CAGR 1Y %\n"
               "use DISCRETE fiscal-year growth. They can disagree — TTM\n"
               "captures the very-recent trajectory; CAGR uses clean FY\n"
               "boundaries. All growth fields capped at ±500% to filter\n"
               "tiny-base yfinance noise on micro-caps."),
    "FIN HEALTH": ("Balance sheet safety",
                   "D/E, ND/EBITDA, int coverage, liquidity ratios, cash, debt,\n"
                   "FCF, CCC days, dividend yield. Strong here = survives downturns.\n"
                   "v10.15: CCC Days capped at ±500. When revenue < ₹0.1 Cr,\n"
                   "CCC computation is skipped (EMAMIREAL was showing 16,821\n"
                   "days = 46 years — arithmetic noise on tiny denominators)."),
    "CAP ALLOC": ("How management deploys cash",
                  "Dividend yield, payout ratio, capex/revenue ratio.\n"
                  "Low capex + high FCF = asset-light compounder."),
    "SHAREHOLDING": ("Who owns the stock — and how that's changing",
                     "Promoter %, pledge %, FII %, DII %, public float + QoQ deltas.\n"
                     "Rising institutional ownership = conviction signal.\n"
                     "v10.15 honest-display model: Pledge % and DII % show '—'\n"
                     "when 0 because free-tier can't populate them reliably\n"
                     "(pledge needs BSE filings, DII needs NSE API that's\n"
                     "blocked on cloud IPs). Pro QoQ Δ shows '—' when no real\n"
                     "delta computable — was showing 0 for 83/86 stocks\n"
                     "indistinguishably pre-v10.15."),
    "QUALITY SCORES": ("Forensic & accounting-quality checks",
                       "Piotroski F-Score, Altman Z, Beneish M, earnings quality.\n"
                       "Red flags here trigger the anti-trigger guard."),
    "PIPELINE / OB": ("Forward revenue visibility",
                      "Order book ratio, pipeline visibility, L1 contract wins/value,\n"
                      "new market entry. Relevant for infra / defence / engineering."),
    "EARLY DETECTION": ("Signals firing NOW — catch stocks before consensus",
                        "Early Signals list, Smart Money flow, Sector Stage.\n"
                        "Stage 2 + signals firing = classic accumulation zone."),
    "TECHNICAL": ("Chart & momentum indicators",
                  "SMA 200, Supertrend, ADX, RSI, MACD, Stoch %K, MFI, OBV, VWAP,\n"
                  "chart pattern. Confirms entry timing for a qualifying BUY."),
    "BALANCE SHEET": ("Balance sheet health flag",
                      "HEALTHY / WATCH / ALERT status + explanatory note.\n"
                      "ALERT suppresses spike signals and penalises safety score."),
    "TRADE PLAN": ("Actionable entry / stop / targets",
                   "Entry range, Stop Loss, Target 1/2/3, Time Horizon, Risk Level.\n"
                   "Only act when CMP is inside the entry range — don't chase."),
    "NEWS & RISK": ("Narrative context for the trade",
                    "Key catalyst, news sentiment, primary risk, SEBI flags.\n"
                    "Read before entering — catches things the numbers miss."),
    "NEWS": ("Sentiment + risk summary",
             "Key catalyst, news sentiment, primary risk."),
    "ANALYSIS SUMMARY": ("Full AI-written investor memo",
                         "150–250-word narrative from Gemini AI covering quality,\n"
                         "ratios, risks, catalysts, and verdict rationale.\n"
                         "Generated fresh each trading day."),
    "KEY METRICS": ("Essential ratios for gold-tier candidates",
                    "P/E, PEG, ROE, D/E, PAT YoY, Piotroski F.\n"
                    "Quick-glance fundamentals for early-mover candidates."),
}


# ════════════════════════════════════════════════════════════════════════════
# CONTEXT-APPROPRIATE ICONS PER METRIC FAMILY
# ════════════════════════════════════════════════════════════════════════════
_ICON_FAMILIES = {
    "🎯": {"Verdict", "Score /100", "Early Entry /100", "Spike Score /6", "Spike /6",
           "Storm Score /10", "Storm /10", "Action Required",
           "Gold-Tier Filter"},
    "💰": {"CFV (₹)", "FV Low (₹)", "FV High (₹)", "MoS %", "MoS Label",
           "P/E TTM", "P/E", "Earn Yield %",
           "P/CF", "PEG Ratio", "PEG", "P/B", "P/S", "EV/EBITDA",
           "Div Yield %", "Payout Ratio %",
           "M1: DCF FV (₹)", "M2: Graham FV (₹)", "M3: PE FV (₹)",
           "M4: PB FV (₹)", "M5: EV FV (₹)", "M6: DDM FV (₹)", "M7: PEG FV (₹)",
           "CFV Safety Cap"},
    "📈": {"CMP (₹)", "Day Chg %", "52W High (₹)", "52W Low (₹)",
           "Chg% [2-Weekly]", "Chg% [4-Weekly]", "Chg% [6-Weekly]", "Chg% [8-Weekly]",
           "Chg% [2-Wk]", "Chg% [4-Wk]", "Chg% [6-Wk]", "Chg% [8-Wk]"},
    "📊": {"SMA 200", "Supertrend", "ADX", "RSI (14)", "MACD Signal",
           "Stoch %K", "MFI", "OBV Signal", "Above VWAP", "Chart Pattern", "Pattern",
           "Vol Spike (×50D)", "Delivery %", "Beta"},
    "🏛": {"Promoter %", "Pro QoQ Δ", "FII %", "FII QoQ Δ", "DII %", "DII QoQ Δ",
           "Public Float %", "Smart Money"},
    "🛡": {"D/E Ratio", "D/E", "ND/EBITDA", "Int Coverage", "Current Ratio",
           "Quick Ratio", "Cash (₹Cr)", "Total Debt (₹Cr)", "FCF (₹Cr)",
           "FCF Yield %", "CCC Days", "Pledge %", "Pledge Direction",
           "BS Health Flag", "BS Health Note", "Altman Z", "Piotroski F /9"},
    "⚠": {"Beneish M", "Primary Risk", "SEBI Flags", "Risk Level", "Earn Quality"},
    "🚀": {"Rev CAGR 1Y %", "Rev CAGR 3Y %", "PAT CAGR 1Y %", "PAT CAGR 3Y %",
           "EBITDA CAGR 1Y %", "Rev YoY %", "PAT YoY %", "Margin Expansion",
           "Q3 Rev (₹Cr)", "Q3 PAT (₹Cr)", "Q3 EBITDA (₹Cr)",
           "Gross Mgn %", "EBITDA Mgn %", "NPM %", "NPM Q (latest) %", "NPM Q-1 %", "NPM Q-2 %",
           "ROE %", "ROCE %", "ROA %", "Early Signals", "Sector Stage",
           "OB/Bill Ratio", "Pipeline Vis", "L1 Wins 90D", "L1 Est (₹Cr)",
           "New Mkt Entry", "Capex / Rev %", "Key Catalyst"},
    "🎚": {"Entry Range (₹)", "Stop Loss (₹)", "Target 1 (₹)", "Target 2 (₹)",
           "Target 3 (₹)", "Time Horizon", "Horizon", "R:R Ratio",
           "Support 1 (₹)", "Support 2 (₹)", "Resist 1 (₹)", "Resist 2 (₹)"},
    "🏷": {"Symbol", "Company Name", "Company", "Sector", "Exchange",
           "Cap Category", "Date", "Time (IST)", "Alert Type", "Trigger Detail",
           "Prev Score", "New Score", "Score Δ", "BSE Code",
           "IDENTITY"},
    "📰": {"News Sentiment", "View Analysis Summary",
           "NEWS", "NEWS & RISK", "ANALYSIS SUMMARY"},
}

# Group-name → icon family mapping (supplements _ICON_MAP above for section tips)
_GROUP_ICON_EXTRA = {
    "SCORES":          "🎯",
    "PRICE & MARKET":  "📈",
    "PRICE":           "📈",
    "WEEKLY CHANGE %": "📈",
    "FAIR VALUE":      "💰",
    "VALUATION":       "💰",
    "PROFITABILITY":   "🚀",
    "GROWTH":          "🚀",
    "KEY METRICS":     "🚀",
    "FIN HEALTH":      "🛡",
    "BALANCE SHEET":   "🛡",
    "CAP ALLOC":       "🛡",
    "SHAREHOLDING":    "🏛",
    "QUALITY SCORES":  "⚠",
    "PIPELINE / OB":   "🚀",
    "EARLY DETECTION": "🚀",
    "TECHNICAL":       "📊",
    "TRADE PLAN":      "🎚",
}

_ICON_MAP: Dict[str, str] = {}
for icon, headers in _ICON_FAMILIES.items():
    for h in headers:
        _ICON_MAP[h] = icon


# ════════════════════════════════════════════════════════════════════════════
# COLOR ACCENTS PER FAMILY
# ════════════════════════════════════════════════════════════════════════════
_FAMILY_COLORS = {
    "🎯": ("B45309", "FEF3C7"),
    "💰": ("065F46", "D1FAE5"),
    "📈": ("1D4ED8", "DBEAFE"),
    "📊": ("6366F1", "EEF2FF"),
    "🏛": ("7C3AED", "EDE9FE"),
    "🛡": ("0F766E", "CCFBF1"),
    "⚠": ("B91C1C", "FEE2E2"),
    "🚀": ("DB2777", "FCE7F3"),
    "🎚": ("0891B2", "CFFAFE"),
    "🏷": ("475569", "F1F5F9"),
    "📰": ("92400E", "FED7AA"),
    "💡": ("64748B", "F8FAFC"),
}


def _icon(header: str) -> str:
    # Session 17: check _ICON_MAP (per-column) first, then group-name extras,
    # then fall back to the neutral 💡 glyph. This way section-header tips
    # render with a family-appropriate icon (💰 for FAIR VALUE, 🛡 for FIN
    # HEALTH, 📊 for TECHNICAL, etc.) instead of the generic bulb.
    if header in _ICON_MAP:
        return _ICON_MAP[header]
    if header in _GROUP_ICON_EXTRA:
        return _GROUP_ICON_EXTRA[header]
    return "💡"


# ════════════════════════════════════════════════════════════════════════════
# TIER 1 — POLISHED HOVER TEXT
# ════════════════════════════════════════════════════════════════════════════
_DIVIDER = "━" * 28


def _bulletise_line(line: str) -> list:
    line = line.strip()
    if not line:
        return []
    if "|" in line:
        parts = [p.strip() for p in line.split("|") if p.strip()]
        return [f"  › {p}" for p in parts]
    return [f"  {line}"]


def format_tooltip(header: str, short: str, full: str) -> str:
    """Render the polished Tier-1 hover text for a single header."""
    icon = _icon(header)
    short_lines = _bulletise_line(short) or [f"  {short.strip()}"]

    full_lines = []
    for raw in full.split("\n"):
        full_lines.extend(_bulletise_line(raw))

    body_parts = [f"{icon}  {header.upper()}", _DIVIDER, "QUICK READ", *short_lines]
    if full_lines:
        body_parts += ["", "DETAIL", *full_lines]
    body_parts += [_DIVIDER, "NSE/BSE Analyser · hover cue"]
    return "\n".join(body_parts)


# ════════════════════════════════════════════════════════════════════════════
# TIER 1+2 — APPLY TO WORKSHEET HEADERS
# ════════════════════════════════════════════════════════════════════════════
_CUE = " ⓘ"


def _comment(text: str, width: int = 420, height: int = None) -> Comment:
    c = Comment(text, "NSE/BSE Analyser")
    line_count = text.count("\n") + 1
    c.width = width
    # v10.12: height is now per-tooltip dynamic. Pre-v10.12 this was hardcoded
    # to max(260, min(18*lines+40, 380)) which forced a 260px floor onto every
    # tooltip — short ones (2-3 lines like Stop Loss) ended up with massive
    # empty vertical space. Now:
    #   - if caller passed an explicit height, honour it
    #   - otherwise compute: max(85, min(17*lines+36, 380))
    # The actual on-screen box is set by excel_generator._patch_tooltip_vml()
    # which uses the same formula per-shape from comments{N}.xml text.
    if height is not None:
        c.height = height
    else:
        c.height = max(85, min(17 * line_count + 36, 380))
    return c


def apply_tooltips(
    ws,
    header_row: int,
    col_headers: Iterable[str],
    *,
    add_cue: bool = True,
    ref_anchors: Optional[Dict[str, int]] = None,
    ref_sheet_name: str = "📖 Tooltip Reference",
) -> None:
    """Attach polished tooltips + ⓘ cue + optional ref-link to every header
    cell that has an entry in TIPS. Headers without entries are skipped (no
    error — keeps backward compatibility with legacy _apply_col_tips behavior).
    """
    for ci, h in enumerate(col_headers, 1):
        if h not in TIPS:
            continue
        short, full = TIPS[h]
        cell = ws.cell(header_row, ci)

        # Tier 1: polished hover
        cell.comment = _comment(format_tooltip(h, short, full))

        # Tier 2: visible ⓘ cue on header text
        if add_cue and isinstance(cell.value, str) and not cell.value.endswith(_CUE):
            cell.value = f"{cell.value}{_CUE}"

        # Tier 3: hyperlink to reference row
        if ref_anchors and h in ref_anchors:
            cell.hyperlink = f"#'{ref_sheet_name}'!A{ref_anchors[h]}"


# ════════════════════════════════════════════════════════════════════════════
# TIER 3 — REFERENCE SHEET BUILDER
# ════════════════════════════════════════════════════════════════════════════

def _fill(hex_): return PatternFill("solid", fgColor=hex_)
def _font(bold=False, color="1F2937", size=10, italic=False):
    return Font(name="Segoe UI", bold=bold, color=color, size=size, italic=italic)
def _thin(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


_FAMILY_ORDER = ["🎯", "💰", "📈", "📊", "🚀", "🛡", "🏛", "⚠", "🎚", "🏷", "📰", "💡"]
_FAMILY_LABEL = {
    "🎯": "Verdicts & Composite Scores",
    "💰": "Valuation & Fair Value",
    "📈": "Price & Momentum",
    "📊": "Technical Indicators",
    "🚀": "Growth & Margins",
    "🛡": "Balance Sheet & Safety",
    "🏛": "Ownership & Institutional",
    "⚠": "Risk & Forensics",
    "🎚": "Trade Plan",
    "🏷": "Identity & Metadata",
    "📰": "Narrative & Sentiment",
    "💡": "Other",
}


def build_reference_sheet(wb, sheet_name: str = "📖 Tooltip Reference") -> Dict[str, int]:
    """Create the rich Tooltip Reference sheet. Returns {header -> row}
    mapping so callers can pass it as ref_anchors to apply_tooltips()."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "475569"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([3, 34, 78, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title banner
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, "TOOLTIP REFERENCE  ·  Full explanations for every analyser metric")
    t.fill = _fill("1E3A8A"); t.font = _font(True, "FFFFFF", 13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Sub-instructions
    ws.merge_cells("A2:D2")
    s = ws.cell(2, 1,
                "Every column header in the dashboard is explained here. "
                "Click the ⓘ on any header to jump straight to its card.")
    s.fill = _fill("F1F5F9"); s.font = _font(False, "334155", 9, italic=True)
    s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    by_family: Dict[str, list] = {k: [] for k in _FAMILY_ORDER}
    for h in TIPS.keys():
        fam = _icon(h)
        by_family.setdefault(fam, []).append(h)

    anchors: Dict[str, int] = {}
    row = 4

    for fam in _FAMILY_ORDER:
        headers = sorted(by_family.get(fam, []))
        if not headers:
            continue
        accent, pale = _FAMILY_COLORS[fam]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row, 1, f"  {fam}   {_FAMILY_LABEL[fam]}  ({len(headers)} metric{'s' if len(headers) != 1 else ''})")
        c.fill = _fill(accent); c.font = _font(True, "FFFFFF", 12)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for h in headers:
            short, full = TIPS[h]
            anchors[h] = row

            lbl = ws.cell(row, 2, h)
            lbl.fill = _fill(pale); lbl.font = _font(True, accent, 11)
            lbl.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            lbl.border = _thin(accent)

            body = f"{short}\n\n{full}" if full else short
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=3)
            cnt = ws.cell(row, 3, body)
            cnt.fill = _fill("FFFFFF"); cnt.font = _font(False, "1F2937", 10)
            cnt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            cnt.border = _thin()

            nav = ws.cell(row, 4, "↑ top")
            nav.fill = _fill("FFFFFF"); nav.font = _font(False, accent, 9, italic=True)
            nav.alignment = Alignment(horizontal="center", vertical="top")
            nav.border = _thin()
            nav.hyperlink = f"#'{sheet_name}'!A1"

            line_count = body.count("\n") + max(1, len(body) // 80)
            ws.row_dimensions[row].height = max(38, min(18 * line_count + 10, 180))
            row += 1

        row += 1

    ws.freeze_panes = "A4"
    return anchors


# ════════════════════════════════════════════════════════════════════════════
# GROUP-HEADER TOOLTIP APPLICATION
# ════════════════════════════════════════════════════════════════════════════

def apply_group_tooltips(ws, header_row: int,
                          groups) -> None:
    """Attach a hover tooltip to each group (section) header.

    `groups` is an iterable of (start_col, name, color, span) tuples —
    same shape as FULL_GROUPS / GOLD_GROUPS in excel_generator. The tooltip
    is placed on the FIRST cell of each merged group (where the label lives).

    Session 17: added so section headers (IDENTITY, SCORES, FAIR VALUE, etc.)
    hover with an explanation of what the whole section covers. Complementary
    to per-column tooltips — the group tip is an orientation aid when scanning
    left-to-right across the wide dashboard.

    Session 28: also append the visible ⓘ cue to the section label so users
    know the section header is hoverable (matches per-column behaviour).
    """
    for sc, nm, _color, _span in groups:
        if nm not in GROUP_TIPS:
            continue
        short, full = GROUP_TIPS[nm]
        cell = ws.cell(header_row, sc)
        cell.comment = _comment(format_tooltip(nm, short, full),
                                 width=340, height=200)
        # Append the ⓘ cue (matches apply_tooltips behaviour for columns)
        if isinstance(cell.value, str) and not cell.value.endswith(_CUE):
            cell.value = f"{cell.value}{_CUE}"