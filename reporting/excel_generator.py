"""
excel_generator.py  —  NSE/BSE Stock Analyser v6 Dashboard Generator
Matches the reference template exactly: 6 sheets, correct colours,
group headers, column order, row heights, and column widths.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime
# Session 16: tooltip system — polished hover + ⓘ cue + Tooltip Reference sheet
from reporting.tooltip_formatter import (
    TIPS as _TT_TIPS,
    apply_tooltips as _tt_apply,
    apply_group_tooltips as _tt_apply_groups,
    build_reference_sheet as _tt_build_ref,
)

NAVY  = "1E293B";  WHITE = "FFFFFF";  LG = "F8FAFC"

VERDICT_STYLES = {
    "DEEP VALUE EARLY MOVER": {"bg":"FAC775","text":"412402"},
    "EARLY MOVER":            {"bg":"FAC775","text":"412402"},
    "DEEP VALUE":             {"bg":"D1FAE5","text":"065F46"},
    "BUY":                    {"bg":"DBEAFE","text":"1E3A5F"},
    "BUY / EARLY MOVER":      {"bg":"DBEAFE","text":"1E3A5F"},
    # Session 24: OVERVALUED distinct from BUY (blue) and WATCHLIST (yellow)
    # Soft orange reads as "good stock, caution on price" — not green, not red
    "OVERVALUED":             {"bg":"FED7AA","text":"7C2D12"},
    "WATCHLIST":              {"bg":"FEF3C7","text":"78350F"},
    "NEUTRAL":                {"bg":"FEF3C7","text":"78350F"},
    "AVOID":                  {"bg":"FEE2E2","text":"7F1D1D"},
    "AVOID / EXIT":           {"bg":"FEE2E2","text":"7F1D1D"},
    "EXIT":                   {"bg":"FEE2E2","text":"7F1D1D"},
}
_DS = {"bg":"F8FAFC","text":"1E293B"}

FULL_GROUPS = [
    # v13.x: SCORES band widened 4→5 to accommodate new "Quick Pick" column
    # inserted right after Verdict. All subsequent starts shift by +1.
    (1,  "IDENTITY",        "1E293B",7),(8, "SCORES",         "7C3AED",5),
    (13, "PRICE & MARKET",  "0369A1",7),(20,"WEEKLY CHANGE %", "0F766E",4),
    # v10.8: FAIR VALUE span reduced 13→12 after removing duplicate 'Upside to FV %' column.
    # All subsequent group starts shift left by 1 (36→35, 43→42, 53→52, ...).
    (24, "FAIR VALUE",      "B45309",12),(36,"VALUATION",      "0891B2",7),
    (43, "PROFITABILITY",   "059669",10),(53,"GROWTH",         "047857",10),
    (63, "FIN HEALTH",      "DC2626",10),(73,"CAP ALLOC",      "6D28D9",3),
    (76, "SHAREHOLDING",    "7C3AED",9),(85,"QUALITY SCORES",  "0D9488",4),
    (89, "PIPELINE / OB",   "1D4ED8",5),(94,"EARLY DETECTION", "B45309",3),
    (97, "TECHNICAL",       "6D28D9",14),(111,"BALANCE SHEET", "D97706",2),
    (113,"TRADE PLAN",      "059669",7),(120,"NEWS & RISK",    "475569",4),
    (124,"ANALYSIS SUMMARY","0F172A",1),
]

FULL_COLS = [
    ("Symbol",12,"symbol"),("Company Name",28,"company_name"),("Sector",22,"sector"),
    ("Exchange",13,"exchange_tag"),("BSE Code",10,"bse_code"),("Cap Category",13,"cap_category"),
    # Session 24: Verdict col now uses verdict_display (e.g., "BUY ●●●") to
    # show confidence dots inline. Plain 'verdict' key is still used by Gold
    # filter, priority_ranker, and styling lookup.
    ("Verdict",26,"verdict_display"),
    # v13.x: Quick Pick — surfaces the archetype assigned by
    # ScoringEngine._assign_quick_pick(). Distinct from Verdict (which is
    # cap-tier+MoS-gated). Possible values:
    #   DEEP VALUE EARLY MOVER · DEEP VALUE · EARLY MOVER · AVOID / EXIT · WATCHLIST
    # The 'label' field is already populated on every stock dict in
    # scoring_engine.calculate_composite_score() (line 324). Color styling
    # falls back to the existing VERDICT_STYLES map keys when the label
    # matches one of those entries.
    ("Quick Pick",26,"label"),
    ("Score /100",10,"composite_score"),("Early Entry /100",14,"early_entry_score"),
    ("Spike Score /6",11,"spike_count"),("Storm Score /10",12,"storm_score"),
    ("CMP (₹)",11,"close"),("Day Chg %",10,"day_change"),("52W High (₹)",12,"high_52w"),
    ("52W Low (₹)",12,"low_52w"),("Vol Spike (×50D)",14,"vol_ratio"),("Delivery %",11,"delivery_pct"),
    ("Beta",8,"beta"),("Chg% [2-Weekly]",14,"2w_chg"),("Chg% [4-Weekly]",14,"4w_chg"),
    ("Chg% [6-Weekly]",14,"6w_chg"),("Chg% [8-Weekly]",14,"8w_chg"),
    ("CFV (₹)",11,"cfv"),("FV Low (₹)",11,"cfv_low"),("FV High (₹)",11,"cfv_high"),
    ("MoS %",10,"mos_pct"),("MoS Label",22,"mos_label"),
    ("M1: DCF FV (₹)",14,"M1_DCF"),("M2: Graham FV (₹)",16,"M2_Graham"),
    ("M3: PE FV (₹)",14,"M3_PE"),("M4: PB FV (₹)",14,"M4_PB"),
    ("M5: EV FV (₹)",14,"M5_EV"),("M6: DDM FV (₹)",14,"M6_DDM"),("M7: PEG FV (₹)",14,"M7_PEG"),
    ("P/E TTM",9,"pe"),("Earn Yield %",11,"earnings_yield"),("P/CF",9,"p_cf"),
    ("PEG Ratio",10,"peg"),("P/B",9,"pb"),("P/S",9,"ps"),("EV/EBITDA",11,"ev_ebitda"),
    ("ROE %",9,"roe"),("ROCE %",9,"roce"),("ROA %",9,"roa"),
    ("Gross Mgn %",11,"gross_margin"),("EBITDA Mgn %",12,"ebitda_margin"),("NPM %",9,"npm"),
    # v12.6 (#11): renamed from NPM Q1/Q2/Q3 → NPM Q (latest)/Q-1/Q-2 to
    # eliminate the inverse-chronological confusion. Pre-v12.6 labels read
    # left-to-right as "Q1 Q2 Q3" which suggested chronological order, but
    # the data was actually most-recent-first (Q1 = latest quarter). New
    # labels make the time order unambiguous: "(latest) → -1 → -2" reads
    # as "this quarter / one ago / two ago". DB keys (npm_q1/q2/q3) and
    # all scoring logic unchanged.
    ("NPM Q (latest) %",13,"npm_q1"),("NPM Q-1 %",10,"npm_q2"),("NPM Q-2 %",10,"npm_q3"),
    ("Margin Expansion",17,"margin_expansion"),
    ("Rev CAGR 1Y %",12,"rev_cagr_1y"),("Rev CAGR 3Y %",12,"rev_cagr_3y"),
    ("PAT CAGR 1Y %",12,"pat_cagr_1y"),("PAT CAGR 3Y %",12,"pat_cagr_3y"),
    ("EBITDA CAGR 1Y %",15,"ebitda_cagr_1y"),("Rev YoY %",10,"rev_yoy"),
    ("PAT YoY %",10,"pat_yoy"),("Q3 Rev (₹Cr)",13,"q3_rev"),
    ("Q3 PAT (₹Cr)",13,"q3_pat"),("Q3 EBITDA (₹Cr)",15,"q3_ebitda"),
    ("D/E Ratio",10,"debt_equity"),("ND/EBITDA",11,"nd_ebitda"),
    ("Int Coverage",12,"int_coverage"),("Current Ratio",13,"current_ratio"),
    ("Quick Ratio",11,"quick_ratio"),("Cash (₹Cr)",11,"cash"),
    ("Total Debt (₹Cr)",15,"total_debt"),("FCF (₹Cr)",11,"fcf"),
    ("FCF Yield %",12,"fcf_yield"),("CCC Days",10,"ccc_days"),
    ("Div Yield %",11,"div_yield"),("Payout Ratio %",13,"payout_ratio"),
    ("Capex / Rev %",12,"capex_rev"),
    ("Promoter %",11,"promoter_pct"),("Pro QoQ Δ",10,"promoter_qoq"),
    ("Pledge %",10,"pledge_pct"),("Pledge Direction",15,"pledge_direction"),
    ("FII %",9,"fii_pct"),("FII QoQ Δ",10,"fii_qoq"),
    ("DII %",9,"dii_pct"),("DII QoQ Δ",10,"dii_qoq"),("Public Float %",13,"public_float"),
    ("Piotroski F /9",13,"piotroski_f"),("Altman Z",10,"altman_z"),
    ("Beneish M",10,"beneish_m"),("Earn Quality",12,"earnings_quality"),
    ("OB/Bill Ratio",12,"ob_bill_ratio"),("Pipeline Vis",12,"pipeline_vis"),
    ("L1 Wins 90D",12,"l1_wins"),("L1 Est (₹Cr)",13,"l1_value"),
    ("New Mkt Entry",26,"new_market_entry"),
    ("Early Signals",42,"early_signals"),("Sector Stage",11,"rotation_stage"),
    ("Smart Money",28,"smart_money_signals"),
    ("SMA 200",10,"sma_200"),("Supertrend",12,"supertrend"),("ADX",8,"adx"),
    ("RSI (14)",9,"rsi"),("MACD Signal",18,"macd_signal"),("Stoch %K",9,"stoch_k"),
    ("MFI",8,"mfi"),("OBV Signal",14,"obv_signal"),("Above VWAP",11,"above_vwap"),
    ("Chart Pattern",22,"chart_pattern"),
    ("Support 1 (₹)",12,"support_1"),("Support 2 (₹)",12,"support_2"),
    ("Resist 1 (₹)",12,"resist_1"),("Resist 2 (₹)",12,"resist_2"),
    ("BS Health Flag",13,"bs_status"),("BS Health Note",40,"bs_flags"),
    ("Entry Range (₹)",15,"entry_range"),("Stop Loss (₹)",13,"stop_loss"),
    ("Target 1 (₹)",12,"t1"),("Target 2 (₹)",12,"t2"),("Target 3 (₹)",12,"t3"),
    ("Time Horizon",22,"horizon"),("Risk Level",11,"risk_level"),
    ("Key Catalyst",42,"key_catalyst"),("News Sentiment",15,"news_sentiment"),
    ("Primary Risk",42,"primary_risk"),("SEBI Flags",22,"sebi_flags"),
    ("View Analysis Summary",70,"Analysis_Summary_Block_H"),
]

GOLD_GROUPS = [
    # Session 25 fix: After Session 23 removed the "Upside %" column from
    # GOLD_COLS, this band table wasn't updated. KEY METRICS was still 7 wide
    # (including the now-removed Upside col), so every band from EARLY DETECTION
    # onward was shifted right by 1 — headers like "Early Signals" appeared
    # under KEY METRICS, "Supertrend" under EARLY DETECTION, etc. Fixed by
    # reducing KEY METRICS span 7→6 and shifting all subsequent starts by -1.
    # Sum of spans now = 6+4+1+4+3+6+3+3+7+2+1 = 40 = len(GOLD_COLS). ✓
    # v13.x: SCORES band widened 4→5 for new "Quick Pick" column inserted
    # right after Verdict. All subsequent starts shift +1. New sum = 41.
    (1,"IDENTITY","1E293B",6),(7,"SCORES","7C3AED",5),(12,"PRICE","0369A1",1),
    (13,"WEEKLY CHANGE %","0F766E",4),(17,"FAIR VALUE","B45309",3),
    (20,"KEY METRICS","0891B2",6),(26,"EARLY DETECTION","B45309",3),
    (29,"TECHNICAL","6D28D9",3),(32,"TRADE PLAN","059669",7),
    (39,"NEWS","475569",2),(41,"ANALYSIS SUMMARY","0F172A",1),
]

GOLD_COLS = [
    ("Symbol",12,"symbol"),("Company Name",28,"company_name"),("Sector",22,"sector"),
    ("Exchange",13,"exchange_tag"),("Cap Category",13,"cap_category"),("Verdict",26,"verdict_display"),
    # v13.x: Quick Pick — same as in FULL_COLS. See note there for full context.
    ("Quick Pick",26,"label"),
    ("Score /100",10,"composite_score"),("Early Entry /100",14,"early_entry_score"),
    ("Spike /6",9,"spike_count"),("Storm /10",9,"storm_score"),
    ("CMP (₹)",11,"close"),("Chg% [2-Wk]",13,"2w_chg"),("Chg% [4-Wk]",13,"4w_chg"),
    ("Chg% [6-Wk]",13,"6w_chg"),("Chg% [8-Wk]",13,"8w_chg"),
    ("CFV (₹)",11,"cfv"),("MoS %",10,"mos_pct"),
    # Session 23: "Upside %" removed — it was always identical to MoS %
    # ((CFV-CMP)/CMP×100 computed once, displayed under two labels).
    # Kept in stock dict as 'upside' key for backward compat with scorer and Trade Summary.
    ("MoS Label",20,"mos_label"),("P/E",9,"pe"),("PEG",9,"peg"),
    ("ROE %",9,"roe"),("D/E",9,"debt_equity"),("PAT YoY %",10,"pat_yoy"),
    # v12.5: renamed from "F-Score /9" to "Piotroski F /9" to match the
    # Full Dashboard. Same data, same key — just a label sync so the
    # Glossary entry "Piotroski F /9" applies to both sheets.
    ("Piotroski F /9",13,"piotroski_f"),
    ("Early Signals",42,"early_signals"),("Smart Money",28,"smart_money_signals"),
    ("Sector Stage",11,"rotation_stage"),("Supertrend",12,"supertrend"),
    ("RSI (14)",9,"rsi"),("Pattern",24,"chart_pattern"),
    ("Entry Range (₹)",16,"entry_range"),("Stop Loss (₹)",12,"stop_loss"),
    ("Target 1 (₹)",12,"t1"),("Target 2 (₹)",12,"t2"),("Target 3 (₹)",12,"t3"),
    ("Time Horizon",22,"horizon"),("Risk Level",10,"risk_level"),
    ("Key Catalyst",42,"key_catalyst"),("Primary Risk",42,"primary_risk"),
    ("View Analysis Summary",70,"Analysis_Summary_Block_H"),
]

GLOSSARY_DATA = [
    ("IDENTITY","Symbol","NSE/BSE ticker symbol. Format: up to 20 chars, no spaces (e.g. TATAMOTORS, HDFCBANK)","All sheets"),
    ("IDENTITY","BSE Code","6-digit BSE scrip code (e.g. 500325 = Reliance). Use on BSE for orders.","Full Dashboard"),
    ("IDENTITY","Cap Category",
     "LARGE CAP = mcap > ₹20,000 Cr (top 100 stocks, lowest risk) | "
     "MID CAP = ₹5,000–20,000 Cr (101–250 stocks, moderate risk) | "
     "SMALL CAP = ₹500–5,000 Cr (251+ stocks, higher risk) | "
     "MICRO CAP = < ₹500 Cr (highest risk, highest potential)","All sheets"),
    ("IDENTITY","Verdict",
     "BUY = score above cap threshold AND MoS ≥ −10% (CMP at/below fair value, act now) | "
     "OVERVALUED = score qualifies for BUY but MoS gate blocks (great business, currently expensive, wait for pullback) | "
     "WATCHLIST = signal building, below BUY threshold (monitor for score improvement) | "
     "NEUTRAL = no clear signal yet, hold off | "
     "AVOID = weak fundamentals + bad technicals + overvalued (stay out) | "
     "DEEP VALUE = significantly undervalued (MoS>25%) with high score | "
     "EARLY MOVER = early signal detected before consensus (act ahead of crowd). "
     "Session 24: confidence dots indicate distance from the decisive threshold — "
     "●●● = HIGH (≥5 points clear), ●●○ = MEDIUM (2-5 above), ●○○ = LOW (<2 above; cliff zone; treat with extra caution). "
     "v10.17: data-completeness guard — a stock is only allowed to carry a BUY verdict when at least 3 of "
     "5 sub-score dimensions (Fundamental / Technical / Safety / Sentiment / Early Entry) actually had real "
     "data move them away from base. If a stock's score qualifies for BUY but only 0–2 dimensions are "
     "informed, the verdict is demoted to WATCHLIST and labelled '(thin data)'. Prevents inflated BUYs on "
     "stocks with too much missing data.","All sheets"),
    ("IDENTITY","Exchange",
     "NSE_ONLY = only on National Stock Exchange | "
     "BSE_ONLY = only on Bombay Stock Exchange | "
     "DUAL_LISTED = on both NSE and BSE (broader institutional access, preferred) | "
     "BSE_SME = BSE Small & Medium Enterprise platform (use limit orders, low liquidity). "
     "Session 22: when BSE bhavcopy download fails (Cloudflare blocks cloud IPs), "
     "a curated allowlist of Nifty 100 + popular mid-caps is used to tag DUAL_LISTED. "
     "Lesser-known small-caps may show NSE_ONLY even if also listed on BSE.","All sheets"),
    ("SCORES","Quick Pick",
     "Archetype label assigned by ScoringEngine._assign_quick_pick(). DISTINCT from the Verdict column: "
     "Verdict is cap-tier-aware and MoS-gated (asks 'should you act?'); Quick Pick is a simpler tag for "
     "the stock's character (asks 'what kind of signal is this?'). The two CAN disagree — that's by design, not a bug. "
     "Possible values (priority order, first match wins): "
     "DEEP VALUE EARLY MOVER = MoS > 25% AND score > 70 AND early_entry ≥ 60 (cheap + breaking out, top conviction) | "
     "DEEP VALUE = MoS > 25% AND score > 70 (cheap vs fair value, no breakout yet) | "
     "EARLY MOVER = early_entry ≥ 70 AND score > 55 (technical breakout signal, may not be deeply discounted) | "
     "AVOID / EXIT = score < 38 OR (score < 45 AND MoS < −30%) (weak fundamentals + bad value) | "
     "WATCHLIST = fallback for everything else (no special archetype, monitor). "
     "Use this column with AutoFilter to instantly isolate top-conviction picks (e.g., filter to 'DEEP VALUE EARLY MOVER'). "
     "Why DEEP VALUE EARLY MOVER uses 3 factors (vs 2 for the others): it's the COMBO of DEEP VALUE and EARLY MOVER — "
     "inherits the value gates from DEEP VALUE (MoS>25% AND Score>70) and adds a momentum check. Note the EE threshold "
     "drops 70 → 60 in the combo: when value AND quality are both high, you don't need an extreme momentum signal — "
     "moderate confirmation (EE≥60) is enough. Pure EARLY MOVER needs EE≥70 because momentum is the only thing speaking "
     "for stocks that may not have great fundamentals (Score>55 floor, no MoS gate). So both EE thresholds are correct "
     "for what each archetype is claiming. "
     "v13.x: Quick Pick is now computed TWICE — once in calculate_composite_score(), then re-computed in master_funnel "
     "AFTER the Score Convergence +8 EE bonus fires. Pre-fix, when the bonus crossed the 60 or 70 archetype threshold, "
     "the displayed Quick Pick disagreed with the post-bonus EE in the same row (3 stocks affected: MOCAPITAL, KIRLFER, KAMAHOLD). "
     "The recompute runs only when the bonus actually fires, so non-trigger stocks see zero change.",
     "All sheets"),
    ("SCORES","Score /100","Composite: Fundamental 35% + Technical 30% + Early 15% + Sentiment 10% + Safety 10%. Session 24: if no paid/AI sentiment signals fired (no FII/Promoter/DII QoQ, no insider buy, no news sentiment, no pledge direction), the 10% sentiment weight redistributes proportionally to the other 4 sub-scores (Fundamental→0.389, Technical→0.333, Early→0.167, Safety→0.111) — no 'free 5 points' for missing data. Spike bonus (+2 per trigger, max +10) is gated on fundamental quality: only awards full +10 when fundamental_score ≥ 55; capped at +3 for weaker stocks to prevent momentum masking weak fundamentals. v10.17: a high composite score alone does not guarantee BUY — the verdict also requires at least 3 of 5 sub-score dimensions to be 'informed' (real data fired, not just sat at base). Otherwise BUY is demoted to WATCHLIST with a 'thin data' annotation. See the Verdict tooltip for details.","All sheets"),
    ("SCORES","Early Entry /100","12-signal system measuring how early vs consensus. ≥50 = EARLY MOVER badge | ≥35 = AHEAD OF CONSENSUS. Session 23: low EE on a Gold-sheet stock is not a bug — Gold admits two archetypes: MOMENTUM (high EE from firing trend/volume signals) and VALUE (low EE but high Score + MoS + clean safety; quietly accumulating without momentum triggers yet)","All sheets"),
    ("SCORES","Spike Score /6","Count of active triggers from 6 IF-THEN spike conditions (Section 3H). Session 23: low Spike on a Gold stock is OK — VALUE-archetype candidates may be accumulating quietly without hot momentum triggers","All sheets"),
    ("SCORES","Storm Score /10","Volatility resilience score. Higher = more defensive in downturns (VIX>18)","Full Dashboard"),
    ("PRICE & MARKET","CMP","Current Market Price in Indian Rupees (₹)","All sheets"),
    ("PRICE & MARKET","52W High","Highest closing price in past 52 weeks","Full Dashboard"),
    ("PRICE & MARKET","52W Low","Lowest closing price in past 52 weeks","Full Dashboard"),
    ("PRICE & MARKET","Vol Spike (×50D)","Today's volume ÷ 50-day average. >3× = institutional activity","Full Dashboard"),
    ("PRICE & MARKET","Delivery %","Share of traded volume with actual delivery. >60% = conviction","Full Dashboard"),
    ("PRICE & MARKET","Beta","Beta vs Nifty 50. <1 = defensive. >1.5 = volatile","Full Dashboard"),
    ("WEEKLY CHANGE %","Chg% [2-Weekly]","Price change % over past 2 weeks (10 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [4-Weekly]","Price change % over past 4 weeks (20 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [6-Weekly]","Price change % over past 6 weeks (30 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [8-Weekly]","+25% on a quality stock over 8W = institutional accumulation","All sheets"),
    ("FAIR VALUE","CFV","Composite Fair Value — sector-weighted blend of 7 models (M1–M7). v12.2: M3/M4/M5 sector resolution rewritten — production sector strings (Basic Materials, Industrials, Communication Services, Consumer Cyclical/Defensive, Financial Services, Real Estate) now canonicalize via SECTOR_ALIASES before substring-matching against benchmark multipliers, fixing 31/100 stocks that were silently using defaults pre-v12.2.","All sheets"),
    ("FAIR VALUE","MoS %",
     "Margin of Safety = (CFV−CMP)/CMP×100. How much cheaper CMP is vs fair value. "
     "EXCEPTIONAL VALUE = MoS > 40% (deeply undervalued, strong BUY signal) | "
     "STRONG VALUE = MoS 25–40% (significantly undervalued) | "
     "GOOD VALUE = MoS 10–25% (moderately undervalued) | "
     "FAIR VALUE = MoS 0–10% (priced fairly) | "
     "SLIGHT PREMIUM = MoS −15–0% (slightly overvalued, ok for quality stocks) | "
     "OVERVALUED = MoS −30–−15% (CMP exceeds FV, caution) | "
     "SIGNIFICANTLY OVERVALUED = MoS < −30% (avoid, major downside risk). "
     "v13.x: Cell renders '—' (not a number) when CFV is unavailable — typical for ETFs / "
     "index funds / very thin-data instruments where no valuation models could fire. "
     "Pre-fix the cell would show '-100%' as a math artifact of (0-CMP)/CMP×100, "
     "misleading readers to think the stock is 100% overvalued when the truth is "
     "'we can't value it with our model set'. The MoS Label cell mirrors this — "
     "both show '—' together when CFV is missing.","All sheets"),
    # Session 27: Removed duplicate "M1: DCF FV" glossary row here — the more
    # complete version remains in the second FAIR VALUE block below (with
    # Session 19 cap details + SBIN β=0.2 example). Leaving both created
    # two consecutive "M1" rows in the Glossary sheet.
    ("FAIR VALUE","M2: Graham FV","Graham Number = √(22.5×EPS×BVPS). Skip if EPS negative. v12.2: eps/bvps now sanitized via _sf() — handles '—' / 'N/A' / None inputs cleanly.","Full Dashboard"),
    ("FAIR VALUE","M3: PE FV","EPS × Sector 5yr median P/E (mean reversion). v12.2: 28-sector benchmarks via SECTOR_ALIASES (Banks 18, Tech 30, FMCG 45, Steel 10, Realty 25, Telecom 22, Defence 40, etc.). 'Basic Materials' → Metals, 'Industrials' → Infra, 'Communication Services' → Telecom.","Full Dashboard"),
    ("FAIR VALUE","M4: PB FV","BVPS × Sector median Price/Book. Good for asset-heavy sectors (banks, metals). v12.2: 28-sector benchmarks via SECTOR_ALIASES. BVPS fallback derives from close/PB if missing.","Full Dashboard"),
    ("FAIR VALUE","M5: EV FV","CMP × (Sector median EV/EBITDA ÷ Stock EV/EBITDA). Good for capital-intensive businesses. v12.2: 28-sector benchmarks via SECTOR_ALIASES. v12.3 Round 2: proper EV-based formula primary — fair_per_share = CMP × ((annual_ebitda × sector_mult − net_debt) / mcap_cr). Three-tier dispatch (proper / shortcut / skip). Banks/NBFCs/Insurance skip M5 entirely (EV/EBITDA not meaningful for financials). _m5_method diagnostic surfaces which tier fired.","Full Dashboard"),
    ("FAIR VALUE","M6: DDM FV","Dividend Discount Model = DPS×(1+g)/(r−g). Only shown for dividend-paying stocks. v12.2 fix: removed the 2% growth floor — declining-earnings stocks (pat_yoy<0) now correctly get 0% div growth (was 2%, which inflated FV by ~26%).","Full Dashboard"),
    ("FAIR VALUE","M7: PEG FV","EPS × EPS Growth Rate × PEG_BENCHMARK. v12.3 Round 2: PEG_BENCHMARK = 1.0 made an explicit named constant (Lynch's rule of thumb). Mathematically identical to v12.2 outputs but the constant is now tunable — strict value setups could use 0.8, growth-tilted mandates 1.2. v12.2 unit guard still active: skips when growth_3yr < 1.0 (catches decimal-fraction unit error).","Full Dashboard"),
    ("VALUATION","P/E TTM","Price-to-Earnings trailing 12 months","All sheets"),
    ("VALUATION","Earn Yield %","EPS/CMP×100. >6% = beats 10Y GSec risk-free rate","All sheets"),
    ("VALUATION","P/E TTM",
     "Price ÷ EPS (trailing 12 months). How much you pay per ₹1 of earnings. "
     "< 10 = Very cheap (PSUs, cyclicals) | "
     "10–20 = Fair value for slow-growth / value stocks | "
     "20–35 = Growth premium (acceptable if ROE>20% and growth>15%) | "
     "35–60 = Expensive (only justified by very high growth) | "
     "> 60 = Very expensive (speculative, future growth already priced in) | "
     "'—' = Display when raw value ≥ 500 (near-zero EPS produces mathematical "
     "PE in thousands; AMAGI raw PE was 1,981 pre-fix — arithmetic noise, "
     "not a real 'expensive' signal). "
     "v10.16 (Option B): thousand-fold clamped values replaced with '—' for "
     "honest display. DB persists numeric (clamped at 500) for scoring; "
     "fundamental_score treats pe_num ≥ 500 as NEUTRAL (no penalty) because "
     "the value represents 'unknown valuation', not 'expensive'.","All sheets"),
    ("VALUATION","PEG Ratio",
     "P/E ÷ EPS Growth Rate. Adjusts PE for growth. "
     "< 0.5 = Deeply undervalued vs growth (excellent BUY) | "
     "0.5–1.0 = Undervalued vs growth (good value) | "
     "1.0 = Fairly valued (Peter Lynch's neutral benchmark) | "
     "1.0–2.0 = Slight premium (acceptable for quality compounder) | "
     "> 2.0 = Expensive relative to growth (avoid unless market leader) | "
     "'—' = Display when any tier yields ≥ 50 (P/E divided by near-zero "
     "growth — pure arithmetic noise). Even extreme glamour stocks rarely "
     "exceed PEG of 10. "
     "v10.16 (Option B): threshold tightened from 100 → 50 for honest '—' "
     "display. 4-tier fallback: direct pegRatio → PE/PAT-growth → PE/Rev-"
     "growth → PE/sustainable-growth. If all tiers yield ≥ 50 or fail, "
     "display is '—'.","All sheets"),
    ("VALUATION","P/B",
     "Price ÷ Book Value per Share. "
     "< 1.0 = Trading below book (potential deep value or value trap) | "
     "1.0–2.0 = Reasonable (good for banks, metals, asset-heavy) | "
     "2.0–5.0 = Premium (justified if ROE>15%) | "
     "> 5.0 = High premium (only for asset-light compounders with high ROE) | "
     "'—' = Display when raw ≥ 500 (tiny book-value distortions). "
     "v10.16 (Option B): thousand-fold noise values shown as '—' instead "
     "of clamped to specific number.","All sheets"),
    ("VALUATION","EV/EBITDA",
     "Enterprise Value ÷ EBITDA. Better than PE as it ignores capital structure. "
     "< 8 = Cheap (cyclicals, turnarounds) | "
     "8–15 = Fair value range | "
     "15–25 = Premium (growth sectors like IT, pharma) | "
     "> 25 = Expensive (only for market leaders or high-growth) | "
     "'—' = Display when raw ≥ 500 (near-zero EBITDA produces mathematical "
     "EV/EBITDA in thousands; RHETAN raw was 1,352 pre-fix). "
     "v10.16 (Option B): honest display instead of misleading clamped number.",
     "Full Dashboard"),
    ("VALUATION","Earn Yield %",
     "EPS ÷ CMP × 100. Inverse of PE — shows what % return you earn per ₹ invested. "
     "< 4% = Very expensive (below risk-free rate) | "
     "4–6% = Below risk-free (10Y G-Sec ~7%) | "
     "> 6% = Beats G-Sec (attractive for value investors) | "
     "> 8% = High yield (deep value territory)","All sheets"),
    ("PROFITABILITY","ROE %",
     "Return on Equity = Net Income ÷ Shareholders' Equity. "
     ">20% excellent | >15% good | >10% acceptable | <10% weak. "
     "Source: yfinance .info['returnOnEquity'] × 100 when available; "
     "else derived as Earnings Yield × P/B (ROE ≈ EPS/BVPS). "
     "v10.15 FIX #1: values now stored as FLOAT (were quoted strings "
     "pre-v10.15 — '12.47' as text, not 12.47 as number — broke Excel "
     "sort/filter/conditional-formatting on this column for 69/86 stocks).",
     "All sheets"),
    ("PROFITABILITY","ROA %",
     "Return on Assets = Net Income ÷ Total Assets. "
     ">10% efficient | 5–10% acceptable | <5% poor (but normal for "
     "banks/utilities which use high leverage). "
     "Source: yfinance .info['returnOnAssets'] × 100 when available; "
     "else derived as ROE / (1 + D/E). v10.15: stored as FLOAT, not string. "
     "v12.4: clamped to ±100 % (yfinance returned 189 % for M&MFIN, "
     "181 % for TATACAP pre-clamp — finance/NBFC unit-mismatch artefacts).",
     "All sheets"),
    ("PROFITABILITY","ROCE %",
     "Return on Capital Employed. ROCE > cost of capital = value-creating. "
     ">15% good | >20% excellent. Not directly in yfinance; derived from "
     "ROE adjusted for leverage when possible.",
     "Full Dashboard"),
    ("PROFITABILITY","Gross Mgn %",
     "Revenue − COGS, ÷ Revenue × 100. >40% = strong moat; >20% = decent. "
     "Source: yfinance .info['grossMargins'] × 100. "
     "v12.4: clamped to [0, 100] % to filter unit-mismatch outliers.",
     "Full Dashboard"),
    ("PROFITABILITY","EBITDA Mgn %",
     "EBITDA ÷ Revenue × 100. >25% excellent | >15% good | <10% tight. "
     "Source: yfinance .info['ebitdaMargins'] × 100. "
     "v12.4: clamped to ±100 % (POWERGRID showed 83.4 %; legitimate values "
     "stay; absurd >100 % values get capped).",
     "Full Dashboard"),
    ("PROFITABILITY","NPM %",
     "Net Profit Margin TTM = Net Income ÷ Revenue × 100. "
     ">15% excellent | >5% decent | <0% unprofitable. "
     "Source: yfinance .info['profitMargins'] × 100. "
     "v12.4: clamped to ±100 %. Six stocks in the prior run had NPM "
     "126–189 % (DGCONTENT 126 %, AMAGI 189 %, MEGASTAR 165 %, "
     "REDINGTON 157 %, RELIGARE 128 %, GCSL −145 %) — thin-revenue "
     "/ one-time-gain rows where the math is meaningless.",
     "All sheets"),
    ("PROFITABILITY","NPM Q (latest) %",
     "Most recent quarter's Net Profit Margin = (Q(latest) PAT ÷ Q(latest) Revenue) × 100. "
     "Source: yfinance quarterly_income_stmt most-recent column. "
     "v12.6 (#11): renamed from 'NPM Q1 %' for clarity — Q1=most-recent in old "
     "scheme but the L→R reading order suggested chronological (oldest→newest). "
     "v10.15 FIX #2: capped at ±500% to prevent tiny-revenue-denominator "
     "distortions (EMAMIREAL Q(latest) hit −762% pre-clamp). Same design as v10.14 "
     "CAGR clamp. Real businesses don't sustain >500% margin even briefly.",
     "Full Dashboard"),
    ("PROFITABILITY","NPM Q-1 %",
     "Previous quarter's NPM (one quarter before the latest report). Track "
     "Q-2 → Q-1 → Q(latest) trend to catch margin expansion early. "
     "v12.6 (#11): renamed from 'NPM Q2 %'. "
     "v10.15 FIX #2: capped at ±500% (EMAMIREAL Q-1 hit −387% pre-clamp).",
     "Full Dashboard"),
    ("PROFITABILITY","NPM Q-2 %",
     "NPM from two quarters before the latest report. Rising "
     "Q-2 < Q-1 < Q(latest) triggers 'Margin Expansion = YES' — strong "
     "compounder signal. v12.6 (#11): renamed from 'NPM Q3 %'. "
     "v10.15 FIX #2: capped at ±500% (EMAMIREAL Q-2 hit −845% pre-clamp).",
     "Full Dashboard"),
    ("PROFITABILITY","Margin Expansion",
     "YES when NPM has risen for 3 consecutive quarters (Q-2 < Q-1 < Q(latest)). "
     "Strong operating leverage / pricing power signal. "
     "Score: Fundamental +5, Safety +3, Storm +1.",
     "All sheets"),
    ("GROWTH","Rev CAGR 1Y %",
     "Latest FY revenue ÷ prior FY revenue − 1 (discrete fiscal years). "
     ">20% high growth | >10% good. "
     "Differs from 'Rev YoY %' which uses rolling TTM — a large divergence "
     "usually means the company is mid-year with a strong/weak recent quarter. "
     "v10.14: capped at ±500% to filter tiny-base distortions.",
     "Full Dashboard"),
    ("GROWTH","Rev CAGR 3Y %",
     "Latest FY ÷ FY-3 revenue, then ^(1/3) − 1. "
     ">15% strong | >8% decent. Requires ≥4 annual columns in yfinance; "
     "shows '—' for newer IPOs. v10.14: capped at ±500%.",
     "Full Dashboard"),
    ("GROWTH","PAT CAGR 1Y %",
     "Latest FY PAT ÷ prior FY PAT − 1. >20% strong earnings momentum. "
     "Shows '—' if either FY had loss (CAGR mathematically undefined). "
     "v10.14: capped at ±500%.",
     "Full Dashboard"),
    ("GROWTH","PAT CAGR 3Y %",
     "Latest FY PAT ÷ FY-3 PAT, ^(1/3) − 1. >20% compounder | >10% good. "
     "Most reliable long-horizon compounding signal — 3Y smooths single-quarter "
     "distortions. v10.14: capped at ±500%.",
     "Full Dashboard"),
    ("GROWTH","EBITDA CAGR 1Y %",
     "Latest FY EBITDA ÷ prior FY EBITDA − 1. >15% strong operating growth. "
     "EBITDA recovery from near-zero prior-year base can produce outsized "
     "percentages — v10.14 caps at ±500% to prevent tiny-base noise.",
     "Full Dashboard"),
    ("GROWTH","Rev YoY %",
     "TRAILING TWELVE-MONTH revenue growth, from yfinance "
     ".info['revenueGrowth'] × 100. Rolling 4-quarter comparison — NOT the "
     "same as Rev CAGR 1Y % which uses discrete fiscal years. The two diverge "
     "most in insurance/NBFC stocks (premium accounting) and post-restructuring "
     "companies. v10.14: capped at ±500% to filter yfinance junk signals "
     "on micro-caps (e.g., 14,000%+ readings from tiny revenue bases).",
     "All sheets"),
    ("GROWTH","PAT YoY %",
     "TRAILING TWELVE-MONTH earnings growth, from yfinance "
     ".info['earningsGrowth'] × 100. Same TTM vs fiscal-year distinction as "
     "Rev YoY %. v10.14: capped at ±500%.",
     "All sheets"),
    ("GROWTH","Q3 Rev (₹Cr)",
     "3rd-most-recent quarter's revenue in ₹ Crore. "
     "Source: yfinance quarterly_income_stmt 3rd column. Shows '—' if "
     "quarterly data unavailable.",
     "Full Dashboard"),
    ("GROWTH","Q3 PAT (₹Cr)",
     "3rd-most-recent quarter's net profit in ₹ Crore. "
     "Shows '—' if loss-making that quarter (NULL in DB).",
     "Full Dashboard"),
    ("GROWTH","Q3 EBITDA (₹Cr)",
     "3rd-most-recent quarter's EBITDA in ₹ Crore. "
     "Falls back to 'operating income' row if EBITDA not reported. "
     "Q3 EBITDA Margin = Q3 EBITDA / Q3 Rev — compare vs TTM.",
     "Full Dashboard"),
    ("FIN HEALTH","D/E Ratio",
     "Debt ÷ Equity. "
     "0 = Zero debt (fortress balance sheet) | "
     "0–0.3 = Very low debt (conservative, safe) | "
     "0.3–1.0 = Moderate (acceptable for most sectors) | "
     "1.0–2.0 = High (acceptable only for banking/NBFC/infra sectors) | "
     "> 2.0 = Very high (risk of default, avoid unless sector-specific justification) | "
     "Note: Banks naturally run higher D/E (deposits = debt)","All sheets"),
    ("FIN HEALTH","Current Ratio",
     "Current Assets ÷ Current Liabilities — short-term liquidity. "
     "< 1.0 = Danger (cannot cover near-term obligations) | "
     "1.0–1.5 = Tight (manage carefully) | "
     "1.5–2.0 = Healthy (comfortable) | "
     "> 2.0 = Strong (ample short-term liquidity)","Full Dashboard"),
    ("FIN HEALTH","FCF (₹Cr)",
     "Free Cash Flow = Operating CF − Capex in ₹ Crore. Positive + growing = "
     "healthy cash generation. Source: yfinance .info['freeCashflow'] ÷ 1e7.",
     "Full Dashboard"),
    ("FIN HEALTH","CCC Days",
     "Cash Conversion Cycle = Days Inventory Outstanding + Days Sales "
     "Outstanding − Days Payables Outstanding. Lower = more efficient. "
     "Negative CCC = collects cash before paying suppliers (ideal — FMCG/retail). "
     "Source: (inventory + receivables − payables) / revenue × 365. "
     "v10.15 FIX #3: capped at ±500 days. Computation skipped when revenue "
     "< ₹0.1 Cr (prevents arithmetic noise — EMAMIREAL was showing 16,821 "
     "days = 46 years before fix). "
     "v12.5: skipped entirely (renders '—') for Banks / NBFCs / HFCs / "
     "Insurance — the metric is meaningless for finance-sector stocks. "
     "TATACAP showed 7,739 days, FUSION 3,216 in prior runs.",
     "Full Dashboard"),
    ("SHAREHOLDING","Promoter %",
     "Promoter holding share. <25% = low conviction (consider Ownership "
     "red flag); 0% = Hard Drop from Gold filter. "
     "Source: NSE corp-info API primary; yfinance heldPercentInsiders fallback. "
     "Rising over time = promoter conviction signal.",
     "All sheets"),
    ("SHAREHOLDING","Pro QoQ Δ",
     "Promoter shareholding change vs previous quarter (percentage points). "
     ">+0.5 = promoter buying (Sentiment +5); <−0.5 = selling (Sentiment −5). "
     "Source: computed as (current promoter_pct − prior-quarter promoter_pct). "
     "v13.0: now populates immediately from NSE corp-info JSON which returns "
     "2–4 quarters of history in a single call (zero new API cost). Pre-v13.0 "
     "depended on a 3-month rolling window of daily runs to accumulate prior "
     "data. '—' = no prior quarter from NSE OR symbol failed corp-info call.",
     "Full Dashboard"),
    ("SHAREHOLDING","Pledge %",
     "Percentage of promoter shares pledged as loan collateral. "
     "0% = clean capital structure (Safety +4); 10–20% = watch (−7); "
     ">20% = RED FLAG (Safety −15 + suppresses all spike signals). "
     "Source: NSE bulk pledge endpoint (corporates-pledgedata, free, daily). "
     "v13.0 makes this real for the first time — pre-v13.0 was hardcoded 0 "
     "from yfinance fallback. Shows '—' when no record in NSE feed (most "
     "stocks have zero pledge so they don't appear in the report). Score "
     "gates fire on numeric values; '—' treated as 0 for guard purposes.",
     "All sheets"),
    ("SHAREHOLDING","Pledge Direction",
     "FALLING = promoters repaying loans (positive); RISING = more shares "
     "pledged (risk signal); STABLE = unchanged at non-zero level; "
     "'—' = no pledge data (free-tier limitation).",
     "Full Dashboard"),
    ("SHAREHOLDING","FII %",
     "Foreign Institutional Investor holding %. >15% = institutional backing; "
     ">25% = high global interest. Rising FII over quarters = smart-money "
     "accumulation signal.",
     "Full Dashboard"),
    ("SHAREHOLDING","FII QoQ Δ",
     "FII holding change vs previous quarter. >+1 pp = strong accumulation "
     "(Early Entry +8); <−1 = distribution. v13.0: populates immediately "
     "from NSE corp-info multi-quarter response (single call, zero new API "
     "cost). '—' when no prior quarter available.",
     "Full Dashboard"),
    ("SHAREHOLDING","DII %",
     "Domestic Institutional Investor (MF/insurance/banks) holding %. "
     ">10% = domestic confidence. Rising DII + FII = dual institutional "
     "accumulation = bullish signal. "
     "Source: NSE corp-info JSON API (heldPercentInstitutions in yfinance "
     "is FII+DII combined; DII alone needs NSE). v10.15 FIX #6: shows '—' "
     "when 0 because NSE API is commonly blocked on cloud IPs (GitHub "
     "Actions etc.) — 0 is indistinguishable from 'API blocked' on free-tier.",
     "Full Dashboard"),
    ("SHAREHOLDING","DII QoQ Δ",
     "DII holding change vs previous quarter. >+0.5 pp = strong domestic "
     "accumulation (Sentiment +6); <−0.3 = distribution (Sentiment −3). "
     "v13.0: populates from NSE corp-info multi-quarter response. '—' when "
     "no prior quarter computable.",
     "Full Dashboard"),
    ("SHAREHOLDING","Public Float %",
     "Percentage of shares NOT held by promoter or institutions. "
     "Derived as 100 − Promoter% − FII% − DII%. Lower = tighter float = "
     "more volatile; higher = more liquid but less concentrated control.",
     "Full Dashboard"),
    ("QUALITY SCORES","Piotroski F /9","9-point business health score computed from free yfinance data (Session 14 wire-up). ≥7=strong, ≤3=weak. Typical distribution on a real run: 4-8 range","All sheets"),
    ("QUALITY SCORES","Altman Z","Bankruptcy predictor. >2.99=safe, <1.81=distress zone. Requires paid balance-sheet feed (working capital, retained earnings, EBIT, total liabilities, total assets) — displays '—' when missing. v12.5: clamped at 10 (Z>7 already signals exceptional safety; >10 values in production were typically unit-mismatch artefacts in the X4 = mcap/total_liab component).","Full Dashboard"),
    ("QUALITY SCORES","Beneish M","Manipulation risk score (Beneish 1999). >-2.22=manipulation flag, <-2.22=likely honest. v12.9: real 8-variable formula (DSRI, GMI, AQI, SGI, DEPI, SGAI, TATA, LVGI) using yfinance prior-year balance-sheet/income/cashflow data. Falls back to 3-bucket accrual proxy (-2.5/-2.22/-1.5) when only current-year data available. Pre-v12.9 was proxy-only (4 unique values across 100 stocks). Displays '—' when forensic inputs missing","Full Dashboard"),
    ("PIPELINE / OB","OB/Bill Ratio","Order Book ÷ Revenue. >1.5× = strong pipeline","All sheets"),
    ("EARLY DETECTION","Early Entry /100","12 signals: quiet accum, SME migration, analyst imminent, sector Stage 1. Session 23: low EE on Gold is OK — VALUE archetype (high Score + MoS + clean safety) qualifies without momentum signals","All sheets"),
    ("EARLY DETECTION","Sector Stage",
     "Stage 1 = Early Accumulation (sector just starting to move, best entry point, low risk/reward) | "
     "Stage 2 = Confirmed Uptrend (momentum building, institutional buying, good entry) | "
     "Stage 3 = Peak / Euphoria (sector fully priced, risk of reversal, tighten stops) | "
     "Stage 4 = Distribution / Decline (smart money exiting, avoid fresh entry)","All sheets"),
    ("TECHNICAL","SMA 200",
     "200-Day Simple Moving Average — long-term trend indicator. "
     "ABOVE = CMP above 200 SMA, stock in long-term uptrend (bullish) | "
     "BELOW = CMP below 200 SMA, stock in long-term downtrend (bearish) | "
     "Golden rule: only buy stocks ABOVE their 200 SMA","Full Dashboard"),
    ("TECHNICAL","Supertrend",
     "ATR-based trend-following indicator. "
     "BUY = price above supertrend line (uptrend confirmed, go long) | "
     "SELL = price below supertrend line (downtrend, exit or avoid) | "
     "NEUTRAL = indeterminate / sideways market","Full Dashboard"),
    ("TECHNICAL","MACD Signal",
     "Moving Average Convergence Divergence. "
     "BUY = MACD line crossed above signal line (bullish momentum) | "
     "SELL = MACD line crossed below signal line (bearish momentum) | "
     "NEUTRAL = no recent crossover","Full Dashboard"),
    ("TECHNICAL","RSI (14)",
     "Relative Strength Index, 0–100. "
     "< 30 = Oversold (potential reversal up, look for buy signal) | "
     "30–50 = Bearish zone (weak, avoid fresh entry) | "
     "50–60 = Neutral (no strong signal) | "
     "60–70 = Bullish zone (momentum building, good entry on dips) | "
     "> 70 = Overbought (potential reversal down, wait for pullback) | "
     "Best entry: RSI 50–65 with MACD BUY and Supertrend BUY","Full Dashboard"),
    ("TECHNICAL","ADX",
     "Average Directional Index — measures trend strength (0–100), not direction. "
     "< 20 = Weak/no trend (sideways, avoid trend strategies) | "
     "20–25 = Emerging trend (watch closely) | "
     "25–40 = Strong trend (good for momentum entry) | "
     "> 40 = Very strong trend (ride with trailing stop)","Full Dashboard"),
    ("TECHNICAL","OBV Signal",
     "On-Balance Volume — tracks smart money flow. "
     "ACCUMULATION = OBV rising even when price flat (institutions buying quietly, bullish) | "
     "DISTRIBUTION = OBV falling while price holds up (institutions selling, bearish) | "
     "NEUTRAL = no clear pattern","Full Dashboard"),
    ("TECHNICAL","Above VWAP",
     "Volume Weighted Average Price — intraday fair value benchmark. "
     "YES = CMP above VWAP (buyers in control, bullish intraday) | "
     "NO = CMP below VWAP (sellers in control, wait for VWAP reclaim)","Full Dashboard"),
    ("BALANCE SHEET","BS Health Flag",
     "HEALTHY = cash > debt, D/E < 1.0, no pledge risk, FCF positive (safe to invest) | "
     "WATCH = moderate debt or pledge 10–20%, monitor quarterly results | "
     "ALERT = high debt (D/E>2), pledge>20%, negative FCF, or interest coverage<1.5 (extra caution)","All sheets"),
    ("TRADE PLAN","R:R Ratio",
     "Risk:Reward = (Target1 − Entry midpoint) ÷ (Entry midpoint − Stop Loss). "
     "< 1:1 = Poor (avoid) | 1:1 to 2:1 = Acceptable | "
     "2:1 to 3:1 = Good (standard for positional trades) | "
     "> 3:1 = Excellent (asymmetric payoff). "
     "Session 22: Target1 auto-derived to keep R:R ≥ 2.0 — uses max of "
     "(Entry + 2× risk distance) and CFV-weighted target.","Trade Summary"),
    ("TRADE PLAN","Risk Level",
     "LOW = large cap, low beta, positive MoS, strong balance sheet | "
     "MEDIUM = mid cap or slight premium or moderate debt | "
     "HIGH = small/micro cap or negative MoS or high D/E or high beta","Trade Summary"),
    ("ANALYSIS SUMMARY","View Analysis Summary","150–250 word AI note with exact ₹ figures, catalysts, risks","All sheets"),

    # ── IDENTITY ──────────────────────────────────────────────────────────────
    ("IDENTITY","Company Name","Full registered company name on NSE/BSE","All sheets"),
    ("IDENTITY","Sector",
     "Sector classified by NSE/BSE/yfinance. Used for sector-median PE and sector rotation stage. "
     "Examples: Information Technology, Financial Services, Automobiles, Pharmaceuticals, Energy","All sheets"),

    # ── PRICE & MARKET ────────────────────────────────────────────────────────
    ("PRICE & MARKET","CMP",
     "Current Market Price — closing price from today's bhav copy (NSE/BSE). "
     "All valuations and ratios are computed using this price.","All sheets"),
    ("PRICE & MARKET","Day Chg %",
     "Price change % today vs yesterday's close. "
     ">2% with high delivery = strong buying signal | <-2% = watch for panic or news","All sheets"),
    ("PRICE & MARKET","52W High",
     "Highest closing price in the past 52 weeks. "
     "CMP within 10% of 52W High = stock near peak, caution. CMP near 52W Low = potential value opportunity.","Full Dashboard"),
    ("PRICE & MARKET","52W Low",
     "Lowest closing price in the past 52 weeks. "
     "CMP at 52W Low with good fundamentals = potential deep value entry.","Full Dashboard"),

    # ── FAIR VALUE ────────────────────────────────────────────────────────────
    ("FAIR VALUE","CFV (₹)",
     "Composite Fair Value — sector-weighted average of all 7 models (M1–M7) that return a valid value. "
     "CFV > CMP = undervalued. CFV < CMP = overvalued.","All sheets"),
    ("FAIR VALUE","FV Low (₹)",
     "Bear-case fair value = CFV × 0.85 (15% margin of safety buffer). "
     "If CMP < FV Low, stock is deeply undervalued even in a pessimistic scenario.","Full Dashboard"),
    ("FAIR VALUE","FV High (₹)",
     "Bull-case fair value = CFV × 1.15. "
     "If CMP > FV High, stock is overvalued even optimistically.","Full Dashboard"),
    # Session 23: "Upside to FV %" glossary entry moved to FAIR VALUE section later
    # (kept single authoritative entry rather than two).
    ("FAIR VALUE","MoS Label",
     "Text label for MoS %: "
     "EXCEPTIONAL (>40%) | STRONG (>25%) | ADEQUATE (>10%) | THIN (0–10%) | "
     "SLIGHT PREMIUM (0 to −15%) | SIGNIFICANT PREMIUM (<−15%). "
     "v12.5: trailing '*' (e.g., 'EXCEPTIONAL*') signals the CFV hit the 3× CMP "
     "safety cap — the underlying models projected even higher upside but the "
     "cap fired. Treat such cells with extra scrutiny. "
     "v12.6: trailing '†' (e.g., 'EXCEPTIONAL†') signals the CFV is based on "
     "fewer than 3 valuation models (M1–M7) firing — the FV evidence is thin. "
     "The CFV value is still shown but the automatic +score bonus is suppressed. "
     "Markers can stack: '*†' means BOTH conditions fired.","All sheets"),
    ("FAIR VALUE","M1: DCF FV (₹)","3-Stage Discounted Cash Flow. WACC = 10Y GSec + Beta×5.5%. Terminal growth 4.5%. Best for steady compounders. Session 19 cap: M1 limited to 4× CMP so low-beta stocks (e.g., SBIN β=0.2) don't produce absurd DCF outputs. v12.2: eps now sanitized via _sf() (handles '—' / 'N/A' / None).","Full Dashboard"),
    ("FAIR VALUE","M2: Graham FV (₹)","Graham Number = √(22.5 × EPS × BVPS). Benjamin Graham's intrinsic value formula. Best for value stocks with positive EPS. v12.2: eps/bvps sanitized via _sf(); BVPS fallback derives from close/PB if missing.","Full Dashboard"),
    ("FAIR VALUE","M3: PE FV (₹)","EPS × Sector 5yr median P/E. Mean-reversion model — assumes P/E reverts to sector norm. v12.2: 28-sector benchmarks via SECTOR_ALIASES — production sectors (Basic Materials → Metals PE 12, Industrials → Infra PE 22, Communication Services → Telecom PE 22, Consumer Cyclical/Defensive → Consumer PE 40, Financial Services → Financial PE 20, Real Estate → Realty PE 25) now resolve correctly.","Full Dashboard"),
    ("FAIR VALUE","M4: PB FV (₹)","BVPS × Sector median Price/Book. Best for asset-heavy sectors: banks, metals, real estate. v12.2: same SECTOR_ALIASES path as M3 — Basic Materials gets PB 1.5 (Metals), Industrials gets PB 2.5 (Infra), Communication Services gets PB 2.5 (Telecom).","Full Dashboard"),
    ("FAIR VALUE","M5: EV FV (₹)","CMP × (Sector median EV/EBITDA ÷ Stock EV/EBITDA). Best for capital-intensive businesses. v12.2: 28-sector benchmarks via SECTOR_ALIASES. v12.3 Round 2: proper EV-based formula primary — fair_per_share = CMP × ((annual_ebitda × sector_mult − net_debt) / mcap_cr). Tier 1 fires when q_ebitda_cr + total_debt_cr + cash_cr + mcap_cr all populated; Tier 2 falls back to v12.2 shortcut; Tier 3 skips for Banks/NBFCs/Insurance entirely. 4× CMP cap on Tier 1 outliers; 70% discount when fair equity goes negative (severely overlevered).","Full Dashboard"),
    ("FAIR VALUE","M6: DDM FV (₹)",
     "Gordon Growth Model = DPS×(1+g)/(r−g). "
     "Only shown for dividend-paying stocks (yield 0.1%–15%). "
     "DPS = CMP × Div Yield. r = GSec + 4.5%. g = max(min(pat_yoy/200, 6%), 0%). "
     "v12.2 fix: removed the 2% growth floor that previously inflated FV by ~26% "
     "for declining-earnings stocks (pat_yoy<0) — they now correctly get 0% div growth.","Full Dashboard"),
    ("FAIR VALUE","M7: PEG FV (₹)","EPS × EPS Growth Rate × PEG_BENCHMARK. PEG=1 baseline (Lynch's rule). Best for high-growth stocks where PE alone overstates expensiveness. v12.3 Round 2: PEG_BENCHMARK = 1.0 now an explicit named constant — tunable for value-tilted (0.8) or growth-tilted (1.2) mandates. v12.2 unit guard still active: skips when growth_3yr < 1.0 (catches decimal-fraction unit error).","Full Dashboard"),

    # ── VALUATION ─────────────────────────────────────────────────────────────
    ("VALUATION","P/B",
     "Price ÷ Book Value per Share. "
     "<1 = trading below book (possible deep value or value trap) | "
     "1–2 = reasonable (banks, metals) | 2–5 = premium (justified if ROE>15%) | "
     ">5 = high (asset-light compounders only) | "
     "'—' = raw value ≥ 500 (tiny-book-value distortion; v10.16 Option B)",
     "All sheets"),
    ("VALUATION","P/S",
     "Price-to-Sales = MCap ÷ Annual Revenue. "
     "<1 = very cheap (cyclicals, PSUs) | 1–3 = reasonable | 3–8 = premium | "
     ">8 = expensive (only for high-margin businesses) | "
     "'—' = raw value ≥ 500 (v10.16 Option B)","Full Dashboard"),
    ("VALUATION","P/CF",
     "Price-to-Cash Flow = MCap ÷ Operating Cash Flow. "
     "Better than P/E as cash flow is harder to manipulate. "
     "<10 = cheap | 10–20 = fair | 20–35 = premium | >35 = expensive. "
     "Derived as P/S ÷ EBITDA margin when direct cash flow data unavailable.","Full Dashboard"),
    ("VALUATION","EV/EBITDA",
     "Enterprise Value ÷ EBITDA. Ignores capital structure — better for comparing levered vs unlevered firms. "
     "<8 = cheap | 8–15 = fair | 15–25 = premium | >25 = expensive","Full Dashboard"),

    # ── PROFITABILITY ─────────────────────────────────────────────────────────
    ("PROFITABILITY","ROCE %",
     "Return on Capital Employed = EBIT ÷ Capital Employed × 100. "
     "Measures how efficiently the company uses ALL capital (debt + equity). "
     ">20% = excellent | 15–20% = good | 10–15% = average | <10% = poor. "
     "Derived from EBITDA margin × Revenue / Capital Employed when direct data unavailable.","Full Dashboard"),
    ("PROFITABILITY","ROA %",
     "Return on Assets = Net Income ÷ Total Assets × 100. "
     ">10% = excellent | 5–10% = good | <5% = poor. "
     "Derived as ROE ÷ (1 + D/E) when direct data unavailable. "
     "v12.4: clamped to ±100 % (filters unit-mismatch outliers).","Full Dashboard"),
    ("PROFITABILITY","Gross Mgn %",
     "Gross Profit ÷ Revenue × 100. Revenue minus direct costs (raw materials, COGS). "
     ">50% = high-margin business (software, pharma) | >30% = good | <15% = commodity/trading. "
     "v12.4: clamped to [0, 100] %.","Full Dashboard"),
    ("PROFITABILITY","EBITDA Mgn %",
     "EBITDA ÷ Revenue × 100. Operating profitability before interest, tax, depreciation. "
     ">25% = excellent | 15–25% = good | 8–15% = average | <8% = tight. "
     "v12.4: clamped to ±100 %.","Full Dashboard"),
    ("PROFITABILITY","NPM %",
     "Net Profit Margin = Net Income ÷ Revenue × 100. Bottom-line profitability after everything. "
     ">15% = excellent | 8–15% = good | 3–8% = average | <3% = thin (watch for debt servicing risk). "
     "v12.4: clamped to ±100 % (filters thin-revenue artefacts).","Full Dashboard"),
    ("PROFITABILITY","NPM Q (latest) %","Net Profit Margin for the most recent reported quarter. Source: yfinance quarterly_income_stmt. Rising trend across Q-2 → Q-1 → Q(latest) signals Margin Expansion. v12.6 (#11): renamed from 'NPM Q1 %' for chronological clarity.","Full Dashboard"),
    ("PROFITABILITY","NPM Q-1 %","Net Profit Margin for the previous quarter (one before latest). Source: yfinance quarterly_income_stmt. Compare with Q(latest) and Q-2 for trend. v12.6 (#11): renamed from 'NPM Q2 %'.","Full Dashboard"),
    ("PROFITABILITY","NPM Q-2 %","Net Profit Margin for two quarters before the latest report. Source: yfinance quarterly_income_stmt. Oldest of the 3 quarters shown. v12.6 (#11): renamed from 'NPM Q3 %'.","Full Dashboard"),
    ("PROFITABILITY","Margin Expansion",
     "YES = NPM has risen for 3 consecutive quarters (Q-2 → Q-1 → Q(latest), oldest to newest). "
     "Source: derived from NPM Q (latest) / Q-1 / Q-2 via yfinance. Strong signal of operational leverage or pricing power.","Full Dashboard"),

    # ── GROWTH ────────────────────────────────────────────────────────────────
    # v10.14: The complete GROWTH glossary entries (10 fields with TTM-vs-FY
    # clarification + cap note) now live higher in this list at the PROFITABILITY
    # → GROWTH transition. The legacy duplicate block that was here has been
    # removed to keep the glossary deduplicated.

    # ── FIN HEALTH ────────────────────────────────────────────────────────────
    ("FIN HEALTH","ND/EBITDA",
     "Net Debt ÷ EBITDA. How many years of earnings needed to repay net debt. "
     "<1× = very safe | 1–2× = manageable | 2–3× = stretched | >3× = high leverage. "
     "No free source — requires balance sheet detail.","Full Dashboard"),
    ("FIN HEALTH","Int Coverage",
     "EBIT ÷ Interest Expense. How many times earnings cover interest payments. "
     ">5× = safe | 3–5× = adequate | 1.5–3× = risky | <1.5× = danger. "
     "No free source — requires income statement detail.","Full Dashboard"),
    ("FIN HEALTH","Cash (₹Cr)",
     "Total cash and equivalents in ₹ Crore. "
     "High cash vs debt = fortress balance sheet. "
     "Cash > Total Debt = net cash company (very safe, often undervalued).","Full Dashboard"),
    ("FIN HEALTH","Total Debt (₹Cr)",
     "Total financial debt (short + long term) in ₹ Crore. "
     "0 = zero-debt company. Derived as D/E × Book Equity when direct data unavailable. "
     "Compare with Cash to get Net Debt.","Full Dashboard"),
    ("FIN HEALTH","FCF Yield %",
     "Free Cash Flow ÷ MCap × 100. How much FCF you get per ₹ invested. "
     ">5% = very good | 3–5% = good | 1–3% = modest | <1% = poor. "
     "Often proxy-computed from Operating Cash Flow when FCF not available.","Full Dashboard"),
    ("FIN HEALTH","CCC Days",
     "Cash Conversion Cycle = DIO + DSO − DPO. Days to convert inventory investment into cash. "
     "Lower = better. <30 days = excellent. No free source — requires AR/AP/inventory data. "
     "v12.5: renders '—' for Banks/NBFCs/HFCs/Insurance (metric not meaningful for finance sector).","Full Dashboard"),

    # ── CAP ALLOC ─────────────────────────────────────────────────────────────
    ("CAP ALLOC","Div Yield %",
     "Annual dividend ÷ CMP × 100. Income return on investment. "
     ">4% = high yield (income stock) | 1–4% = moderate | 0–1% = growth-focused | "
     "0% = no dividend (all profits reinvested). Normalised from yfinance fractions.","All sheets"),
    ("CAP ALLOC","Payout Ratio %",
     "Dividends ÷ Net Profit × 100. Percentage of earnings paid as dividends. "
     "<30% = growth company (retains earnings) | 30–60% = balanced | "
     ">70% = high payout (mature/PSU/FMCG). >100% = paying from reserves (unsustainable).","Full Dashboard"),
    ("CAP ALLOC","Capex / Rev %",
     "Capital Expenditure ÷ Revenue × 100. Reinvestment intensity. "
     ">15% = capital-heavy (infra, steel, telecom) | 3–15% = moderate | "
     "<3% = asset-light (IT, FMCG). No free source — requires cash flow statement.","Full Dashboard"),

    # ── SHAREHOLDING ──────────────────────────────────────────────────────────
    ("SHAREHOLDING","Pro QoQ Δ",
     "Promoter holding change quarter-over-quarter (%). "
     "Increasing = promoters buying → bullish signal. "
     "Decreasing = promoters selling → investigate reason. "
     "v13.0: populated from NSE corp-info multi-quarter response.","Full Dashboard"),
    ("SHAREHOLDING","Pledge Direction",
     "Direction of pledge change: FALLING / RISING / STABLE. "
     "RISING pledge is a red flag — promoters may be under financial stress. "
     "v13.0: vocabulary aligned with rest of pipeline (was IMPROVING/DETERIORATING).","Full Dashboard"),
    ("SHAREHOLDING","DII %",
     "Domestic Institutional Investor holding %. "
     "DII (mutual funds, insurance) rising = domestic smart money accumulating. "
     "v13.0: separated from FII via NSE corp-info diisTotal field.","Full Dashboard"),
    ("SHAREHOLDING","DII QoQ Δ","DII holding change quarter-over-quarter. "
     "v13.0: from NSE corp-info multi-quarter history.","Full Dashboard"),
    ("SHAREHOLDING","FII QoQ Δ","FII holding change quarter-over-quarter. "
     "v13.0: from NSE corp-info multi-quarter history.","Full Dashboard"),
    ("SHAREHOLDING","Public Float %",
     "% of shares held by retail/public (100% − Promoter% − Institutional%). "
     "Higher float = more liquid, lower impact cost for large orders.","Full Dashboard"),

    # ── QUALITY SCORES ────────────────────────────────────────────────────────
    ("QUALITY SCORES","Earn Quality",
     "Qualitative assessment: HIGH / MODERATE / LOW / —. "
     "Checks if earnings are backed by cash flow (CFO/PAT ratio). "
     "HIGH ≥ 0.8 (cash-backed), 0.5–0.8 MODERATE, < 0.5 LOW (red flag — profits not converting to cash). "
     "v12.9: PAT now annualized (q_pat × 4) to match yfinance's annual CFO — "
     "pre-v12.9 raw quarterly comparison gave 4× inflated ratios, "
     "causing 70% of stocks to falsely score HIGH.","Full Dashboard"),

    # ── PIPELINE / OB ─────────────────────────────────────────────────────────
    ("PIPELINE / OB","Pipeline Vis",
     "Pipeline Visibility: HIGH / MEDIUM / LOW / NONE. "
     "For defence, infra, capital goods stocks — is the revenue pipeline visible for next 12–24 months?","Full Dashboard"),
    ("PIPELINE / OB","L1 Wins 90D",
     "Number of L1 (lowest bid) wins in government tenders in the last 90 days. "
     "Only relevant for defence, EPC, infra companies. Source: CPP portal / BSE announcements.","Full Dashboard"),
    ("PIPELINE / OB","L1 Est (₹Cr)",
     "Estimated value of L1 wins in ₹ Crore over last 90 days. "
     "Compare with annual revenue for order-book visibility.","Full Dashboard"),
    ("PIPELINE / OB","New Mkt Entry",
     "Has the company announced entry into a new market/geography/product in the last quarter? "
     "YES = potential re-rating catalyst.","Full Dashboard"),

    # ── EARLY DETECTION ───────────────────────────────────────────────────────
    ("EARLY DETECTION","Early Signals",
     "Pipe-separated list of active early-mover triggers detected. "
     "Examples: INSTITUTIONAL FOOTPRINT | VOL SURGE + RSI ACCUMULATION | TREND CONFLUENCE | "
     "MOMENTUM BUILDING | DUAL-LISTED DISCOVERY","Full Dashboard"),
    ("EARLY DETECTION","Smart Money",
     "ACCUMULATION = FII+DII buying trend detected (3Q rising) | "
     "DISTRIBUTION = Institutional selling | NEUTRAL = no clear trend. "
     "Signals smart money positioning ahead of price move.","Full Dashboard"),

    # ── TECHNICAL ─────────────────────────────────────────────────────────────
    ("TECHNICAL","Stoch %K",
     "Stochastic Oscillator %K — momentum indicator comparing closing price to price range. "
     "<20 = Oversold (potential reversal up) | 20–40 = Bearish zone | "
     "40–60 = Neutral | 60–80 = Bullish zone | >80 = Overbought (potential reversal down). "
     "Best signal: %K crossing above 20 from below = buy; crossing below 80 from above = sell.","Full Dashboard"),
    ("TECHNICAL","MFI",
     "Money Flow Index — volume-weighted RSI. Uses both price AND volume. "
     "<20 = Oversold (strong reversal signal, especially with rising volume) | "
     ">80 = Overbought | 40–60 = Neutral. "
     "MFI divergence from price = early warning of trend reversal.","Full Dashboard"),
    ("TECHNICAL","Chart Pattern",
     "Most recent chart pattern detected from OHLC: "
     "BULLISH CANDLE / BEARISH CANDLE / DOJI (indecision) / "
     "HAMMER (bullish reversal) / HANGING MAN (bearish reversal) / "
     "SHOOTING STAR (bearish reversal) / NEUTRAL / "
     "UPPER CIRCUIT (hit upper price band) / LOWER CIRCUIT (hit lower band) / "
     "'—' when OHLC data incomplete. "
     "Candlestick patterns are more reliable near key support/resistance levels.","Full Dashboard"),
    ("TECHNICAL","Support 1 (₹)",
     "Short-term support = 20-day rolling low. Nearest price floor — "
     "place stop loss slightly below Support 1.","Full Dashboard"),
    ("TECHNICAL","Support 2 (₹)",
     "Major long-term support = prior 52-week low (v12.4; excludes last 20d). "
     "Computed as the rolling 252-day min over bars BEFORE the most recent 20, "
     "so a fresh breakdown doesn't make S1 == S2. "
     "If price breaks Support 1 with volume, next target is Support 2. "
     "Falls back to '—' for stocks with < 80 days of price history.","Full Dashboard"),
    ("TECHNICAL","Resist 1 (₹)",
     "Short-term resistance = 20-day rolling high. Nearest price ceiling — "
     "Target 1 is typically set at Resist 1.","Full Dashboard"),
    ("TECHNICAL","Resist 2 (₹)",
     "Major long-term resistance = prior 52-week high (v12.4; excludes last 20d). "
     "Computed as the rolling 252-day max over bars BEFORE the most recent 20. "
     "Earlier v10.9 logic (rolling-252 over the full series) silently mirrored "
     "Resist 1 whenever the 52-week max landed inside the last 20 days — "
     "observed in 87.9 % of production rows. v12.4 separates them. "
     "Target 2 / T3 set at Resist 2.","Full Dashboard"),

    # ── BALANCE SHEET ─────────────────────────────────────────────────────────
    ("BALANCE SHEET","BS Health Note",
     "Detailed note explaining BS Health Flag. "
     "Examples: 'No red flags detected' | 'High D/E 2.1x — monitor debt serviceability' | "
     "'Pledge 23% — elevated risk'","Full Dashboard"),

    # ── TRADE PLAN ────────────────────────────────────────────────────────────
    ("TRADE PLAN","Entry Range (₹)",
     "Suggested entry price band: Low–High in ₹. "
     "Based on CMP ± 1.5% to account for intraday spread. "
     "Use limit orders within this range for best execution.","All sheets"),
    ("TRADE PLAN","Stop Loss (₹)",
     "Mandatory stop loss price. Exit the trade if CMP closes below this on a daily basis. "
     "Set at ~3–5% below entry, or just below key support level. "
     "ALWAYS place stop loss before entering any trade.","All sheets"),
    ("TRADE PLAN","Target 1 (₹)",
     "First price target = nearest resistance level. "
     "Book 40–50% of position at T1. Let rest run with trailing stop.","All sheets"),
    ("TRADE PLAN","Target 2 (₹)",
     "Second price target = next resistance zone. "
     "Book another 30% at T2. Trail stop to entry cost.","All sheets"),
    ("TRADE PLAN","Target 3 (₹)",
     "Final price target = CFV (Composite Fair Value). "
     "Hold remaining 20–30% position for full re-rating. "
     "Only applicable for BUY stocks with positive MoS.","All sheets"),

    # ── NEWS & RISK ───────────────────────────────────────────────────────────
    ("NEWS & RISK","Key Catalyst",
     "Primary upcoming event that could trigger price re-rating. "
     "Examples: order win, product launch, QIP, promoter buyback, index inclusion. "
     "Requires Gemini API credits — populated by AI analyst (aistudio.google.com).","Full Dashboard"),
    ("NEWS & RISK","News Sentiment",
     "AI-assessed recent news tone: BULLISH / NEUTRAL / BEARISH. "
     "Based on last 30 days of BSE announcements and news. "
     "Requires Gemini API credits.","Full Dashboard"),
    ("NEWS & RISK","Primary Risk",
     "The single most important risk factor for this stock right now. "
     "Examples: regulatory overhang, promoter pledge, client concentration, commodity exposure. "
     "Requires Gemini API credits.","Full Dashboard"),
    ("NEWS & RISK","SEBI Flags",
     "Any active SEBI actions, adjudication orders, or exchange surveillance flags. "
     "NONE = clean | Any other value = investigate before investing. "
     "Requires Gemini API credits.","Full Dashboard"),

    # ── GOLD SHEET SPECIFIC ────────────────────────────────────────────────────
    # v12.5: removed the duplicate "F-Score /9" glossary entry that used
    # to live here for the Gold sheet. The Gold column is now labelled
    # "Piotroski F /9" (matches Full Dashboard), so the QUALITY SCORES
    # entry above this block applies to both sheets.
    ("SCORES","Spike /6",
     "Spike Score 0–6: count of institutional triggers fired. "
     "Triggers: Value Breakout | Perfect Storm | Institutional Accumulation | "
     "Technical Breakout | RSI Accumulation | Dual-Listed Discovery. "
     "≥3=Strong setup | 1–2=Watch | 0=No spike.","All sheets"),
    ("SCORES","Storm /10",
     "Storm Score 0–10: defensive quality in volatile markets. "
     "Rewards low beta, low D/E, positive FCF, high promoter holding, "
     "low pledge and strong cash cover. Higher = safer defensive pick.","All sheets"),
    ("TRADE PLAN","Time Horizon",
     "How long the system claims to hold this pick. "
     "SHORT TERM = 2–4 weeks (spike-driven entry) | "
     "POSITIONAL = 1–3 months (trend confirmed) | "
     "LONG TERM = 6–12 months (value accumulation). "
     "Drives the v14.1 outcome-tracking expiry window: "
     "SHORT TERM → 30 days, POSITIONAL → 90 days, LONG TERM → 270 days.","All sheets"),
    ("VALUATION","P/E",
     "Price-to-Earnings ratio (abbreviated in Gold sheet). "
     "CMP ÷ Trailing 12-month EPS. "
     "<15=Cheap | 15–25=Fair | 25–40=Premium | >40=Expensive.","Gold Sheet"),
    ("VALUATION","PEG",
     "PEG Ratio (abbreviated). P/E ÷ Earnings Growth Rate. "
     "<1=Undervalued vs growth | 1–2=Fair | >2=Expensive vs growth.","Gold Sheet"),
    ("VALUATION","D/E",
     "Debt-to-Equity ratio (abbreviated in Gold sheet). "
     "<0.5=Low debt | 0.5–1.5=Moderate | >2=High leverage risk.","Gold Sheet"),
    ("TECHNICAL","Pattern",
     "Latest candlestick pattern (abbreviated in Gold sheet — same as Chart Pattern). "
     "BULLISH/BEARISH CANDLE | HAMMER | HANGING MAN | SHOOTING STAR | DOJI | "
     "NEUTRAL | UPPER/LOWER CIRCUIT | '—' if OHLC missing.","Gold Sheet"),
    ("FIN HEALTH","Quick Ratio",
     "Quick Ratio = (Current Assets − Inventory) ÷ Current Liabilities. "
     "More conservative than Current Ratio — excludes slow inventory. "
     ">1=Can meet short-term obligations | <1=Potential liquidity risk.","Full Dashboard"),

    # ── PRICE & MARKET ─────────────────────────────────────────────────────────
    ("PRICE & MARKET","CMP (₹)",
     "Current Market Price — closing price from today's bhav copy (NSE/BSE). "
     "All FV models and ratio calculations use this as the base price.","All sheets"),
    ("PRICE & MARKET","52W High (₹)",
     "Highest closing price in the past 52 weeks. "
     "CMP within 5% of 52W High = near resistance peak, caution on entry. "
     "Breakout above 52W High with high volume = very strong signal.","Full Dashboard"),
    ("PRICE & MARKET","52W Low (₹)",
     "Lowest closing price in the past 52 weeks. "
     "CMP near 52W Low + good fundamentals = potential deep value opportunity.","Full Dashboard"),

    # ── WEEKLY CHANGE ──────────────────────────────────────────────────────────
    ("WEEKLY CHANGE %","Chg% [2-Wk]",
     "Price return over last 2 weeks (10 trading days). "
     "2-Wk rising faster than 4-Wk = accelerating momentum — bullish signal.","All sheets"),
    ("WEEKLY CHANGE %","Chg% [4-Wk]",
     "Price return over last 4 weeks (20 trading days). "
     "Used in Sector Stage scoring and Early Entry detection. "
     ">5%=Strong uptrend | <-5%=Downtrend caution.","All sheets"),
    ("WEEKLY CHANGE %","Chg% [6-Wk]",
     "Price return over last 6 weeks. Medium-term trend confirmation.","Full Dashboard"),
    ("WEEKLY CHANGE %","Chg% [8-Wk]",
     "Price return over last 8 weeks. Confirms whether trend is sustained.","Full Dashboard"),

    # ── GOLD-TIER FILTER (Session 19) ─────────────────────────────────────────
    # v10.8: The 'FAIR VALUE / Upside to FV %' glossary entry was removed alongside
    # the duplicate column itself (MoS % is the single source of truth).
    # Documents why the Gold sheet only shows a small number of stocks per day.
    # Any user looking at the Gold sheet with few or zero stocks can reference
    # this to understand the strict 8-condition filter.
    ("GOLD FILTER","Gold-Tier Definition",
     "A stock qualifies for the Gold – Early Movers sheet only if ALL 11 "
     "conditions are met: (1) Verdict = BUY (not WATCHLIST), (2) Composite "
     "Score ≥ 70, (3) Margin of Safety between 15% and 100%, (4) Storm Score "
     "≥ 5 (defensively sound), (5) RSI ≤ 70 (not overbought), (6) BS Health "
     "Flag ≠ ALERT, (7) Pledge % ≤ 10, (8) not spike-suppressed, "
     "(9) Altman Z ≥ 1.8 or missing (v10.11 — not in distress zone), "
     "(10) Earn Quality ≠ LOW (v10.11 — no accounting concern), "
     "(11) Int Coverage ≥ 1.5× or missing (v10.11 — can service interest). "
     "Some days may show 0-3 stocks; other days 8-12. This is by design — the "
     "filter reflects market reality, not a fixed daily quota.","Gold Sheet"),
    ("GOLD FILTER","Why so few Gold stocks?",
     "The Gold filter is strict by design: 'patient upside, healthy stocks "
     "only'. It rejects (a) stocks the system isn't confident enough to BUY, "
     "(b) stocks already overbought (RSI > 70), (c) stocks with inflated or "
     "shallow margins of safety, (d) anything with balance-sheet red flags "
     "or high pledge, (e) v10.11: stocks in Altman distress zone, with LOW "
     "earnings quality, or with weak interest coverage. Most days the top "
     "3-8 stocks by score will also pass; days with broad distress signals "
     "will produce fewer.","Gold Sheet"),

    # ── FAIR VALUE SAFETY CAP (Session 19) ────────────────────────────────────
    ("FAIR VALUE","CFV Cap (3× CMP)",
     "Composite Fair Value is capped at 3× Current Market Price as a "
     "safety net. This means the MoS column will never exceed approximately "
     "200% even if individual fair-value models produce higher outputs. "
     "Prevents a single broken model (e.g., DCF on a low-beta stock with "
     "tiny terminal-value denominator) from distorting the composite. "
     "If you see MoS near 200%, treat it as 'deeply undervalued by model, "
     "verify inputs' rather than a guaranteed bargain.","Full Dashboard"),
    # Session 27: Removed standalone "M1 DCF Cap (4× CMP)" row — its content
    # is already covered by the main "M1: DCF FV (₹)" row above (Session 26
    # added the Session 19 cap + SBIN β=0.2 example inline to that entry).
    # ──────────────────────────────────────────────────────────────────────
    # v14.0: PERFORMANCE SHEET column entries
    # ──────────────────────────────────────────────────────────────────────
    ("PERFORMANCE","Total Tracked",
     "Total Gold-sheet picks ever logged to gold_recommendations. Counts every stock that has cleared the 11-condition Gold filter at least once. "
     "First-appearance rule: a stock is logged ONCE on its first Gold appearance; re-appearances are skipped until the original recommendation closes "
     "(T1/T2/T3 hit, SL break, or 90-day expiry). This prevents the same stock from inflating sample size with correlated outcomes.","🎯 Performance"),
    ("PERFORMANCE","Closed",
     "Picks where outcome is final — sum of T1_HIT + T2_HIT + T3_HIT + SL_HIT + EXPIRED. Hit rate / SL rate / expiry rate are computed against this denominator. "
     "Closed rows are immutable — the tracker never re-evaluates them, so once a stock hits T1, subsequent T2 movement won't update the row.","🎯 Performance"),
    ("PERFORMANCE","Open",
     "Picks still being monitored — within 90-day tracking window with no SL/T1/T2/T3 event yet. The tracker walks each open row's daily price history "
     "every run, refreshing current_price + max_runup + max_drawdown. Open rows graduate to EXPIRED automatically at day 90 if no event fires.","🎯 Performance"),
    ("PERFORMANCE","Hit Rate (T1+)",
     "(T1_HIT + T2_HIT + T3_HIT) / CLOSED × 100. The fraction of closed picks where price reached at least the first target. The headline measure of system accuracy. "
     "Heuristics: ≥60% = strong predictive value; 40-60% = useful but mixed; <40% = weak signal, review filter logic. "
     "Wait for ≥30 closed picks before drawing conclusions — sample size matters.","🎯 Performance"),
    ("PERFORMANCE","SL Rate",
     "SL_HIT / CLOSED × 100. The fraction of closed picks that broke down through stop loss before any target hit. Lower is better. "
     "System is calibrated for ~6.5% risk per trade; SL rate of 25-35% is normal even for a healthy strategy. >50% suggests entries are too late "
     "(chasing momentum) or SL is set too tight.","🎯 Performance"),
    ("PERFORMANCE","T1 Hit / T2 Hit / T3 Hit",
     "Counts of closed picks bucketed by which target the price reached first. Note: a pick that reached T2 is NOT also counted in T1 — these are mutually exclusive buckets. "
     "T1 = first target (~2× the risk distance from entry); T2 = T1 × 1.05; T3 = CFV-anchored (often 25-30% above entry).","🎯 Performance"),
    ("PERFORMANCE","SL Hit",
     "Count of closed picks where the daily LOW touched/breached the stop loss before any target was hit. We use daily OHLC, so on same-day ties (low ≤ SL AND high ≥ T1), "
     "SL wins by convention — we can't tell intra-day order from daily bars.","🎯 Performance"),
    ("PERFORMANCE","Expired",
     "Count of closed picks that crossed 90 calendar days from recommendation_date without hitting any target or SL. Bucketed separately because they're neither wins nor losses — "
     "the stock simply drifted sideways. High Expired rate suggests targets may be too ambitious for the chosen time horizon.","🎯 Performance"),
    ("PERFORMANCE","Avg Days → T1 / T2 / T3 / SL",
     "Mean calendar days from recommendation_date to the event firing. Useful for setting realistic expectations: if Avg Days→T1 is 20, plan to hold positions ~3 weeks. "
     "Long Avg→SL (>30d) suggests gradual drift after entry; short Avg→SL (<10d) suggests entries on weak setups.","🎯 Performance"),
    ("PERFORMANCE","Max Runup %",
     "Highest unrealized gain reached during the tracking period. Computed as max((daily_high - cmp_at_recommendation) / cmp × 100) across the walk. "
     "For closed rows, this is the peak before the close event. For open rows, this is the running max as of the last tracker run. "
     "Useful diagnostic: if SL fires after a +20% runup, the system held a winner too long — consider trailing-stop logic.","🎯 Performance"),
    ("PERFORMANCE","Max DD %",
     "Worst unrealized loss reached during the tracking period — max negative excursion from cmp_at_recommendation. Computed as min((daily_low - cmp) / cmp × 100). "
     "For SL_HIT rows, this matches the SL distance. For winners, a deep DD before the win shows the system held conviction through volatility.","🎯 Performance"),
    ("PERFORMANCE","Days Held",
     "For open positions only. Calendar days since recommendation_date. When this hits 90 with no SL/T1/T2/T3 event, the next tracker run bumps the row to EXPIRED.","🎯 Performance"),
    ("PERFORMANCE","Archetype (in Performance sheet)",
     "Quick Pick label captured at the moment the stock was logged (frozen — not updated even if today's score/EE would change it). Lets the BY QUICK PICK ARCHETYPE table "
     "measure whether DEEP VALUE EARLY MOVER picks actually outperform DEEP VALUE — answers the question 'are the combo-archetype calls meaningfully different?'","🎯 Performance"),
    # ──────────────────────────────────────────────────────────────────────
    # v14.1: Horizon-aware expiry + reappearance tracking
    # ──────────────────────────────────────────────────────────────────────
    ("PERFORMANCE","Time Horizon",
     "Time-horizon classification at recommendation time, set by master_funnel based on verdict + spike + supertrend + MACD: "
     "SHORT TERM = BUY + Spike count ≥ 2 (high momentum); POSITIONAL = BUY + supertrend BUY + MACD BUY (confirmed trend) OR BUY + Score ≥ 68; LONG TERM = BUY (lower conviction) OR WATCHLIST/NEUTRAL/AVOID. "
     "v14.1 drives the per-recommendation expiry window: SHORT TERM → 30 days, POSITIONAL → 90 days, LONG TERM → 270 days. Frozen at log time so changing constants later doesn't retroactively re-bucket existing rows.","🎯 Performance"),
    ("PERFORMANCE","Days Left",
     "Calendar days remaining until hard expiry. Computed as expiry_days minus days_held. When this hits 0 with no SL/T1/T2/T3 event, the next tracker run buckets the row as EXPIRED at the exact horizon-derived cutoff. "
     "HARD CUTOFF — no grace period, no extensions. When ≤14, the open-positions row is highlighted in pale yellow with a ⚠ flag to surface positions you may want to manually review before forced expiry.","🎯 Performance"),
    ("PERFORMANCE","Re-app",
     "Times this stock re-appeared in Gold sheet on subsequent days while the original recommendation is still being tracked. First-appearance rule keeps the original entry/SL/T1/T2/T3 frozen — re-appearances do NOT update targets "
     "(preserves measurement integrity by not letting later optimism rescue earlier calls). The counter exists for diagnostic visibility: high count = 'the system kept saying buy this' which is a stronger conviction signal than a single appearance. "
     "Idempotent within a calendar day: same-day pipeline re-runs don't double-count.","🎯 Performance"),
    ("PERFORMANCE","Approaching Expiry Warning",
     "⚠ flag (last column of Open Positions table) and pale-yellow row highlight when Days Left ≤ 14. Surfaces positions nearing the hard cutoff so you can decide whether to manually exit at current P&L or accept the EXPIRED outcome. "
     "End-of-table summary line counts how many positions are approaching expiry — useful at-a-glance metric for daily review.","🎯 Performance"),
    ("PERFORMANCE","BY TIME HORIZON breakdown",
     "Fourth diagnostic table in the Performance sheet (after Score Band, Quick Pick Archetype, Sector). Shows hit rate / SL rate / expired count per Horizon class. "
     "Answers: 'are SHORT TERM picks actually winning in their stated 30-day window? are LONG TERM picks really earning their 270 days of patience?' If SHORT TERM has lower hit rate than POSITIONAL despite stricter triggers, "
     "the SHORT TERM classification may need recalibration. v14.1 only — appears once data accumulates.","🎯 Performance"),
    ("PERFORMANCE","Avg / Peak Missed Runup",
     "Diagnostic for EXPIRED rows: how much max gain (max_runup_pct) the stock reached before being bucketed as EXPIRED. AVG MISSED RUNUP averages across all expired rows; PEAK MISSED RUNUP shows the worst single case. "
     "Reading guide: AVG >15% suggests targets are too far OR expiry too early — stocks rallied during tracking but failed to reach T1 within the horizon window. AVG <5% means expiry was the right call (truly sideways stocks). "
     "EXPIRED w/ ≥10% RUNUP shows count/total of expired rows that crossed 10% runup before failing — high ratio means the system identifies winners but mistimes them.","🎯 Performance"),
    # v14.4: SL/T1/T2/T3 columns in Open Positions table
    ("PERFORMANCE","SL / T1 / T2 / T3 columns",
     "v14.4: each open position row shows the recommended stop-loss and three profit targets. Format: ₹price (distance%). The distance percentage is from CURRENT price (not entry), so it updates dynamically as the tracker refreshes price each pipeline run. "
     "Example: PETRONET entered at ₹282.10 with T1 ₹310.30. After rallying to ₹283.80, T1 reads '₹310.30 (+9.3%)' — meaning we still need +9.3% more from current to hit T1. SL distance is naturally negative; target distances naturally positive while we're below them. "
     "Levels are frozen at recommendation time — re-appearances do NOT update them, preserving measurement integrity. SL renders in red text (downside risk); T1/T2/T3 render in green (upside).","🎯 Performance"),
    # v14.5: CLOSED POSITIONS section
    ("PERFORMANCE","Closed Positions section",
     "v14.5: chronological log of every closed pick (SL_HIT/T1_HIT/T2_HIT/T3_HIT/EXPIRED), sorted most-recent-first. 12 columns: Symbol, Rec Date, Time Horizon, Outcome (color-coded — green for targets, red for SL, amber for EXPIRED), Outcome Date, Days to Outcome, Entry CMP, Outcome Price, P&L %, Max Runup %, Max Drawdown %, Score. "
     "Realised P&L is computed live from entry vs outcome_price (so it reflects what actually happened, not what was projected). Max Runup % is especially revealing for SL_HIT rows: a high value (e.g. +9%) means the stock rallied significantly BEFORE hitting SL — possible SL too tight. "
     "Max Drawdown % matters for T-HIT rows: shows how much pain you'd have suffered before winning, useful for sizing future positions. Section appears between Diagnostic Breakdowns and Open Positions in the sheet — natural reading flow from aggregates to detail to current state. "
     "Footer summarises counts per outcome bucket.","🎯 Performance"),
    # v15.0: Multi-factor SL/T derivation
    ("PERFORMANCE","v15.0 SL/T derivation methodology",
     "v15.0 replaces the pre-v14.6 fixed -7%/+12.5% rule with a multi-factor formula that derives SL/T1/T2/T3 from six inputs: (1) ATR-14 daily volatility, (2) cap-category fallback (LARGE/MID/SMALL/MICRO), (3) time horizon (SHORT TERM=2.5×ATR, POSITIONAL=3.5×, LONG TERM=5×), (4) sector tier (very-high/high/neutral/low/very-low, ±0.6 to ±0.35), (5) ATR-percentile regime (current 14-day ATR vs 252-day baseline: high regime widens SL 10%, low tightens 10%), (6) volume-confirmed support floor (support1 only used when vol_ratio ≥ 1.20). "
     "SL bounded [4.5%, 12.0%]. Earnings within 5 days → SL widened 20% (infrastructure ready, data source pending). T1 = max(1.5 × SL_pct, 0.40 × CFV_upside) — guarantees R:R ≥ 1.5:1. T2 = max(2.5 × SL_pct, 0.70 × CFV_upside) ≥ T1 × 1.35. T3 = max(4.0 × SL_pct, 1.00 × CFV_upside) ≥ T2 × 1.35. "
     "Hard T3 caps prevent absurd long-horizon stretches: SHORT TERM 35%, POSITIONAL 80%, LONG TERM 200%. Grade: A- (smart heuristics, not walk-forward backtested).","🎯 Performance"),
    ("PERFORMANCE","Trailing Stop logic (v15.0)",
     "v15.0 outcome tracker now ratchets up the stop-loss as the position gains. Peak gain thresholds: ≥+5% → trailing SL moves to break-even (entry); ≥+10% → moves to entry+3% (locks small profit); ≥+15% → moves to entry+7% (locks bigger profit). "
     "Effective SL = MAX(original_sl, trailing_sl_price) — once activated, trailing SL only moves UP, never down. Zero look-ahead bias: trailing-SL update happens at END of each bar after the day's event check, so today's high cannot tighten today's stop and immediately trip it on today's low. "
     "Persisted in gold_outcomes: trailing_sl_pct (current trailing SL as %), trailing_sl_price (absolute), peak_price_seen (highest price observed). Original SL preserved in gold_recommendations.original_stop_loss for audit. SL_HIT outcomes now record the effective SL price (which may be above entry for trailing fires).","🎯 Performance"),
    ("PERFORMANCE","Sector Tier / Regime / Volume confirmation columns",
     "v15.0 audit columns logged with each recommendation. regime_at_rec: 'high' (current 14-day ATR ≥ 1.20× 252-day baseline → SL widened 10%), 'low' (≤ 0.80× → SL tightened 10%), or 'neutral'. atr_at_rec: ATR-14 as % of CMP at log time (debug/audit value). "
     "Sector tier comes from a 5-tier static map: very-high (+0.6: Realty/Sugar/Aviation/Real Estate), high (+0.3: Metals/PSU Bank/Coal/Power/etc), neutral (0: default for sectors not in map), low (-0.2: Banking/IT-Services/Auto/Chemicals), very-low (-0.35: FMCG/Pharma/Healthcare/Utilities/etc). "
     "Volume-confirmed support: support1 (20-day low from technicals) only used as SL floor when vol_ratio (today's vol / 50-day avg) ≥ 1.20, filtering out random lows with no buying conviction.","🎯 Performance"),
]

GRP_COLORS = {
    "IDENTITY":"1E293B","SCORES":"7C3AED","PRICE & MARKET":"0369A1",
    "WEEKLY CHANGE %":"0F766E","FAIR VALUE":"B45309","VALUATION":"0891B2",
    "PROFITABILITY":"059669","GROWTH":"047857","FIN HEALTH":"DC2626",
    "CAP ALLOC":"6D28D9","SHAREHOLDING":"7C3AED","QUALITY SCORES":"0D9488",
    "PIPELINE / OB":"1D4ED8","EARLY DETECTION":"B45309","TECHNICAL":"6D28D9",
    "BALANCE SHEET":"D97706","TRADE PLAN":"059669","NEWS & RISK":"475569",
    "ANALYSIS SUMMARY":"0F172A",
    # Session 19: new group for documenting the Gold-Tier filter definition
    "GOLD FILTER":"B45309",
    # v14.0: new group for the Performance sheet
    "PERFORMANCE":"B45309",
}

# Columns permanently blank — no free data source available
# These are highlighted with bold red headers so user knows at a glance
NO_FREE_SOURCE_COLS = {
    # Financial ratios needing balance sheet detail (BSE filings / paid API)
    "ND/EBITDA","Int Coverage","CCC Days","Capex / Rev %",
    # Shareholding QoQ changes (need quarterly filing history). Public Float % is derived and shown in normal colour.
    "Pro QoQ Δ","Pledge %","Pledge Direction","DII %","DII QoQ Δ",
    "FII QoQ Δ",
    # Forensic / quality scores (need multi-year filed financials)
    # Session 20: Piotroski F /9 removed — it now computes from free data
    # via FundamentalEngine.calculate_piotroski_f_score (Session 14 wire-up).
    # Typical output 4-8 of 9 on free data; no longer a paid-source column.
    "Altman Z","Beneish M","Earn Quality",
    # Intelligence / pipeline (needs company-specific filed data)
    "OB/Bill Ratio","Pipeline Vis","L1 Wins 90D","L1 Est (₹Cr)","New Mkt Entry",
    # AI-generated text fields (needs Gemini credits — separate amber set below)
    "Key Catalyst","News Sentiment","Primary Risk","SEBI Flags",
    # NOTE: NPM Q (latest)/Q-1/Q-2, Margin Expansion, CAGRs, Q3 Rev/PAT/EBITDA
    # were previously red but are now calculated via yfinance — moved to normal.
    # (Pre-v12.6 these were labelled NPM Q1/Q2/Q3.)
}
# Needs Gemini API credits — amber highlight
NEEDS_AI_CREDITS = {"View Analysis Summary"}

def _sf(val, default=0.0):
    if val is None or val == "" or str(val) in ("—", "--", "N/A"):
        return float(default)
    try:
        return float(val)
    except (ValueError, TypeError):
        return float(default)


def _f(h): return PatternFill("solid", fgColor=h)

# Thin border on all 4 sides — used for header rows and data cells
_THIN = Side(style="thin", color="FFFFFF")       # white border between same-section cols
_SECT = Side(style="medium", color="FFFFFF")      # medium white border between sections
_BDR  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BDR_SECT = Border(left=_SECT, right=_SECT, top=_THIN, bottom=_THIN)

def _border(is_section_edge=False):
    return _BDR_SECT if is_section_edge else _BDR
def _ft(bold=False,color=NAVY,size=9,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def _al(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _sv(v):
    if v is None: return None
    if isinstance(v,(str,int,float,bool)): return v
    if isinstance(v,(list,tuple)): return " | ".join(str(x) for x in v) if v else ""
    if isinstance(v,dict): return str(v)
    return str(v)
def _g(s,k,d="—"):
    v=s.get(k,d); return _sv(v) if v is not None else d

def _alt(bg):
    return {"FAC775":"FEF3C7","D1FAE5":"ECFDF5","DBEAFE":"EFF6FF",
            "FEE2E2":"FEF2F2","FEF3C7":"FFFBEB"}.get(bg,bg)

# ── Column header tooltips (module-scope so all sheets can use them) ──
_HDR_TIPS = {
    "Verdict":("✅ BUY=strong | WATCHLIST=wait | AVOID=skip","BUY: Score clears cap-tier threshold + MoS > -10%\nWATCHLIST: Score qualifies but MoS gate blocks\nAVOID: Score < 38 (universal floor)\n\nBUY thresholds: LARGE>=60 | MID>=63 | SMALL>=66 | MICRO>=70\nTech Override: MoS gate relaxes to -20% when Score>=70+ST=BUY+Stage2"),
    "Quick Pick":("⭐ Archetype label — distinct from Verdict","DEEP VALUE EARLY MOVER: MoS>25% + Score>70 + EE>=60 (top conviction)\nDEEP VALUE: MoS>25% + Score>70 (cheap vs FV)\nEARLY MOVER: EE>=70 + Score>55 (breakout signal)\nAVOID/EXIT: Score<38 OR (Score<45 AND MoS<-30%)\nWATCHLIST: fallback (no special archetype)\n\nFirst-match-wins priority order. AutoFilter to isolate top picks.\n\nWhy DEEP VALUE EARLY MOVER uses 3 factors (others use 2):\nIt's the COMBO of DEEP VALUE + EARLY MOVER — inherits both gates.\nNote EE drops 70→60 in combo: when value+quality both high, you don't\nneed an extreme momentum signal — moderate confirmation is enough.\nPure EARLY MOVER needs EE>=70 because momentum is the only thing\nspeaking for stocks that may not have great fundamentals (Score>55 floor).\n\nSource: scoring_engine._assign_quick_pick (L494-507)"),
    "Score /100":(">=70 strong | >=60 watch | <38 avoid","Fundamental*35%+Technical*30%+EarlyEntry*15%+Sentiment*10%+Safety*10%\n+MoS adj(-10 to +12)+Spike bonus(x2 capped +10)\n+Early Mover+5(if EE>=50)-Risk penalty-10\n\n>=80:Exceptional | >=70:Strong BUY | >=60:Watch | <38:Avoid"),
    "Early Entry /100":(">=50=Early Mover | >=35=Ahead of Consensus","Detects stocks 4-12 wks BEFORE institutional coverage.\nVol Surge+RSI +15 | Trend Confluence +12 | Momentum +10\n52W Breakout +10 | Deep Value+BUY +10 | Inst Footprint +10\nScore Convergence +8 | FII Accum +8 | Promoter Accum +8\n\n>=50:EARLY MOVER | >=35:AHEAD OF CONSENSUS | <35:EMERGING\nMax ~55 with free data(FII/Promoter QoQ needs paid source)"),
    "Spike Score /6":(">=2 notable | >=4 strong | 6 very rare","6 momentum triggers-how many fire simultaneously:\nT1:CMP within 3% of 52W High+vol>2x | T2:MACD+ST=BUY+vol>1.5x\nT3:ADX>25+delivery>60%+vol>1.5x | T4:RSI 45-65+vol>2x\nT5:vol>3x+delivery>60% | T6:2w_chg>3%+2w>4w+vol>1.5x\n\n0:None | 1:Weak | 2-3:Notable | 4-5:Strong | 6:Extreme\nSuppressed to 0 if pledge>20% or Beneish/Altman flags active"),
    "Storm Score /10":(">=8 Storm Safe | >=5 Moderate | <5 High Risk","Defensive quality-how safe in a crash?\nBeta<0.8+2 | D/E<0.3+2 | FCF positive+2\nDiv yield>2%+1 | Rev growth>10%+1 | Margin Expansion+1\nPromoter QoQ up+1 | FII buying 3Q+1\n\n>=8:STORM SAFE | 5-7:MODERATE | <5:HIGH RISK"),
    "CMP (\u20b9)":("Current market price","Compare with CFV. CMP < CFV = undervalued (buying opportunity)"),
    "Day Chg %":("Today's move | >+3% strong | <-3% weak",">5% may trigger circuit rules"),
    "52W High (\u20b9)":("52-wk peak-CMP near here = breakout","CMP near 52W High = breakout territory (T1 spike fires)"),
    "52W Low (\u20b9)":("52-wk floor-value or continued decline","Verify fundamentals before buying near 52W Low"),
    "Vol Spike (\xd750D)":(">=2x unusual | >=3x institutional","1x:Normal | 1.5-2x:Above avg | 2-3x:Unusual | >3x:Major event\nHigh vol+rising=buying | High vol+falling=selling"),
    "Delivery %":(">=60% institutional | <40% speculative",">=70%:Strong institutional conviction\n40-70%:Mixed | <40%:Mostly speculative-caution"),
    "Beta":("<0.8 defensive | >1.2 volatile","Beta<0.8:Less volatile-good in downturns(+2 Storm)"),
    "Chg% [2-Weekly]":("2-wk return | >3% strong momentum",">2%:EE Momentum signal fires. 2W>4W=accelerating(bullish)"),
    "Chg% [4-Weekly]":("4-wk return | 2W>4W = acceleration","2W>4W=accelerating | 2W<4W=decelerating"),
    "Chg% [6-Weekly]":("6-wk trend","Positive+rising across 2W/4W/6W/8W=sustained uptrend"),
    "Chg% [8-Weekly]":("8-wk return | best for trend direction","Consistently positive=confirmed uptrend"),
    "CFV (\u20b9)":("Composite Fair Value(7 models weighted)","M1 DCF 30%|M2 Graham 15%|M3 PE 20%|M4 PB 15%\nM5 EV/EBITDA 10%|M6 DDM 5%|M7 PEG 5%\nCMP<CFV=undervalued | CMP>CFV=overvalued"),
    "FV Low (\u20b9)":("Conservative FV = CFV x0.85","CMP below FV Low=very deeply undervalued"),
    "FV High (\u20b9)":("Optimistic FV = CFV x1.15","CMP above FV High=significantly overvalued"),
    "MoS %":(">25% strong buy | <-15% overvalued","(CFV-CMP)/CMP*100\n>40%:Exceptional(+12) | >25%:Strong(+8) | 10-25%:Adequate(+4)\n-15 to -30%:Overvalued(-5) | <-30%:Significant premium(-10)"),
    "MoS Label":("Valuation summary (* = capped, † = thin-FV)","EXCEPTIONAL>40%|STRONG>25%|ADEQUATE>10%|THIN 0-10%|PREMIUM<0%; *=CFV hit 3× cap; †=<3 models fired"),    "P/E TTM":("<20 cheap | 20-40 fair | >40 expensive","Score: <=20=+12 | <=40=+7 | >60=-8"),
    "Earn Yield %":(">6% undervalued vs bonds","EPS/CMP*100. >6%:Cheap. Compare to 10Y bond yield"),
    "P/CF":("<15 value | >25 expensive","More reliable than P/E(cash harder to fake)"),
    "PEG Ratio":("<1 undervalued | >2 expensive","P/E / Growth. <1:Undervalued(Peter Lynch favourite)"),
    "P/B":("<2 value | >5 expensive","<1:Below asset value | >5:Only justified by very high ROE"),
    "P/S":("<3 cheap | >10 expensive","Useful when P/E unavailable"),
    "EV/EBITDA":("<12 value | >20 expensive","IT=20|Pharma=18|FMCG=30|Banks=12|Metals=8"),
    "ROE %":(">20% excellent | <10% weak","Score: >20%=+12 | >10%=+6 | <5%=-5\nHigh ROE+Low D/E=ideal combination"),
    "ROCE %":(">15% good capital allocation","Return on Capital Employed. ROCE>cost of capital=value-creating"),
    "ROA %":(">10% efficient | <5% poor","Low for Banks/Utilities is normal"),
    "Gross Mgn %":(">40% strong moat | >20% decent","Score: >40%=+8 | >20%=+4"),
    "EBITDA Mgn %":(">25% excellent | >15% good",">30%:Excellent | >20%:Good | <10%:Tight"),
    "NPM %":(">15% excellent | <5% thin","Score: >15%=+8 | >5%=+4 | <0%=-8"),
    "NPM Q (latest) %":("Most recent quarter margin vs TTM","Q(latest)>NPM(TTM):margins accelerating"),
    "NPM Q-1 %":("Previous quarter margin","Track Q-2 → Q-1 → Q(latest) trend for margin direction"),
    "NPM Q-2 %":("Quarter from 2 reports ago — rising = Margin Expansion","Rising Q-2 → Q-1 → Q(latest) triggers Margin Expansion=YES"),
    "Margin Expansion":("YES=3 consecutive qtrs of rising NPM","Score: Fundamental+5 | Safety+3 | Storm+1. YES is rare(~10%)"),
    "Rev CAGR 1Y %":(">20% high growth | >10% good","1Y>3Y CAGR=growth accelerating"),
    "Rev CAGR 3Y %":(">15% strong | >8% decent","Score: >15%=+5 | >8%=+3. More reliable than 1Y."),
    "PAT CAGR 1Y %":(">20% strong earnings momentum","1Y>3Y=accelerating profitability"),
    "PAT CAGR 3Y %":(">20% compounder | >10% good","Score: >20%=+8 | >10%=+4. PAT CAGR>Rev CAGR=improving margins"),
    "EBITDA CAGR 1Y %":(">15% strong operating growth","Score: >15%=+4 | >8%=+2"),
    "Rev YoY %":(">10% growing | <0% declining","Score: >15%=+5 | >8%=+3 | <-5%=-4"),
    "PAT YoY %":(">20% strong | >10% good","Score: >20%=+8 | >10%=+4 | >0%=+2 | <-10%=-7"),
    "Q3 Rev (\u20b9Cr)":("Latest quarter revenue","Rising QoQ=business growing"),
    "Q3 PAT (\u20b9Cr)":("Latest quarter net profit","Positive and growing=healthy earnings"),
    "Q3 EBITDA (\u20b9Cr)":("Latest quarter operating profit","Q3 EBITDA Margin=Q3 EBITDA/Q3 Rev vs TTM"),
    "D/E Ratio":("<0.3 excellent | >2 risky | >3 danger","<0.3:+8 Fundamental+2 Storm | 0.3-1:+4 | >2:-10 | >3:BS ALERT\nException:Banks/NBFCs naturally high D/E"),
    "ND/EBITDA":("<2 safe | >4 risky","Net Debt/EBITDA=years to repay. <0:Net cash | >4:High leverage"),
    "Int Coverage":(">5 safe | <2 danger","EBIT/Interest. >10:Very safe | <2:Danger | <1:Critical"),
    "Current Ratio":(">2 healthy | 1-2 adequate | <1 risky","Score: >2=+6 | >1.5=+3 | <1=-7"),
    "Quick Ratio":(">1 safe | <0.5 risky","(Current Assets-Inventory)/Current Liabilities"),
    "Cash (\u20b9Cr)":("Higher=stronger safety net","Cash>Total Debt=NET CASH COMPANY"),
    "Total Debt (\u20b9Cr)":("Lower=better | 0=ideal","Compare with Cash and EBITDA"),
    "FCF (\u20b9Cr)":(">0 cash generator | <0 cash consuming","Score: >0=+2 Storm+3 Safety | <0 cash consuming"),
    "FCF Yield %":(">6% undervalued | >3% fair","Score: >6%=+6 | >3%=+3 | <0%=-5"),
    "CCC Days":("Lower/negative=more efficient","Negative CCC=collects cash before paying suppliers"),
    "Div Yield %":(">2% good income | >4% check sustainability",">2%:+1 Storm Score"),
    "Payout Ratio %":("40-60% balanced | >80% unsustainable","30-60%:Balanced | >80%:Check FCF coverage"),
    "Capex / Rev %":("<5% asset-light | >15% capital-heavy","<3%:Asset-light=high FCF"),
    "Promoter %":(">50% aligned | <20% concern","Score: >50%=+5 | >35%=+2 | <20%=-3"),
    "Pro QoQ \u0394":(">0.3% buying signal | negative=selling","+1 Storm Score if >0.3%"),
    "Pledge %":("0% ideal | >10% watch | >20% RED FLAG","Safety:>20%=-15 | >20% suppresses ALL Spike signals"),
    "Pledge Direction":("FALLING=positive | RISING=risk","FALLING:Repaying loans | RISING:More pledging=risk"),
    "FII %":(">15% institutional backed",">25%:High global interest | Rising FII=strong signal"),
    "FII QoQ \u0394":(">1% accumulation | <-1% selling","+8 EE if >1% | +1 Storm if >0.3%. Often blank(free sources)"),
    "DII %":(">10% domestic confidence","Rising DII+FII=dual institutional accumulation=bullish"),
    "DII QoQ \u0394":(">0.5% domestic accumulation","Rising=domestic MFs accumulating"),
    "Public Float %":(">50% good liquidity | <20% manipulation risk","<20%:Volatile, easier to manipulate"),
    "Piotroski F /9":(">=7 strong | <=3 weak","9 criteria:Profitability(3)+Leverage(2)+Efficiency(4)\n8-9:Excellent | 6-7:Good | <=3:Avoid"),
    "Altman Z":(">2.99 safe | <1.81 distress zone","<1.81:Triggers anti-trigger guard(Spike suppressed)"),
    "Beneish M":("<-2.22 honest | >-2.22 possible manipulation","v12.9 real 8-var formula (DSRI/GMI/AQI/SGI/DEPI/SGAI/TATA/LVGI) | >-2.22 triggers anti-trigger guard (Spike suppressed)"),
    "Earn Quality":("HIGH=cash-backed earnings","CFO/PAT ratio (PAT annualized in v12.9) | HIGH≥0.8 | MODERATE 0.5-0.8 | LOW<0.5"),
    "OB/Bill Ratio":(">1 strong pipeline | >3 excellent visibility",">3:3+ year revenue visibility. For infra/defence/engineering"),
    "Pipeline Vis":("HIGH=strong revenue visibility","HIGH:Strong order book or recurring revenue"),
    "L1 Wins 90D":("Recent govt contract wins",">3:Active and winning bidder"),
    "L1 Est (\u20b9Cr)":("Estimated govt contract value","Higher=more near-term revenue locked in"),
    "New Mkt Entry":("YES=new revenue stream potential","YES:New geography or product launch(growth catalyst)"),
    "Early Signals":("Signals fired today-more=higher conviction","EE+spike signals: VOL SURGE+RSI|TREND CONFLUENCE\nTECHNICAL BREAKOUT|INSTITUTIONAL FOOTPRINT|52W BREAKOUT"),
    "Sector Stage":("Stage 2=best entry | Stage 4=avoid/exit","STAGE1 EARLY ACCUM:Smart money entering(good entry)\nSTAGE2 CONFIRMED UPTREND:All signals aligned(BEST ENTRY)\nSTAGE3 MOMENTUM PEAK:Overbought(caution)\nSTAGE4 DISTRIBUTION:Smart money exiting(avoid)"),
    "Smart Money":("ACCUMULATION=institutional buying","INST ACCUMULATION|INSIDER BUYING|FII INCREASING\nPROMOTER BUYING|HIGH DELIVERY BUYING|RSI ACCUM ZONE|NEUTRAL"),
    "SMA 200":("CMP>SMA200=bull trend confirmed","200-Day SMA. Golden Cross=major buy | Death Cross=major sell"),
    "Supertrend":("BUY=uptrend | SELL=downtrend | NEUTRAL=sideways","BUY:Price>SMA20+0.5xATR14 | SELL:Price<SMA20-0.5xATR14\nUsed in T2 Spike, Trend Confluence EE, Tech Override"),
    "ADX":(">25 strong trend | <20 weak/sideways","Measures TREND STRENGTH(not direction)\n>25:Strong(+5) | 20-25:Moderate(+2) | <20:Weak"),
    "RSI (14)":("45-65 sweet spot | >70 overbought | <30 oversold",">70:Overbought | 50-70:Bullish(+4 to +8)\n45-65:SWEET SPOT for entry | <30:Oversold"),
    "MACD Signal":("BUY=bullish crossover | SELL=bearish","BUY:+6 Technical | SELL:-6 Technical"),
    "Stoch %K":("20-40 accumulation zone | >80 overbought","20-40:Accum zone(+5) | >80:Overbought(-3)"),
    "MFI":(">60 money inflow | <30 outflow",">60:+4 Technical | <30:-3 Technical"),
    "OBV Signal":("RISING=accumulation | FALLING=distribution","RISING:+4 Technical | FALLING:-4 Technical"),
    "Above VWAP":("YES=institutional support | NO=weak","YES:+4 Technical | NO:-2 Technical"),
    "Chart Pattern":("BULLISH REVERSAL=buy signal","BULLISH REVERSAL|BEARISH CANDLE|DOJI(indecision)"),
    "BS Health Flag":("HEALTHY=safe | WATCH=monitor | ALERT=danger","HEALTHY:No red flags\nWATCH:One concern(D/E>2 or low liq or neg FCF)\nALERT:Serious(pledge>20% or D/E>3 or leveraged+neg FCF)\nALERT:suppresses entry signals+Safety-15"),
    "BS Health Note":("Explains the health flag","Examples:NET CASH COMPANY|HIGH D/E 2.5x|NEGATIVE FCF|HIGH PLEDGE"),
    "Entry Range (\u20b9)":("Ideal buy zone = CMP +/- 0.5xATR","Avoid chasing if CMP moves significantly above upper bound"),
    "Stop Loss (\u20b9)":("Exit if CMP closes below this level","Never risk >2-3% of portfolio per trade"),
    "Target 1 (\u20b9)":("First target=Resistance 1","Book 30-50% of position here. R:R should be >1:2"),
    "Target 2 (\u20b9)":("Second target=Resistance 2","Hold remainder after Target 1"),
    "Target 3 (\u20b9)":("Final target=Fair Value(CFV)","High MoS stocks can give 20-50% upside"),
    "Time Horizon":("How long to hold","SHORT TERM:2-4wks(BUY+Spike>=2)\nPOSITIONAL:1-3mo(Score>=68+ST=BUY)\nLONG TERM:3-12mo(Score>=72+no spike)"),
    "Risk Level":("LOW=safest | VERY HIGH=speculative only","LOW:High score+low beta+low D/E+no pledge\nMEDIUM:Acceptable | HIGH:Small/micro | VERY HIGH:Speculative"),
    "Exchange":("DUAL_LISTED=best liquidity","DUAL_LISTED:NSE+BSE(best) | NSE_ONLY:Good\nBSE_ONLY:Lower liq | BSE_SME:Very low-high impact cost"),
    "Cap Category":("LARGE=safest | MICRO=speculative","BUY thresholds: LARGE>=60|MID>=63|SMALL>=66|MICRO>=70"),
    "Support 1 (\u20b9)":("Nearest support=buy zone floor","20-day rolling low. Breach=bearish. Used for Stop Loss."),
    "Support 2 (\u20b9)":("Deeper support level","40-day rolling low. Next level if Support 1 breaks."),
    "Resist 1 (\u20b9)":("First resistance=Target 1","20-day rolling high. Breakout with volume=bullish."),
    "Resist 2 (\u20b9)":("Stronger resistance=Target 2","40-day rolling high. Used as Target 2."),
    "Key Catalyst":("Primary near-term growth driver","Product launch,order win,policy tailwind,expansion"),
    "News Sentiment":("POSITIVE=tailwind | NEGATIVE=headwind","POSITIVE:Favourable | NEUTRAL:No news | NEGATIVE:Headwinds"),
    "Primary Risk":("Biggest downside risk","Always read before investing"),
    "SEBI Flags":("NONE=clean | Any flag=investigate first","Any flag=investigate before buying"),
    "View Analysis Summary":("Gemini AI investor narrative(150-250 words)","Business quality,ratios,risks,catalysts,verdict rationale.\nGenerated fresh each trading day."),
}


def _patch_tooltip_vml(xlsx_path):
    """Session 27+28+29+v10.12: Post-process .xlsx to fix tooltip VML shapes.

    Four things this function does that openpyxl won't:

    1. Box dimensions, per-tooltip (v10.12 dynamic sizing).
       openpyxl writes VML shapes with hardcoded 144×79px regardless of what
       Comment.width/height are set to. Pre-v10.12 we rewrote every shape to
       a fixed 420×380 px — which caused short tooltips (e.g., Stop Loss:
       2 short lines) to show with 60-70% empty space below. Now we parse
       xl/comments1.xml to get the text of each comment, map it to its shape
       via (row, col) anchor, and size the box to fit the actual content:
         width  = 420px  (standard)
         height = clamp(17 * line_count + 36, min=85, max=380)
       Line heuristic: tooltip text uses \\n separators; each line ~17px high
       with ~36px chrome (title bar + padding).

    2. Anchor direction for right-side columns (Session 29). openpyxl sets
       `margin-left:59.25pt` on every shape, which pushes the comment box
       to the RIGHT of the anchor cell. On the rightmost few columns of a
       sheet, the box gets clipped or hidden off-screen. We scan each
       drawing's shapes to find the rightmost anchor column, then for any
       shape whose column is in the right portion of the sheet, flip the
       margin to a negative value so the box opens to the LEFT of the cell.
       Threshold: any column at >= 70% of the rightmost tooltipped column
       index on that sheet gets flipped.

    3. Nothing for scrolling. VML comments do not support internal
       scrollbars (that's a threaded-comments feature). We size the box
       for the known-tallest tooltip content plus padding.
    """
    import re
    import shutil
    import zipfile
    import os as _os
    import tempfile
    import xml.etree.ElementTree as ET

    tmpdir = tempfile.mkdtemp(prefix="xlsx_patch_")
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            zin.extractall(tmpdir)

        vml_dir = _os.path.join(tmpdir, "xl", "drawings")
        if not _os.path.isdir(vml_dir):
            return  # no comments; nothing to patch

        # ---- v10.12: parse comments*.xml to get per-cell line counts --------
        # Build dict: (sheet_idx, row, col) → line_count
        # VML commentsdrawingN.vml corresponds to xl/comments{N}.xml
        def _excel_ref_to_rc(ref: str):
            """Convert 'B7' → (row=7, col=2). Col header letters 1-based."""
            m = re.match(r'^([A-Z]+)(\d+)$', ref.strip())
            if not m: return None, None
            letters, rownum = m.group(1), int(m.group(2))
            col = 0
            for ch in letters:
                col = col * 26 + (ord(ch) - ord('A') + 1)
            return rownum, col

        def _extract_text(elem):
            """Concatenate all <t> text under a <comment> element."""
            parts = []
            # namespace-agnostic search
            for sub in elem.iter():
                if sub.tag.endswith('}t') or sub.tag == 't':
                    if sub.text:
                        parts.append(sub.text)
            return ''.join(parts)

        # comments_dir lookup: map each comments{N}.xml → {(row,col): line_count}
        comments_dir_map = {}   # key: 'N' (e.g., '1', '2'), val: dict[(r,c)] → lc
        # openpyxl writes to xl/comments/commentN.xml (newer) or xl/commentsN.xml
        # (older path). Scan both locations.
        comments_search_dirs = [
            _os.path.join(tmpdir, "xl"),
            _os.path.join(tmpdir, "xl", "comments"),
        ]
        for comments_root in comments_search_dirs:
            if not _os.path.isdir(comments_root): continue
            for fn in _os.listdir(comments_root):
                # Match: comments1.xml, comment1.xml, Comments1.xml, Comment1.xml
                m_id = re.match(r'comments?(\d+)\.xml$', fn, re.IGNORECASE)
                if not m_id: continue
                cid = m_id.group(1)
                line_map = {}
                try:
                    tree = ET.parse(_os.path.join(comments_root, fn))
                    # Walk through <comment ref="X7"> elements
                    for comment_el in tree.iter():
                        tag = comment_el.tag
                        local_tag = tag.split('}', 1)[1] if '}' in tag else tag
                        if local_tag != 'comment':
                            continue
                        ref = comment_el.get('ref')
                        if not ref: continue
                        r, c = _excel_ref_to_rc(ref)
                        if r is None: continue
                        text = _extract_text(comment_el)
                        # openpyxl joins lines with \n internally; in XML it
                        # appears as literal newlines within <t xml:space="preserve">
                        line_count = text.count('\n') + 1 if text else 1
                        line_map[(r, c)] = line_count
                except Exception:
                    pass
                if line_map:
                    comments_dir_map[cid] = line_map

        # ---- v10.12 dynamic dimensions + Session 28 bounds ------------------
        BOX_W_PX = 420
        LINE_PX  = 17
        CHROME_PX = 36    # title bar + vertical padding
        MIN_H_PX = 85     # short 2-line tooltips still need title bar room
        MAX_H_PX = 380    # safety cap for very long tooltips

        def _height_for_lines(lc: int) -> int:
            return max(MIN_H_PX, min(LINE_PX * lc + CHROME_PX, MAX_H_PX))

        # ---- Session 29 per-shape margin flip ------------------------------
        # Default openpyxl emits: margin-left:59.25pt (positive → box opens
        # to the right of the anchor cell). For cells on the right edge of
        # a sheet, we flip to a negative value so Excel draws the box to the
        # left of the cell instead. 420px box + a small buffer → flip value
        # pushes the box left-of-cell with a 15pt gap.
        LEFT_FLIP_MARGIN_PT = -(BOX_W_PX * 0.75 + 15)  # ≈ -330pt (420px ≈ 315pt)
        # "Right portion" threshold: any column ≥ 70% of max-tooltipped-col
        # gets flipped to the left side.
        FLIP_THRESHOLD = 0.70

        # Regex components
        style_re = re.compile(
            r'(<[^:]*:shape\b[^>]*\bstyle=")([^"]+)("[^>]*>.*?</[^:]*:shape>)',
            re.DOTALL
        )
        dim_re = re.compile(
            r'width\s*:\s*\d+px\s*;\s*height\s*:\s*\d+px',
            re.IGNORECASE
        )
        row_re = re.compile(r'<[^:]*:Row>\s*(\d+)\s*</[^:]*:Row>')
        col_re = re.compile(r'<[^:]*:Column>\s*(\d+)\s*</[^:]*:Column>')
        margin_left_re = re.compile(r'margin-left\s*:\s*[^;"]+')

        patched_count = 0
        flipped_count = 0
        for fn in _os.listdir(vml_dir):
            if not fn.lower().startswith("commentsdrawing") or not fn.lower().endswith(".vml"):
                continue

            # Extract drawing number and match to comments{N}.xml line map
            m_id = re.match(r'commentsDrawing(\d+)\.vml', fn, re.IGNORECASE)
            cid = m_id.group(1) if m_id else None
            line_map = comments_dir_map.get(cid, {}) if cid else {}

            fpath = _os.path.join(vml_dir, fn)
            with open(fpath, "r", encoding="utf-8") as f:
                content = f.read()

            # Pass 1: find max anchor column across all shapes in this file
            max_col = 0
            for shape_m in style_re.finditer(content):
                col_m = col_re.search(shape_m.group(0))
                if col_m:
                    max_col = max(max_col, int(col_m.group(1)))
            flip_col_threshold = max_col * FLIP_THRESHOLD if max_col > 0 else float("inf")

            # Pass 2: rewrite each shape's style
            def _fix_shape(m):
                nonlocal patched_count, flipped_count
                full_shape = m.group(0)
                style = m.group(2)

                # v10.12: look up this shape's (row, col) and comments line count
                row_m = row_re.search(full_shape)
                col_m = col_re.search(full_shape)
                # VML Row/Column are 0-indexed; openpyxl writes 0-based here.
                # Spreadsheet refs use 1-based rows and columns. Convert:
                if row_m and col_m:
                    vml_row_0 = int(row_m.group(1))
                    vml_col_0 = int(col_m.group(1))
                    row_1based = vml_row_0 + 1
                    col_1based = vml_col_0 + 1
                    line_count = line_map.get((row_1based, col_1based), 0)
                else:
                    line_count = 0

                # Compute dynamic height; fallback to MAX_H_PX if we couldn't
                # locate the text (preserves previous behavior for edge cases)
                if line_count > 0:
                    box_h = _height_for_lines(line_count)
                else:
                    box_h = MAX_H_PX
                dim_replacement = f"width:{BOX_W_PX}px;height:{box_h}px"

                # Fix dimensions
                new_style, n_dim = dim_re.subn(dim_replacement, style)
                if n_dim > 0:
                    patched_count += n_dim
                # Decide if margin should be flipped
                if col_m:
                    col_idx = int(col_m.group(1))
                    if col_idx >= flip_col_threshold and col_idx > 0:
                        new_style = margin_left_re.sub(
                            f"margin-left:{LEFT_FLIP_MARGIN_PT:.2f}pt",
                            new_style
                        )
                        flipped_count += 1
                return m.group(1) + new_style + m.group(3)

            new_content = style_re.sub(_fix_shape, content)
            if new_content != content:
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(new_content)

        # Rewrite the .xlsx from the patched directory
        backup = xlsx_path + ".orig"
        shutil.move(xlsx_path, backup)
        try:
            with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for root, _, files in _os.walk(tmpdir):
                    for fn in files:
                        full = _os.path.join(root, fn)
                        rel = _os.path.relpath(full, tmpdir)
                        # zip paths use forward slashes
                        zout.write(full, rel.replace(_os.sep, "/"))
            _os.remove(backup)
        except Exception:
            # Restore original on any error
            if _os.path.exists(backup):
                shutil.move(backup, xlsx_path)
            raise
    finally:
        import shutil as _sh
        _sh.rmtree(tmpdir, ignore_errors=True)


class ExcelGeneratorV6:
    REQUIRED_COLS = {
        "early_entry_score":0,"composite_score":0,"spike_count":0,"storm_score":0,
        "mos_pct":0.0,"upside":0.0,"verdict":"WATCHLIST","spike_suppressed":False,
        "symbol":"","close":0.0,"company_name":"","sector":"General",
        "exchange_tag":"NSE","Analysis_Summary_Block_H":"—",
        "2w_chg":0,"4w_chg":0,"6w_chg":0,"8w_chg":0,"cfv":0.0,
        "entry_range":"—","stop_loss":"—","t1":0,"t2":0,"t3":0,
        "horizon":"POSITIONAL","risk_level":"MEDIUM","mos_label":"—",
        "smart_money_signals":"","rotation_stage":"NEUTRAL","vol_ratio":1.0,
        "bs_status":"HEALTHY","bs_flags":"—",
    }

    def __init__(self,data,date_str,run_time=None,prev_scores=None,gap_days=0):
        self.df=pd.DataFrame(data) if data else pd.DataFrame()
        self.date_str=date_str
        self.gap_days=int(gap_days or 0)  # trading days missed since last run
        try: self.dlbl=datetime.strptime(date_str,"%Y%m%d").strftime("%-d %b %Y")
        except: self.dlbl=date_str
        # Actual pipeline run time in IST — used in Alert Log and Delivery Preview
        if run_time:
            self.run_time = run_time
        else:
            try:
                import pytz as _ptz2
                _ist2 = _ptz2.timezone("Asia/Kolkata")
                self.run_time = datetime.now(_ist2).strftime("%H:%M IST")
            except Exception:
                self.run_time = "—"
        # Previous day scores for Score Δ computation in Alert Log
        # Suppress if gap > 5 trading days — deltas are meaningless after a long gap
        self.prev_scores = {} if self.gap_days > 5 else (prev_scores or {})
        for col,dflt in self.REQUIRED_COLS.items():
            if col not in self.df.columns: self.df[col]=dflt

        # ── NEUTRAL filter: REMOVED in v12.0 ────────────────────────────────
        # Previously this block dropped NEUTRAL-verdict stocks unless they met
        # an "exceptional" bar (ROE>20% AND PE<30 AND MoS>10% AND ts>62). This
        # caused the dashboard to silently shrink below 100 whenever Gemini
        # quota was exhausted: stocks that didn't get an AI card stayed at the
        # default NEUTRAL label and got filtered out here, even though Stage 3
        # had already selected them as worth analysing.
        #
        # Stage 3 (priority_ranker.get_top_100_candidates) is the single
        # authoritative quality gate. If a stock makes it into final_100_list,
        # it belongs in the dashboard regardless of verdict tier — including
        # NEUTRAL. Sorting by VERDICT_ORDER in priority_ranker already places
        # BUY first and AVOID last, so NEUTRAL rows surface in their natural
        # position without crowding out high-conviction picks.

    @staticmethod
    def _safe_val(v): return _sv(v)

    def generate_excel_reports(self):
        """Generates a single Excel file with all 6 sheets.
        Sheet 2 (Gold – Early Movers) is embedded inside the Full Dashboard file.
        No separate Gold file is produced — avoids duplication.
        Returns (master_file, None) — None is filtered out by master_funnel
        so the xlsx appears only once in the email attachment list.
        """
        master = self._full()
        return master, None

    def _full(self):
        wb=Workbook(); wb.active.title="📊 Full Dashboard"
        # Session 16: build reference sheet FIRST so header hyperlinks have
        # row anchors to target. Rearranged at the end so it sits after all
        # the primary data sheets in the tab order.
        self._tt_ref_anchors = _tt_build_ref(wb)
        self._full_sheet(wb.active)
        self._gold_sheet(wb)
        self._trade_summary(wb)
        self._alert_log(wb)
        self._delivery_preview(wb)
        self._performance_sheet(wb)   # v14.0: Gold-pick outcome dashboard
        self._glossary(wb)
        # Move Tooltip Reference to the end so tab order is:
        # Full Dashboard → Gold → Trade Summary → Alert Log → Delivery → Performance → Glossary → Reference
        _ref = wb["📖 Tooltip Reference"]
        wb._sheets.remove(_ref); wb._sheets.append(_ref)
        for ws in wb.worksheets: ws.sheet_view.showGridLines=False
        fn=f"NSE_BSE_Full_Dashboard_{self.date_str}.xlsx"; wb.save(fn)
        # Session 27 fix: openpyxl does NOT persist Comment.width/height to the
        # saved VML drawing — it always writes 144×79 (its internal defaults)
        # regardless of what _comment() set. This makes every tooltip box cramped:
        # 27-line tooltips (Score/EE/Verdict) overflow invisibly into an 79pt-tall
        # container so users see only the first 3-4 lines. Post-process the .xlsx
        # VML files to:  (a) set generous dimensions (380×320 px),
        #                (b) enable internal wrap so long lines don't clip horizontally,
        #                (c) add auto-scroll so content remains reachable.
        _patch_tooltip_vml(fn)
        return fn

    def _gold_file(self):
        """Kept for backward compatibility — returns master file path."""
        return f"NSE_BSE_Full_Dashboard_{self.date_str}.xlsx"


    def _apply_col_tips(self, ws, header_row, col_headers):
        """Add polished hover tooltips + visible ⓘ cue + ref-sheet hyperlink
        to every header cell that has an entry in tooltip_formatter.TIPS.

        Session 16: delegates to reporting.tooltip_formatter for the full
        Tier 1+2+3 treatment. Legacy _HDR_TIPS (still present for backward
        compatibility with any external reader) is no longer used by this
        method — tooltip_formatter.TIPS is the active source of truth and
        has richer, curated content for 150+ headers.
        """
        _tt_apply(
            ws, header_row, col_headers,
            add_cue=True,
            ref_anchors=getattr(self, "_tt_ref_anchors", None),
        )

    def _full_sheet(self,ws):
        N=len(FULL_COLS)
        # R1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
        c=ws.cell(1,1,f"NSE / BSE STOCK ANALYSER  ·  FULL RESEARCH DASHBOARD  ·  v6.2  ·  {self.dlbl}")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,13); c.alignment=_al()
        ws.row_dimensions[1].height=34
        # R2
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=N)
        c2=ws.cell(2,1,"AutoFilter (row 4): Exchange · Cap Category · Sector · Verdict · MoS Label · BS Flag · Risk · Storm · Sector Stage · Weekly Change   |  Last column = 'View Analysis Summary' — scroll right to see full AI reasoning with recent company facts   |  GOLD=Early Mover · GREEN=Deep Value · BLUE=Buy · AMBER=Watch · RED=Avoid   |  RED column header = No free data source (requires paid API / BSE filings). AMBER header = Needs Gemini API credits. Normal header = calculated from free sources (yfinance / NSE).")
        c2.fill=_f(LG); c2.font=_ft(False,"475569",8,True); c2.alignment=_al("left","center")
        ws.row_dimensions[2].height=16
        # R3 groups
        ws.row_dimensions[3].height=20
        for sc,nm,col,sp in FULL_GROUPS:
            ec=sc+sp-1
            if sp>1: ws.merge_cells(start_row=3,start_column=sc,end_row=3,end_column=ec)
            c=ws.cell(3,sc,nm); c.fill=_f(col); c.font=_ft(True,WHITE,8); c.alignment=_al()
            c.border=_border(is_section_edge=True)
        # Session 17: section-header tooltips (IDENTITY, SCORES, FAIR VALUE, …)
        _tt_apply_groups(ws, 3, FULL_GROUPS)
        # R4 headers
        # Build col→group_color map so each header cell matches its section
        _col_color = {}
        for _sc,_nm,_col,_sp in FULL_GROUPS:
            for _ci in range(_sc, _sc+_sp):
                _col_color[_ci] = _col

        ws.row_dimensions[4].height=40

        # v12.4: Dynamically detect which NO_FREE_SOURCE_COLS actually have
        # populated data in this run. A column is considered "covered" only
        # when at least 30 % of rows carry real data — a single fluke value
        # used to be enough to demote (v10.4 bug), which hid sparse columns
        # like Pro QoQ Δ (2/99 populated) and FII QoQ Δ (22/99) behind a
        # normal-coloured header.
        _stks_preview = self.df.to_dict("records")
        _row_total    = max(1, len(_stks_preview))
        _COVERAGE_MIN = 0.30   # ≥30 % of rows must carry real data
        _header_has_data = {}   # header_name → True if column meets coverage
        for (_h, _w, _key) in FULL_COLS:
            if _h not in NO_FREE_SOURCE_COLS:
                continue
            _real_count = 0
            for _stk in _stks_preview:
                _v = _stk.get(_key)
                if _v is None:
                    continue
                if _v in ("", "—", "--", "N/A", "STABLE"):
                    continue
                if _v in (0, 0.0, "0", "0.0"):
                    continue
                _real_count += 1
            _header_has_data[_h] = (_real_count / _row_total) >= _COVERAGE_MIN

        for i,(h,w,_) in enumerate(FULL_COLS,1):
            ws.column_dimensions[get_column_letter(i)].width=w
            hdr_bg = _col_color.get(i, NAVY)
            if h in NO_FREE_SOURCE_COLS and not _header_has_data.get(h, False):
                # Bold red — column has no real data in this run
                c=ws.cell(4,i,h); c.fill=_f("991B1B"); c.font=_ft(True,"FEE2E2",8)
                c.alignment=_al("center","center",True); c.border=_border()
            elif h in NEEDS_AI_CREDITS:
                # Amber — populated only when Gemini API credits are loaded
                c=ws.cell(4,i,h); c.fill=_f("92400E"); c.font=_ft(True,"FEF3C7",8)
                c.alignment=_al("center","center",True); c.border=_border()
            else:
                # Normal section color — includes previously-red cols that NOW have data
                c=ws.cell(4,i,h); c.fill=_f(hdr_bg); c.font=_ft(True,WHITE,8)
                c.alignment=_al("center","center",True); c.border=_border()
        # Session 16: polished tooltips + ⓘ cue + reference-sheet hyperlinks
        # for all headers. Replaces the legacy 💡-prefixed 300×180 yellow note
        # with a structured QUICK READ / DETAIL layout and a Tooltip Reference
        # tab for headers that want full-colour cards.
        self._apply_col_tips(ws, 4, [h for h,w,_ in FULL_COLS])
        ws.freeze_panes="A5"
        ws.auto_filter.ref=f"A4:{get_column_letter(N)}4"
        # Data
        # FV model keys that show "—" when value is 0 (model not applicable)
        FV_MODEL_KEYS = {"M1_DCF","M2_Graham","M3_PE","M4_PB","M5_EV","M6_DDM","M7_PEG",
                         "cfv","cfv_low","cfv_high"}

        stks=self.df.to_dict("records")
        for ri,stk in enumerate(stks):
            rn=ri+5; ws.row_dimensions[rn].height=20
            verd=str(stk.get("verdict","WATCHLIST"))
            st=VERDICT_STYLES.get(verd,_DS)
            bg,tx=st["bg"],st["text"]
            ubg=bg if ri%2==0 else _alt(bg)
            # v13.x: Quick Pick gets its OWN color (not the row's verdict color)
            # so the archetype label (DEEP VALUE / EARLY MOVER / AVOID / EXIT / etc.)
            # visually pops at a glance. Falls back to row's color if label
            # is unknown or missing — defensive against future label additions.
            qp_label=str(stk.get("label","WATCHLIST"))
            qp_st=VERDICT_STYLES.get(qp_label,st)            # default: row color
            qp_bg,qp_tx=qp_st["bg"],qp_st["text"]
            qp_ubg=qp_bg if ri%2==0 else _alt(qp_bg)
            # v13.x fix: when CFV could not be computed (no models fired,
            # typical for ETFs / index funds / very thin-data instruments),
            # the FV engine returns mos_pct=-100 as a math-only result of
            # (0-CMP)/CMP*100. Without this guard the cell shows "-100%"
            # alongside an "SIGNIFICANT PREMIUM" label — misleading the
            # reader to think the stock is 100% overvalued, when the truth
            # is "we can't value this with our model set". Mirror the CFV
            # display rule and render both fields as "—" in that case.
            # Internal dict values (stock["mos_pct"], stock["mos_label"])
            # stay unchanged — DB write, AI analyst, sort, score gates and
            # all other numeric consumers continue to work normally.
            _cfv_for_display = stk.get("cfv", 0)
            _cfv_missing = (_cfv_for_display in (0, 0.0, None, "", "—"))
            for ci,(_,_,key) in enumerate(FULL_COLS,1):
                val=_g(stk,key)
                # FV models: 0 means "not applicable" → show "—" not 0
                if key in FV_MODEL_KEYS and (val == 0 or val == 0.0):
                    val = "—"
                # v13.x: tie MoS % / MoS Label display to CFV availability —
                # when CFV is missing the MoS math is a meaningless -100.
                if _cfv_missing and key in ("mos_pct", "mos_label"):
                    val = "—"
                cell=ws.cell(rn,ci,val)
                # v13.x: Quick Pick column uses its own archetype color
                if key == "label":
                    cell.fill=_f(qp_ubg); cell.font=_ft(True,qp_tx,9)
                else:
                    cell.fill=_f(ubg); cell.font=_ft(False,tx,9)
                # v13.x: col 8 added (Quick Pick, left-aligned like Verdict).
                # Cols 28+ shifted by +1 due to new column insertion.
                wrap_cols={2,3,7,8,29,95,97,113,121,123,125}
                cell.alignment=_al("left" if ci in wrap_cols else "center","center",wrap=(ci==N))
                cell.border=Border(
                    left=Side(style="thin",color="E2E8F0"),
                    right=Side(style="thin",color="E2E8F0"),
                    top=Side(style="thin",color="E2E8F0"),
                    bottom=Side(style="thin",color="E2E8F0")
                )
        # Cond format weekly changes
        # v13.x: shifted from [19,20,21,22] → [20,21,22,23] after Quick Pick
        # column was inserted at position 8.
        lr=4+len(stks)
        if lr>4:
            for col_num in [20,21,22,23]:
                cl=get_column_letter(col_num)
                ws.conditional_formatting.add(f"{cl}5:{cl}{lr}",
                    ColorScaleRule(start_type="min",start_color="FEE2E2",
                    mid_type="num",mid_value=0,mid_color="FFFFFF",
                    end_type="max",end_color="D1FAE5"))

    def _gold_sheet(self,wb):
        ws=wb.create_sheet("⭐ Gold – Early Movers"); self._gold_ws(ws)

    def _gold_ws(self,ws):
        gdf=self._get_gold(); N=len(GOLD_COLS)
        gc=len(gdf)
        ae=gdf["early_entry_score"].mean() if not gdf.empty else 0
        au=gdf["upside"].mean()            if not gdf.empty else 0
        asp=gdf["spike_count"].mean()      if not gdf.empty else 0
        # R1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
        c=ws.cell(1,1,f"⭐  NSE / BSE  —  GOLD CATEGORY  ·  EARLY MOVER STOCKS  ·  Highest Upside Before Consensus  ·  {self.dlbl}")
        c.fill=_f("B45309"); c.font=_ft(True,WHITE,13); c.alignment=_al()
        ws.row_dimensions[1].height=30
        # R2
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=N)
        # v10.11: criteria text expanded from 8 → 11 conditions to match the
        # _get_gold() filter. New gates: Altman Z≥1.8 | EQ≠LOW | IntCov≥1.5×.
        c2=ws.cell(2,1,f"Gold-Tier Criteria (ALL 11 must pass): BUY verdict · Score≥70 · 15%≤MoS≤100% · Storm≥5 · RSI≤70 · BS not ALERT · Pledge≤10% · not spike-suppressed · Altman Z≥1.8 · EQ≠LOW · Int Coverage≥1.5×  ·  {gc} stocks qualify")
        c2.fill=_f("FEF3C7"); c2.font=_ft(False,"92400E",8,True); c2.alignment=_al("left","center")
        ws.row_dimensions[2].height=14
        # R3 summary strip
        strips=[(1,5,f"⭐ GOLD STOCKS: {gc}"),(6,11,f"AVG EARLY SCORE: {ae:.0f}/100"),
                (12,17,f"AVG UPSIDE TO FV: +{au:.1f}%"),(18,23,f"AVG SPIKE SCORE: {asp:.1f}/6"),
                (24,N,f"DELIVERED: {self.run_time} · WhatsApp + Email")]
        ws.row_dimensions[3].height=22
        for s,e,t in strips:
            if s!=e: ws.merge_cells(start_row=3,start_column=s,end_row=3,end_column=e)
            c=ws.cell(3,s,t); c.fill=_f("B45309"); c.font=_ft(True,WHITE,9); c.alignment=_al()
        # R4 groups
        ws.row_dimensions[4].height=16
        for sc,nm,col,sp in GOLD_GROUPS:
            ec=sc+sp-1
            if sp>1: ws.merge_cells(start_row=4,start_column=sc,end_row=4,end_column=ec)
            c=ws.cell(4,sc,nm); c.fill=_f(col); c.font=_ft(True,WHITE,8); c.alignment=_al()
        # Session 17: section-header tooltips on Gold sheet too
        _tt_apply_groups(ws, 4, GOLD_GROUPS)
        # R5 headers
        # Build col→group_color map for gold headers
        _gcol_color = {}
        for _sc,_nm,_col,_sp in GOLD_GROUPS:
            for _ci in range(_sc, _sc+_sp):
                _gcol_color[_ci] = _col

        ws.row_dimensions[5].height=38

        # v12.5: dynamic red-header demotion — same coverage logic as the
        # Full Dashboard (≥30 % of rows must carry real data). Pre-v12.5
        # the Gold sheet used a static `if h in NO_FREE_SOURCE_COLS` check,
        # which made columns red even when populated. Now matches the
        # Full Dashboard so a column populated for the (small) Gold cohort
        # demotes from red regardless of its NO_FREE_SOURCE_COLS membership.
        _gold_preview = gdf.to_dict("records")
        _gold_total   = max(1, len(_gold_preview))
        _GOLD_COV_MIN = 0.30
        _gold_has_data = {}
        for (_h, _w, _key) in GOLD_COLS:
            if _h not in NO_FREE_SOURCE_COLS:
                continue
            _real = 0
            for _stk in _gold_preview:
                _v = _stk.get(_key)
                if _v is None: continue
                if _v in ("", "—", "--", "N/A", "STABLE"): continue
                if _v in (0, 0.0, "0", "0.0"): continue
                _real += 1
            _gold_has_data[_h] = (_real / _gold_total) >= _GOLD_COV_MIN

        for i,(h,w,_) in enumerate(GOLD_COLS,1):
            ws.column_dimensions[get_column_letter(i)].width=w
            hdr_bg = _gcol_color.get(i, "92400E")
            if h in NO_FREE_SOURCE_COLS and not _gold_has_data.get(h, False):
                # Bold red — column has no real data in this run
                c=ws.cell(5,i,h); c.fill=_f("991B1B"); c.font=_ft(True,"FEE2E2",8)
            else:
                # Normal section colour — includes previously-red cols that
                # NOW have data on the Gold cohort.
                c=ws.cell(5,i,h); c.fill=_f(hdr_bg); c.font=_ft(True,WHITE,8)
            c.border=_border()
            c.alignment=_al("center","center",True)
        # Tooltips on row 5 headers
        self._apply_col_tips(ws, 5, [h for h,w,_ in GOLD_COLS])
        ws.freeze_panes="A6"
        ws.auto_filter.ref=f"A5:{get_column_letter(N)}5"
        # Data
        for ri,stk in enumerate(gdf.to_dict("records")):
            rn=ri+6; ws.row_dimensions[rn].height=20
            bg="92400E" if ri%2==0 else "FAC775"
            tx=WHITE   if ri%2==0 else "412402"
            # v13.x: Quick Pick gets its own archetype color (same logic as Full sheet)
            qp_label=str(stk.get("label","WATCHLIST"))
            qp_st=VERDICT_STYLES.get(qp_label,{"bg":bg,"text":tx})
            qp_bg,qp_tx=qp_st["bg"],qp_st["text"]
            qp_ubg=qp_bg if ri%2==0 else _alt(qp_bg)
            for ci,(_,_,key) in enumerate(GOLD_COLS,1):
                val=_g(stk,key); cell=ws.cell(rn,ci,val)
                # v13.x: Quick Pick column uses archetype color, not row color
                if key == "label":
                    cell.fill=_f(qp_ubg); cell.font=_ft(True,qp_tx,9)
                else:
                    cell.fill=_f(bg); cell.font=_ft(False,tx,9)
                cell.alignment=_al("left" if ci in {2,3,N} else "center","center",wrap=(ci==N))
                cell.border=Border(
                    left=Side(style="thin",color="E2E8F0"),
                    right=Side(style="thin",color="E2E8F0"),
                    top=Side(style="thin",color="E2E8F0"),
                    bottom=Side(style="thin",color="E2E8F0")
                )

    def _trade_summary(self,wb):
        ws=wb.create_sheet("📊 Trade Summary"); ws.sheet_properties.tabColor="059669"
        gdf=self._get_gold()
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=16)
        c=ws.cell(1,1,f"GOLD STOCKS — TRADE PLAN SUMMARY  ·  Entry / SL / Targets / R:R Ratio  ·  {self.dlbl}")
        c.fill=_f("059669"); c.font=_ft(True,WHITE,11); c.alignment=_al()
        ws.row_dimensions[1].height=26
        hdrs=[("Symbol",12),("Company",25),("CMP (₹)",11),("CFV (₹)",12),
              ("MoS %",10),("Chg% [2-Wk]",13),("Chg% [4-Wk]",13),
              ("Chg% [8-Wk]",13),("Entry Range (₹)",16),("Stop Loss (₹)",13),
              ("Target 1 (₹)",12),("Target 2 (₹)",12),("Target 3 (₹)",12),
              ("R:R Ratio",11),("Time Horizon",22),("Risk Level",11)]
        ws.row_dimensions[2].height=32
        for ci,(h,w) in enumerate(hdrs,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f("059669"); c.font=_ft(True,WHITE,9)
            c.alignment=_al("center","center",True)
        # Tooltips on row 2 headers
        # Session 23: "Upside %" removed from Trade Summary too (duplicated MoS %)
        _ts_hdrs=["Symbol","Company","CMP (₹)","CFV (₹)","MoS %",
                   "Chg% [2-Wk]","Chg% [4-Wk]","Chg% [8-Wk]",
                   "Entry Range (₹)","Stop Loss (₹)","Target 1 (₹)",
                   "Target 2 (₹)","Target 3 (₹)","R:R Ratio",
                   "Time Horizon","Risk Level"]
        self._apply_col_tips(ws, 2, _ts_hdrs)
        ws.freeze_panes="A3"
        for ri,stk in enumerate(gdf.to_dict("records")):
            rn=ri+3; ws.row_dimensions[rn].height=22
            bg="D1FAE5" if ri%2==0 else "ECFDF5"; tx="065F46"
            cmp=_g(stk,"close",0); cfv=_g(stk,"cfv",0)
            mos=_g(stk,"mos_pct",0)
            # Session 23: Upside column removed — same value as MoS. The
            # stock['upside'] key is kept in data dict for backward compat
            # but not displayed as a separate column in Trade Summary.
            vals=[_g(stk,"symbol"),_g(stk,"company_name",""),
                  f"₹{cmp:,}" if isinstance(cmp,(int,float)) and cmp else cmp,
                  f"₹{cfv:,}" if isinstance(cfv,(int,float)) and cfv else cfv,
                  f"+{mos:.1f}%" if isinstance(mos,(int,float)) else mos,
                  _g(stk,"2w_chg"),_g(stk,"4w_chg"),_g(stk,"8w_chg"),
                  _g(stk,"entry_range"),_g(stk,"stop_loss"),
                  _g(stk,"t1"),_g(stk,"t2"),_g(stk,"t3"),
                  None,_g(stk,"horizon"),_g(stk,"risk_level")]
            for ci,val in enumerate(vals,1):
                cell=ws.cell(rn,ci,val); cell.fill=_f(bg); cell.font=_ft(False,tx,9)
                cell.alignment=_al("center","center")
                # Column 10 is now Stop Loss (was 11); red highlight kept
                if ci==10: cell.fill=_f("FEE2E2"); cell.font=_ft(False,"7F1D1D",9)
            # Compute R:R numerically — entry_range is text "353.8–364.7",
            # parse midpoint so Excel formula doesn't fail on text input
            try:
                _er = str(stk.get("entry_range","") or "")
                # Parse "low–high" or "low-high" range text
                for _sep in ["–","-","—","to"]:
                    if _sep in _er:
                        _parts = _er.split(_sep)
                        _entry_mid = (float(_parts[0].replace("₹","").strip()) +
                                      float(_parts[-1].replace("₹","").strip())) / 2
                        break
                else:
                    _entry_mid = float(_er.replace("₹","").strip() or 0)
                _sl_v  = float(str(stk.get("stop_loss","0") or 0).replace("₹","").replace(",",""))
                _t1_v  = float(str(stk.get("t1","0") or 0))
                if _entry_mid > 0 and _sl_v > 0 and _entry_mid > _sl_v and _t1_v > _entry_mid:
                    _rr_val = round((_t1_v - _entry_mid) / (_entry_mid - _sl_v), 2)
                else:
                    _rr_val = "—"
            except Exception:
                _rr_val = "—"
            rr=ws.cell(rn,14,_rr_val); rr.fill=_f(bg); rr.font=_ft(True,"065F46",9)
            rr.alignment=_al()
            if isinstance(_rr_val,(int,float)):
                rr.number_format="0.00"
                # Colour code R:R: green>2, amber 1-2, red<1
                if _rr_val >= 3:   rr.fill=_f("D1FAE5"); rr.font=_ft(True,"065F46",9)
                elif _rr_val >= 2: rr.fill=_f("FEF3C7"); rr.font=_ft(True,"92400E",9)
                else:              rr.fill=_f("FEE2E2"); rr.font=_ft(True,"7F1D1D",9)

    def _alert_log(self,wb):
        ws=wb.create_sheet("🔔 Alert Log"); ws.sheet_properties.tabColor="7C3AED"
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
        c=ws.cell(1,1,f"ALERT LOG  ·  Triggered Signals & Score Changes  ·  Generated {self.run_time}")
        c.fill=_f("7C3AED"); c.font=_ft(True,WHITE,11); c.alignment=_al()
        ws.row_dimensions[1].height=24
        hdrs=[("Date",12),("Time (IST)",11),("Symbol",12),("Alert Type",26),
              ("Trigger Detail",46),("Prev Score",11),("New Score",11),
              ("Score Δ",10),("Action Required",26),("Exchange",12)]
        ws.row_dimensions[2].height=28
        for ci,(h,w) in enumerate(hdrs,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f("7C3AED"); c.font=_ft(True,WHITE,9)
            c.alignment=_al("center","center",True)
        # Tooltips on row 2 headers
        _al_hdrs=["Date","Time (IST)","Symbol","Alert Type","Trigger Detail",
                   "Prev Score","New Score","Score Δ","Action Required","Exchange"]
        self._apply_col_tips(ws, 2, _al_hdrs)
        ws.freeze_panes="A3"
        ACOLS={"⭐ EARLY MOVER DETECTED":"FAC775","🔔 SPIKE FIRED":"D1FAE5",
               "💰 SMART MONEY ENTRY":"D1FAE5","📈 SECTOR STAGE CHANGE":"FAC775",
               "⬇ SCORE DEGRADED":"FEE2E2","🛑 EXIT ALERT":"FEE2E2"}
        ri=3
        for stk in self.df.to_dict("records"):
            sym  = stk.get("symbol","")
            spk  = int(stk.get("spike_count",0) or 0)
            early= _sf(stk.get("early_entry_score",0))
            comp = _sf(stk.get("composite_score",0))
            verd = str(stk.get("verdict","WATCHLIST"))
            mos  = _sf(stk.get("mos_pct", stk.get("upside", 0)))
            vol  = _sf(stk.get("vol_ratio", 1.0))

            if spk>=1 or early>=70 or comp<30:
                # ── Prev Score + Delta ─────────────────────────────────────
                prev = self.prev_scores.get(sym, None)
                if prev is not None and prev > 0:
                    prev_disp = f"{prev:.0f}"
                    delta     = comp - prev
                    delta_disp= f"+{delta:.0f}" if delta > 0 else f"{delta:.0f}"
                else:
                    prev_disp = "—"
                    delta_disp= "—"

                # ── Alert Type ─────────────────────────────────────────────
                if comp < 30:
                    at  = "⬇ SCORE DEGRADED"
                    det = f"Score {comp:.0f}/100 below threshold — review position"
                elif spk >= 1:
                    at  = "🔔 SPIKE FIRED"
                    det = f"Spike {spk}/6 | Score: {comp:.0f} | Early: {early:.0f}/100"
                else:
                    at  = "⭐ EARLY MOVER DETECTED"
                    det = f"Early Entry {early:.0f}/100 | Score: {comp:.0f}"

                # ── Action Required — logic based on verdict/score/MoS/Δ ──
                # Session 24: handle new OVERVALUED verdict (score qualifies
                # for BUY but MoS gate blocks — "great stock, expensive").
                if comp < 30:
                    act = "REVIEW FOR EXIT"
                elif verd == "BUY" and mos > 10 and comp >= 65:
                    act = "CONSIDER ENTRY"
                elif verd == "BUY":
                    act = "MONITOR FOR ENTRY"
                elif verd == "OVERVALUED":
                    act = "STRONG STOCK — WAIT FOR PULLBACK"
                elif verd == "WATCHLIST" and delta_disp != "—" and float(delta_disp) >= 3:
                    act = "SCORE IMPROVING — WATCH"
                elif verd == "WATCHLIST" and delta_disp != "—" and float(delta_disp) <= -3:
                    act = "SCORE DECLINING — CAUTION"
                elif vol >= 3.0:
                    act = "VOLUME ALERT — INVESTIGATE"
                elif early >= 70:
                    act = "EARLY MOVER — ACCUMULATE"
                elif verd == "WATCHLIST":
                    act = "MONITOR CLOSELY"
                else:
                    act = "MONITOR CLOSELY"

                row_data=[self.dlbl, self.run_time, sym, at, det,
                          prev_disp, f"{comp:.0f}", delta_disp, act,
                          stk.get("exchange_tag","NSE")]
                ac=ACOLS.get(at,"FFFFFF"); ws.row_dimensions[ri].height=18
                for ci,val in enumerate(row_data,1):
                    cell=ws.cell(ri,ci,_sv(val))
                    if ci==4: cell.fill=_f(ac); cell.font=_ft(True,NAVY,9)
                    else: cell.fill=_f(LG); cell.font=_ft(False,NAVY,9)
                    cell.alignment=_al("left" if ci==5 else "center","center")
                ri+=1

    def _delivery_preview(self,wb):
        ws=wb.create_sheet("📱 Delivery Preview"); ws.sheet_properties.tabColor="0D9488"
        ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=72
        gdf=self._get_gold(); gc=len(gdf)
        top=gdf.iloc[0] if not gdf.empty else {}
        ws.row_dimensions[1].height=24
        c=ws.cell(1,1,f"DELIVERY PREVIEW  ·  WhatsApp & Email Format  ·  Generated {self.run_time}")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,11)
        def R(r,a,b,bf=None,bb=False,h=16):
            ws.row_dimensions[r].height=h
            if a: ca=ws.cell(r,1,a); ca.font=_ft(True,NAVY,9)
            if b is not None:
                cb=ws.cell(r,2,b); cb.fill=_f(bf or LG)
                cb.font=_ft(bb,NAVY,10); cb.alignment=_al("left","center")
            return r+1
        r=3
        r=R(r,None,"WHATSAPP MESSAGE PREVIEW","0D9488",True,18)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"⭐  NSE/BSE GOLD STOCKS  ·  {self.dlbl}  ·  {self.run_time}","B45309",True)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"⭐  EARLY MOVERS TODAY: {gc} stocks identified","D1FAE5",False)
        if not gdf.empty:
            sym=top.get("symbol","—"); verd=top.get("verdict","—")
            cfv=top.get("cfv",0); cmp_=top.get("close",0)
            mos=top.get("mos_pct",0); up=top.get("upside",0)
            w8=top.get("8w_chg",0); w4=top.get("4w_chg",0); w2=top.get("2w_chg",0)
            spk=int(top.get("spike_count",0) or 0); rot=top.get("sector","—")
            r=R(r,None,f"📈  TOP PICK: {sym} — {verd}","FAC775",True)
            r=R(r,None,f"     CFV: ₹{cfv}  |  CMP: ₹{cmp_}  |  MoS: +{mos:.1f}%  |  Upside: +{up:.1f}%  |  Spike: {spk}/6","FEF3C7",False)
            r=R(r,None,f"     8-Week Chg: {w8}%  |  4-Week Chg: {w4}%  |  2-Week Chg: {w2}%","FEF3C7",False)
            r=R(r,None,f"🔔  SPIKE ALERTS: {spk}  |  🔁  SECTOR: {rot} — Stage {top.get('rotation_stage','—')}","D1FAE5",False)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"📎  NSE_BSE_Gold_EarlyMovers_{self.date_str}.xlsx  (attached)","DBEAFE",False)
        r=R(r,None,f"📎  NSE_BSE_Full_Dashboard_{self.date_str}.xlsx  (subscribers)","DBEAFE",False)
        r+=1; r=R(r,"EMAIL SUBJECT LINE PREVIEW",None,h=18)
        ts=top.get("symbol","—") if not gdf.empty else "—"
        tu=top.get("upside",0)   if not gdf.empty else 0
        r=R(r,None,f"NSE/BSE Research | {self.dlbl} | {gc} Early Movers | Top: {ts} +{tu:.1f}% Upside","EFF6FF",True)
        r+=1; r=R(r,"EXCEL TRIGGER COMMANDS",None,h=18)
        for cmd,desc in [('  "generate excel"',"→  Both files today"),
                         ('  "excel gold only"',"→  Gold file only"),
                         ('  "send to whatsapp"',"→  WhatsApp delivery"),
                         ('  "excel sector: Pharma"',"→  Sector-filtered export"),
                         ('  "refresh excel"',"→  Regenerate with latest data")]:
            ws.row_dimensions[r].height=15
            ws.cell(r,2,f"{cmd:<28}{desc}").font=_ft(False,NAVY,9); r+=1

    def _performance_sheet(self, wb):
        """v14.0: Performance dashboard for Gold-pick outcome tracking.

        Shows historical hit-rate stats from gold_recommendations × gold_outcomes
        DB tables (populated by master_funnel logging + track_outcomes.py walk).

        Sections:
            1. Headline — total tracked, hit rate, SL rate, expiry rate
            2. Speed   — avg days to T1/T2/T3
            3. Diagnostic breakdowns — by score band, archetype, sector
            4. Open positions — currently-tracked stocks with running P&L

        Handles "no data yet" gracefully — when fewer than 30 closed picks
        exist, shows a banner explaining stats need ≥30 closed picks for
        meaningful interpretation. We still display whatever data is there
        so the user can see the system is working.
        """
        ws = wb.create_sheet("🎯 Performance"); ws.sheet_properties.tabColor = "B45309"

        # Pull data from DB
        try:
            from database.data_bridge import get_outcome_stats
            _stats = get_outcome_stats()
            _df_all = _stats.get("all_recommendations", pd.DataFrame())
        except Exception as _pe:
            _df_all = pd.DataFrame()
            print(f"   ⚠️  Performance sheet — DB read failed: {_pe}")

        # Column widths
        for col, w in [("A",26),("B",16),("C",16),("D",16),("E",16),("F",16),
                       ("G",16),("H",16),("I",16),("J",16)]:
            ws.column_dimensions[col].width = w

        # ── Title row ─────────────────────────────────────────────────────
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
        ws.row_dimensions[1].height = 26
        c = ws.cell(1,1,
            f"🎯 GOLD-PICK PERFORMANCE TRACKER  ·  Outcome stats since launch  ·  Generated {self.run_time}")
        c.fill = _f(NAVY); c.font = _ft(True,WHITE,12); c.alignment = _al()

        # If table is empty (first run, no data yet)
        if _df_all.empty:
            ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=10)
            c = ws.cell(3,1,
                "📋 No Gold-pick history yet — tracking starts the first time a stock makes the Gold sheet. "
                "Run the daily pipeline + track_outcomes.py for at least 30 days to see meaningful stats.")
            c.fill = _f("FEF3C7"); c.font = _ft(False,NAVY,10,True); c.alignment = _al("left","center",True)
            ws.row_dimensions[3].height = 36
            return

        # ── Helper to write a labeled metric cell ─────────────────────────
        def _metric(row, col, label, value, *, value_fill="EFF6FF", label_fill="DBEAFE",
                    label_color="1E3A8A", val_bold=True, val_color=NAVY, val_size=14):
            cl = ws.cell(row, col, label)
            cl.fill = _f(label_fill); cl.font = _ft(False, label_color, 9); cl.alignment = _al()
            cv = ws.cell(row+1, col, value)
            cv.fill = _f(value_fill); cv.font = _ft(val_bold, val_color, val_size); cv.alignment = _al()
            ws.row_dimensions[row].height = 16
            ws.row_dimensions[row+1].height = 22

        # Compute headline metrics
        n_total = len(_df_all)
        closed_mask = _df_all["outcome_type"].isin(["SL_HIT","T1_HIT","T2_HIT","T3_HIT","EXPIRED"])
        n_closed = int(closed_mask.sum())
        n_open   = int((_df_all["outcome_type"] == "OPEN").sum())

        if n_closed > 0:
            n_t1 = int((_df_all["outcome_type"] == "T1_HIT").sum())
            n_t2 = int((_df_all["outcome_type"] == "T2_HIT").sum())
            n_t3 = int((_df_all["outcome_type"] == "T3_HIT").sum())
            n_sl = int((_df_all["outcome_type"] == "SL_HIT").sum())
            n_ex = int((_df_all["outcome_type"] == "EXPIRED").sum())
            # T1+ hit rate = anything that reached at least T1 (T1, T2, or T3)
            hit_rate = (n_t1 + n_t2 + n_t3) / n_closed * 100
            sl_rate  = n_sl / n_closed * 100
            ex_rate  = n_ex / n_closed * 100
        else:
            n_t1 = n_t2 = n_t3 = n_sl = n_ex = 0
            hit_rate = sl_rate = ex_rate = 0.0

        # Insufficient-sample banner
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
        if n_closed < 30:
            banner = (f"⚠️ Only {n_closed} closed pick(s) so far — stats below are preliminary. "
                      f"Wait for ≥30 closed picks for statistically meaningful interpretation.")
            ws.cell(3,1,banner).fill = _f("FEF3C7")
            ws.cell(3,1).font = _ft(False,"92400E",9,True)
        else:
            banner = (f"✅ {n_closed} closed picks tracked — sample size is meaningful for "
                      f"diagnostic interpretation. {n_open} positions still open.")
            ws.cell(3,1,banner).fill = _f("D1FAE5")
            ws.cell(3,1).font = _ft(False,"065F46",9,True)
        ws.cell(3,1).alignment = _al("left","center",True)
        ws.row_dimensions[3].height = 24

        # ── Section 1: HEADLINE METRICS ───────────────────────────────────
        ws.merge_cells(start_row=5,start_column=1,end_row=5,end_column=10)
        c = ws.cell(5,1,"📊  HEADLINE METRICS")
        c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
        ws.row_dimensions[5].height = 22

        # Row 6/7: 5 metric cells
        _metric(6, 1, "TOTAL TRACKED",     n_total)
        _metric(6, 2, "CLOSED",            n_closed)
        _metric(6, 3, "OPEN",              n_open,
                value_fill="DBEAFE", label_fill="BFDBFE")
        _metric(6, 4, "HIT RATE (T1+)",    f"{hit_rate:.1f}%",
                value_fill="D1FAE5", label_fill="BBF7D0", val_color="065F46")
        _metric(6, 5, "SL RATE",           f"{sl_rate:.1f}%",
                value_fill="FEE2E2", label_fill="FECACA", val_color="991B1B")

        # Row 9/10: outcome counts
        ws.merge_cells(start_row=9,start_column=1,end_row=9,end_column=10)
        c = ws.cell(9,1,"  Outcome breakdown of closed picks")
        c.fill = _f(LG); c.font = _ft(False,NAVY,9,True); c.alignment = _al("left")
        ws.row_dimensions[9].height = 16
        _metric(10, 1, "T1 HIT", n_t1, value_fill="D1FAE5",  label_fill="A7F3D0", val_color="065F46")
        _metric(10, 2, "T2 HIT", n_t2, value_fill="A7F3D0",  label_fill="6EE7B7", val_color="065F46")
        _metric(10, 3, "T3 HIT", n_t3, value_fill="6EE7B7",  label_fill="34D399", val_color="065F46")
        _metric(10, 4, "SL HIT", n_sl, value_fill="FECACA",  label_fill="FCA5A5", val_color="991B1B")
        _metric(10, 5, "EXPIRED",n_ex, value_fill="E5E7EB",  label_fill="D1D5DB", val_color="374151")

        # ── Section 2: SPEED ──────────────────────────────────────────────
        ws.merge_cells(start_row=13,start_column=1,end_row=13,end_column=10)
        c = ws.cell(13,1,"⏱  SPEED METRICS  ·  Average days from recommendation to event")
        c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
        ws.row_dimensions[13].height = 22

        def _avg_days(outcome):
            sub = _df_all[_df_all["outcome_type"] == outcome]
            if sub.empty: return "—"
            return f"{sub['days_to_outcome'].mean():.1f} days"

        _metric(14, 1, "AVG DAYS → T1", _avg_days("T1_HIT"))
        _metric(14, 2, "AVG DAYS → T2", _avg_days("T2_HIT"))
        _metric(14, 3, "AVG DAYS → T3", _avg_days("T3_HIT"))
        _metric(14, 4, "AVG DAYS → SL", _avg_days("SL_HIT"),
                value_fill="FEE2E2", label_fill="FECACA", val_color="991B1B")

        # ── Section 3: DIAGNOSTIC BREAKDOWNS ──────────────────────────────
        ws.merge_cells(start_row=17,start_column=1,end_row=17,end_column=10)
        c = ws.cell(17,1,"🔬  DIAGNOSTIC BREAKDOWNS  ·  Hit rate by Score band / Archetype / Sector")
        c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
        ws.row_dimensions[17].height = 22

        # Score band breakdown
        def _bucket_score(s):
            try: s = float(s)
            except: return "Unknown"
            if s >= 90: return "≥90 (top tier)"
            if s >= 80: return "80-89"
            if s >= 70: return "70-79"
            return "<70"

        _df_all["_score_band"] = _df_all["composite_score"].apply(_bucket_score)
        _closed = _df_all[closed_mask].copy()
        _closed["_score_band"] = _closed["composite_score"].apply(_bucket_score)

        def _build_breakdown_table(start_row, header, group_col):
            """Render a 5-column table: group | total | T1+/SL/EXP / hit%"""
            ws.cell(start_row, 1, header).font = _ft(True, NAVY, 10)
            ws.cell(start_row, 1).fill = _f("FEF3C7")
            ws.row_dimensions[start_row].height = 18
            # column headers
            for ci, h in enumerate(["Group","Total","T1+ Hits","SL Hits","Expired","Hit Rate"], 1):
                cc = ws.cell(start_row + 1, ci, h)
                cc.fill = _f("FED7AA"); cc.font = _ft(True, NAVY, 9); cc.alignment = _al()
            ws.row_dimensions[start_row + 1].height = 18
            # group rows
            if _closed.empty:
                ws.cell(start_row + 2, 1, "  (no closed picks yet)").font = _ft(False, "6B7280", 9, True)
                return start_row + 3
            grouped = _closed.groupby(group_col).agg(
                total=(group_col, "size"),
                t1_plus=("outcome_type", lambda s: int(s.isin(["T1_HIT","T2_HIT","T3_HIT"]).sum())),
                sl=("outcome_type", lambda s: int((s == "SL_HIT").sum())),
                expired=("outcome_type", lambda s: int((s == "EXPIRED").sum())),
            ).reset_index().sort_values("total", ascending=False)
            r = start_row + 2
            for _, gr in grouped.iterrows():
                tot = int(gr["total"])
                if tot == 0: continue
                hit_pct = gr["t1_plus"] / tot * 100
                bg = LG if r % 2 == 0 else WHITE
                ws.cell(r,1, str(gr[group_col])).fill = _f(bg)
                ws.cell(r,1).font = _ft(False, NAVY, 9); ws.cell(r,1).alignment = _al("left")
                for ci, val in [(2, tot), (3, int(gr["t1_plus"])),
                                (4, int(gr["sl"])), (5, int(gr["expired"]))]:
                    cc = ws.cell(r, ci, val); cc.fill = _f(bg); cc.font = _ft(False, NAVY, 9)
                    cc.alignment = _al()
                # Hit rate cell — colored by quality
                if hit_pct >= 60:    rate_bg = "D1FAE5"; rate_color = "065F46"
                elif hit_pct >= 40:  rate_bg = "FEF3C7"; rate_color = "92400E"
                else:                rate_bg = "FEE2E2"; rate_color = "991B1B"
                cc = ws.cell(r, 6, f"{hit_pct:.1f}%")
                cc.fill = _f(rate_bg); cc.font = _ft(True, rate_color, 9); cc.alignment = _al()
                r += 1
            return r + 1

        next_row = _build_breakdown_table(18, "BY COMPOSITE SCORE BAND", "_score_band")
        next_row = _build_breakdown_table(next_row, "BY QUICK PICK ARCHETYPE", "quick_pick_label")
        next_row = _build_breakdown_table(next_row, "BY SECTOR", "sector")
        # v14.1: 4th breakdown — by time horizon (SHORT TERM / POSITIONAL / LONG TERM)
        # Lets you measure whether the system's own horizon classifications hold up
        # (do POSITIONAL picks actually resolve in 1-3 months? do SHORT TERM picks
        # really hit in 2-4 weeks?). Empty/missing horizon rows roll into "—" group.
        if "time_horizon" in _df_all.columns:
            _df_all["_horizon_clean"] = _df_all["time_horizon"].fillna("—").astype(str).replace("", "—")
            if not _closed.empty:
                _closed["_horizon_clean"] = _closed["time_horizon"].fillna("—").astype(str).replace("", "—")
            next_row = _build_breakdown_table(next_row, "BY TIME HORIZON", "_horizon_clean")

        # ── Section 4: CLOSED POSITIONS (v14.5) ───────────────────────────
        # Chronological detail log of every closed pick (SL_HIT/T1_HIT/T2_HIT/
        # T3_HIT/EXPIRED). The headline counts (T1 HIT: 21, SL HIT: 10, etc.)
        # tell you the aggregate score, but they don't tell you which stocks
        # contributed or how each one played out. This section closes that gap:
        # one row per closed pick, sorted most-recent-first, color-coded by
        # outcome type (green for targets, red for SL, amber for EXPIRED).
        #
        # Sample-size guard: if zero closed rows in DB, render the empty-state
        # line and skip the table entirely (same pattern as OPEN POSITIONS).
        _closed_df = _df_all[_df_all["outcome_type"].isin(
            ["SL_HIT","T1_HIT","T2_HIT","T3_HIT","EXPIRED"])].copy()
        ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=16)
        c = ws.cell(next_row,1,
            "📜  CLOSED POSITIONS  ·  Realised outcomes — historical track record")
        c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
        ws.row_dimensions[next_row].height = 22
        next_row += 1

        # 12-column header — matches OPEN POSITIONS density. Outcome column
        # gets color treatment based on the actual outcome_type value.
        closed_cols = [("Symbol",14),("Rec Date",12),("Time Horizon",13),
                       ("Outcome",10),("Outcome Date",13),("Days to Outcome",14),
                       ("Entry CMP",12),("Outcome Price",13),("P&L %",10),
                       ("Max Runup %",12),("Max Drawdown %",14),("Score",8)]
        for ci,(h,w) in enumerate(closed_cols, 1):
            cc = ws.cell(next_row, ci, h)
            cc.fill = _f(NAVY); cc.font = _ft(True, WHITE, 9); cc.alignment = _al()
            # Note: column widths above 12 may be left untouched — OPEN POSITIONS
            # in the next section will re-set them with its own widths anyway.
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[next_row].height = 20
        next_row += 1

        if _closed_df.empty:
            ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=16)
            c = ws.cell(next_row,1,"  No closed positions yet — track record will populate as picks resolve.")
            c.fill = _f(LG); c.font = _ft(False,"6B7280",9,True); c.alignment = _al("left")
            next_row += 1
        else:
            # Sort by outcome_date DESC (most recent closures first). Falls back
            # gracefully if outcome_date is missing/malformed.
            _closed_df = _closed_df.sort_values(
                by="outcome_date", ascending=False, na_position="last")

            # Color map for outcome column (cell background)
            _outcome_bg = {
                "T1_HIT":  "D1FAE5",   # light green
                "T2_HIT":  "A7F3D0",   # medium green
                "T3_HIT":  "6EE7B7",   # darker green
                "SL_HIT":  "FEE2E2",   # light red
                "EXPIRED": "FEF3C7",   # amber
            }
            _outcome_fg = {
                "T1_HIT":  "065F46",
                "T2_HIT":  "065F46",
                "T3_HIT":  "065F46",
                "SL_HIT":  "991B1B",
                "EXPIRED": "92400E",
            }

            for _, row_c in _closed_df.iterrows():
                _outcome = str(row_c.get("outcome_type","") or "")
                _horizon_str = str(row_c.get("time_horizon","") or "—")
                # Compute P&L % at outcome from entry CMP and outcome_price
                _entry_cmp = float(row_c.get("cmp_at_recommendation", 0) or 0)
                _outcome_price = float(row_c.get("outcome_price", 0) or 0)
                if _entry_cmp > 0 and _outcome_price > 0:
                    _pnl_realised = (_outcome_price - _entry_cmp) / _entry_cmp * 100
                    _pnl_str = f"{_pnl_realised:+.1f}%"
                else:
                    _pnl_realised = 0.0
                    _pnl_str = "—"
                _days_to = row_c.get("days_to_outcome", "")
                _days_to_str = str(int(_days_to)) if _days_to not in ("", None) else "—"
                bg = LG if next_row % 2 == 0 else WHITE
                vals = [
                    str(row_c.get("symbol","—")),
                    str(row_c.get("recommendation_date","—")),
                    _horizon_str,
                    _outcome,
                    str(row_c.get("outcome_date","—") or "—"),
                    _days_to_str,
                    round(_entry_cmp, 2) if _entry_cmp > 0 else "—",
                    round(_outcome_price, 2) if _outcome_price > 0 else "—",
                    _pnl_str,
                    f"{float(row_c.get('max_runup_pct',0) or 0):+.1f}%",
                    f"{float(row_c.get('max_drawdown_pct',0) or 0):+.1f}%",
                    round(float(row_c.get("composite_score",0) or 0), 1),
                ]
                for ci, v in enumerate(vals, 1):
                    cc = ws.cell(next_row, ci, v); cc.fill = _f(bg)
                    cc.font = _ft(False, NAVY, 9); cc.alignment = _al()
                # Color the Outcome column (col 4) with semantic bg + bold fg
                if _outcome in _outcome_bg:
                    cc = ws.cell(next_row, 4)
                    cc.fill = _f(_outcome_bg[_outcome])
                    cc.font = _ft(True, _outcome_fg[_outcome], 9)
                # Color P&L column (col 9): green for positive, red for negative
                if _pnl_realised > 0:
                    ws.cell(next_row, 9).font = _ft(True, "065F46", 9)
                elif _pnl_realised < 0:
                    ws.cell(next_row, 9).font = _ft(True, "991B1B", 9)
                next_row += 1

            # End-of-table summary count by outcome bucket (one-liner)
            _summary_parts = []
            for _otype, _emoji in [("T1_HIT","🎯"),("T2_HIT","🎯🎯"),("T3_HIT","🎯🎯🎯"),
                                    ("SL_HIT","🛑"),("EXPIRED","⏱")]:
                _n = int((_closed_df["outcome_type"] == _otype).sum())
                if _n > 0:
                    _summary_parts.append(f"{_emoji} {_otype}: {_n}")
            if _summary_parts:
                ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=16)
                c = ws.cell(next_row,1,
                    f"  Total {len(_closed_df)} closed · " + " · ".join(_summary_parts))
                c.fill = _f(LG); c.font = _ft(False,NAVY,9,True); c.alignment = _al("left")
                next_row += 1

        next_row += 1   # blank spacer before OPEN POSITIONS

        # ── Section 5: OPEN POSITIONS ─────────────────────────────────────
        # v14.1: enhanced with Horizon, Days Left, Re-appearances, ⚠ approaching-expiry flag
        # v14.4: added SL / T1 / T2 / T3 columns showing target prices + dynamic
        #        distance-from-current-price hints. A reader can now see at a
        #        glance "where are we vs where do we expect to go?" without
        #        cross-referencing the Gold sheet.
        # v15.0: span widened 16→18 to cover new Trailing + Regime columns.
        ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=18)
        c = ws.cell(next_row,1,
            "📂  OPEN POSITIONS  ·  Currently-tracked stocks with running P&L vs targets")
        c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
        ws.row_dimensions[next_row].height = 22
        next_row += 1

        # v14.4: SL/T1/T2/T3 columns between Max Runup % and Score.
        # v15.0: Two new columns appended at end — "Trailing" (current trailing
        # state: '—', 'BE locked', '+3% locked', '+7% locked') and "Regime"
        # (high/neutral/low at log time). Total now 18 columns.
        # Each price/level cell shows "₹price (±X.X%)" — the % is distance
        # from CURRENT price (not entry), updated dynamically each run.
        open_cols = [("Symbol",14),("Rec Date",12),("Time Horizon",15),("Days Held",10),
                     ("Days Left",10),("Re-app",8),("CMP at Rec",12),("Current Price",12),
                     ("P&L %",10),("Max Runup %",12),
                     ("SL",18),("T1",18),("T2",18),("T3",18),
                     ("Score",8),("⚠",6),
                     ("Trailing",14),("Regime",10)]
        for ci,(h,w) in enumerate(open_cols, 1):
            cc = ws.cell(next_row, ci, h)
            cc.fill = _f(NAVY); cc.font = _ft(True, WHITE, 9); cc.alignment = _al()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[next_row].height = 20
        next_row += 1

        _open_df = _df_all[_df_all["outcome_type"] == "OPEN"].copy()
        _approaching_count = 0   # for an end-of-table summary
        if _open_df.empty:
            ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=18)
            c = ws.cell(next_row,1,"  No currently-open positions.")
            c.fill = _f(LG); c.font = _ft(False,"6B7280",9,True); c.alignment = _al("left")
            next_row += 1
        else:
            # Compute days held from recommendation_date to today
            _today_dt = datetime.now()
            for _, row_o in _open_df.iterrows():
                try:
                    _rd = datetime.strptime(str(row_o.get("recommendation_date","")), "%Y-%m-%d")
                    _days_held = (_today_dt - _rd).days
                except Exception:
                    _days_held = 0
                # v14.1: per-rec expiry window — defaults to 90 if column missing (legacy rows)
                _exp_days = int(row_o.get("expiry_days", 90) or 90)
                if _exp_days <= 0: _exp_days = 90
                _days_left = max(0, _exp_days - _days_held)
                _is_approaching = _days_left <= 14
                if _is_approaching: _approaching_count += 1
                _times_reap = int(row_o.get("times_reappeared", 0) or 0)
                _horizon_str = str(row_o.get("time_horizon","") or "—")
                bg = LG if next_row % 2 == 0 else WHITE
                # Highlight whole row in pale yellow if approaching expiry
                if _is_approaching: bg = "FEF3C7"
                # v14.4: format SL/T1/T2/T3 as "₹price (±X.X%)" — distance from
                # current price. SL distance is naturally negative (price would
                # need to drop for SL to fire); target distances naturally positive
                # while we're below them.
                _cur_price = float(row_o.get("current_price",0) or 0)
                _sl_v = float(row_o.get("stop_loss",0) or 0)
                _t1_v = float(row_o.get("t1",0) or 0)
                _t2_v = float(row_o.get("t2",0) or 0)
                _t3_v = float(row_o.get("t3",0) or 0)
                def _level_str(level_price):
                    """Format a target/SL level: '₹P (±D.D%)' where D is distance
                    from current price. Returns '—' for missing levels."""
                    if not level_price or level_price <= 0 or _cur_price <= 0:
                        return "—"
                    dist_pct = (level_price - _cur_price) / _cur_price * 100
                    return f"₹{level_price:,.2f} ({dist_pct:+.1f}%)"
                _sl_str = _level_str(_sl_v)
                _t1_str = _level_str(_t1_v)
                _t2_str = _level_str(_t2_v)
                _t3_str = _level_str(_t3_v)
                # v15.0: derive Trailing-SL label + Regime label for columns 17/18.
                _trailing_pct   = float(row_o.get("trailing_sl_pct", 0) or 0)
                _trailing_price = float(row_o.get("trailing_sl_price", 0) or 0)
                if _trailing_price <= 0:
                    _trail_label = "—"
                elif abs(_trailing_pct) < 0.5:
                    _trail_label = "BE locked"
                elif _trailing_pct >= 6.5:
                    _trail_label = "+7% locked"
                elif _trailing_pct >= 2.5:
                    _trail_label = "+3% locked"
                else:
                    _trail_label = f"+{_trailing_pct:.1f}% locked"
                _regime = str(row_o.get("regime_at_rec", "") or "").strip().lower()
                _regime_label = _regime.upper() if _regime in ("high","low","neutral") else "—"
                vals = [
                    str(row_o.get("symbol","—")),
                    str(row_o.get("recommendation_date","—")),
                    _horizon_str,
                    _days_held,
                    _days_left,
                    _times_reap if _times_reap > 0 else "—",
                    round(float(row_o.get("cmp_at_recommendation",0) or 0), 2),
                    round(_cur_price, 2),
                    f"{float(row_o.get('current_pnl_pct',0) or 0):+.1f}%",
                    f"{float(row_o.get('max_runup_pct',0) or 0):+.1f}%",
                    _sl_str, _t1_str, _t2_str, _t3_str,
                    round(float(row_o.get("composite_score",0) or 0), 1),
                    "⚠" if _is_approaching else "",
                    _trail_label,           # v15.0 col 17
                    _regime_label,          # v15.0 col 18
                ]
                for ci, v in enumerate(vals, 1):
                    cc = ws.cell(next_row, ci, v); cc.fill = _f(bg)
                    cc.font = _ft(False, NAVY, 9); cc.alignment = _al()
                # Color P&L cell (column 9)
                pnl = float(row_o.get("current_pnl_pct",0) or 0)
                if pnl > 0:    ws.cell(next_row, 9).font = _ft(True, "065F46", 9)
                elif pnl < 0:  ws.cell(next_row, 9).font = _ft(True, "991B1B", 9)
                # v14.4: color SL column (11) red — always represents downside risk.
                # Color target columns (12/13/14) green — always represents upside.
                # Subtle text color, not bold (avoids visual overwhelm with so many cols).
                if _sl_v > 0:
                    ws.cell(next_row, 11).font = _ft(False, "991B1B", 9)
                if _t1_v > 0:
                    ws.cell(next_row, 12).font = _ft(False, "065F46", 9)
                if _t2_v > 0:
                    ws.cell(next_row, 13).font = _ft(False, "065F46", 9)
                if _t3_v > 0:
                    ws.cell(next_row, 14).font = _ft(False, "065F46", 9)
                # Bold the warning emoji (column 16 in v14.4, was 12 in v14.1)
                if _is_approaching:
                    ws.cell(next_row, 16).font = _ft(True, "92400E", 11)
                # v15.0: color Trailing column (17) green when locked, Regime
                # column (18) red for HIGH, green for LOW, default for NEUTRAL.
                if _trailing_price > 0:
                    ws.cell(next_row, 17).font = _ft(True, "065F46", 9)
                if _regime == "high":
                    ws.cell(next_row, 18).font = _ft(True, "991B1B", 9)
                elif _regime == "low":
                    ws.cell(next_row, 18).font = _ft(True, "065F46", 9)
                next_row += 1
            # End-of-table summary line if any approaching expiry
            if _approaching_count > 0:
                ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=18)
                c = ws.cell(next_row,1,
                    f"  ⚠ {_approaching_count} position(s) within 14 days of expiry — "
                    "review for manual exit or accept EXPIRED outcome at the hard cutoff.")
                c.fill = _f("FEF3C7"); c.font = _ft(False,"92400E",9,True); c.alignment = _al("left")
                ws.row_dimensions[next_row].height = 20
                next_row += 1

        # ── v14.1 Section 5: EXPIRED — MISSED RUNUP DIAGNOSTIC ────────────
        # For EXPIRED rows, show what max runup the stock reached during
        # tracking. High avg = "we're leaving money on the table by expiring
        # positions too early or with too-tight targets". Low avg = "the
        # stock genuinely went sideways, expiry was the right call."
        _expired_df = _df_all[_df_all["outcome_type"] == "EXPIRED"].copy()
        if not _expired_df.empty and len(_expired_df) >= 3:
            next_row += 1   # spacer
            ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=16)
            c = ws.cell(next_row,1,
                "💸  EXPIRED — MISSED RUNUP DIAGNOSTIC  ·  How much was 'left on the table' before expiry")
            c.fill = _f("B45309"); c.font = _ft(True,WHITE,11); c.alignment = _al("left")
            ws.row_dimensions[next_row].height = 22
            next_row += 1

            avg_missed = _expired_df["max_runup_pct"].astype(float).mean()
            max_missed = _expired_df["max_runup_pct"].astype(float).max()
            count_significant = int((_expired_df["max_runup_pct"].astype(float) >= 10).sum())
            # 3 summary cells
            _metric(next_row, 1, "AVG MISSED RUNUP", f"{avg_missed:+.1f}%",
                    value_fill="FEF3C7", label_fill="FED7AA", val_color="92400E")
            _metric(next_row, 2, "PEAK MISSED RUNUP", f"{max_missed:+.1f}%",
                    value_fill="FEF3C7", label_fill="FED7AA", val_color="92400E")
            _metric(next_row, 3, "EXPIRED w/ ≥10% RUNUP",
                    f"{count_significant}/{len(_expired_df)}",
                    value_fill="FEE2E2", label_fill="FECACA", val_color="991B1B")
            next_row += 3
            # Interpretive note
            ws.merge_cells(start_row=next_row,start_column=1,end_row=next_row,end_column=16)
            c = ws.cell(next_row, 1,
                "  Reading guide: high AVG MISSED RUNUP (>15%) suggests targets are too far "
                "or expiry too early. Stocks rallied during tracking but failed to reach T1 within "
                "the horizon's window. Investigate top offenders manually.")
            c.fill = _f(LG); c.font = _ft(False, "475569", 9, True); c.alignment = _al("left", "center", True)
            ws.row_dimensions[next_row].height = 30
            next_row += 1

    def _glossary(self,wb):
        ws=wb.create_sheet("📖 Glossary"); ws.sheet_properties.tabColor="475569"
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)
        c=ws.cell(1,1,"GLOSSARY  ·  Full Forms & Descriptions of All Column Abbreviations  ·  NSE/BSE Stock Analyser v6.2")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,12); c.alignment=_al()
        ws.row_dimensions[1].height=28
        ws.column_dimensions["A"].width=3
        for ci,(h,w) in enumerate([("Group",18),("Short Name / Abbreviation",28),
                                    ("Full Form & Description",70),("Where Used",16)],2):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f(NAVY); c.font=_ft(True,WHITE,9); c.alignment=_al()
            ws.row_dimensions[2].height=28
        ws.freeze_panes="B3"
        import math as _math_gl
        for ri,(grp,short,desc,where) in enumerate(GLOSSARY_DATA,3):
            # Auto-height: col D is 70 units wide, font-9 ≈ 85 chars/line
            # Each wrapped line ≈ 13pt; +4pt padding; 16pt minimum
            _desc_str = str(desc) if desc else ""
            _lines    = max(1, -(-len(_desc_str) // 85))  # ceiling division
            ws.row_dimensions[ri].height = max(16, _lines * 13 + 4)
            bg=LG if ri%2==0 else WHITE
            gc=GRP_COLORS.get(grp,"475569")
            c=ws.cell(ri,2,grp); c.fill=_f(gc); c.font=_ft(True,WHITE,9); c.alignment=_al()
            c=ws.cell(ri,3,str(short) if short else ""); c.fill=_f(bg); c.font=_ft(True,NAVY,9); c.alignment=_al("left","top"); c.data_type="s"
            c=ws.cell(ri,4,_desc_str); c.fill=_f(bg); c.font=_ft(False,"475569",9); c.alignment=_al("left","top",True); c.data_type="s"
            c=ws.cell(ri,5,where); c.fill=_f(bg); c.font=_ft(False,NAVY,9); c.alignment=_al("center","top")

    def _get_gold(self):
        if self.df.empty: return pd.DataFrame()
        try:
            # Session 19 + v10.11 strict Gold-tier filter.
            # Session 19 (8 conditions): Verdict=BUY, Score>=70, 15<=MoS<=100,
            #   Storm>=5, RSI<=70, BS!=ALERT, Pledge<=10, not suppressed.
            # v10.11 (3 new forensic gates using v10.8+v10.9 populated fields):
            #   9. Altman Z >= 1.8  (not in distress zone — if field populated)
            #   10. Earn Quality != "LOW"  (no accounting concern)
            #   11. Int Coverage >= 1.5x  (can service interest — if populated)
            # Fields populated as "—" (missing data) PASS these gates so small
            # caps without forensic data aren't unfairly excluded — the existing
            # 8 gates already cover them via BS Health + Pledge.
            #
            # All 11 conditions must be true for Gold-tier:
            #  1. Verdict = BUY                 — system-confident, not WATCHLIST
            #  2. Score >= 70                   — uniform Gold bar, not cap-adjusted
            #  3. 15 <= MoS <= 100              — real upside, not phantom inflation
            #  4. Storm Score >= 5              — defensively sound
            #  5. RSI <= 70                     — not already overbought
            #  6. BS Health Flag != ALERT       — no balance-sheet red flags
            #  7. Pledge % <= 10                — Gold = clean, not just "not awful"
            #  8. spike_suppressed == False     — Altman/Beneish/pledge all clear
            #  9. Altman Z >= 1.8 OR "—"         — v10.11: not in distress zone
            # 10. Earn Quality != "LOW"         — v10.11: no accounting concern
            # 11. Int Coverage >= 1.5 OR "—"    — v10.11: can service interest
            _mos = self.df["mos_pct"]
            _rsi = self.df.get("rsi", pd.Series([50]*len(self.df)))
            _storm = self.df.get("storm_score", pd.Series([0]*len(self.df)))
            _pledge = pd.to_numeric(self.df.get("pledge_pct",
                                                pd.Series([0]*len(self.df))),
                                    errors="coerce").fillna(0)
            _bs = self.df.get("bs_status", pd.Series([""]*len(self.df))) \
                       .astype(str).str.upper()

            # v10.11 new gates — tolerant of "—" (missing data passes)
            _alt_raw = self.df.get("altman_z", pd.Series(["—"]*len(self.df)))
            _alt_num = pd.to_numeric(_alt_raw, errors="coerce")
            # Pass if Altman ≥ 1.8 OR is missing (NaN)
            _alt_gate = (_alt_num >= 1.8) | _alt_num.isna()

            _eq = self.df.get("earnings_quality",
                              pd.Series(["—"]*len(self.df))).astype(str).str.upper()
            # Pass if Earn Quality is NOT "LOW" (HIGH/MODERATE/— all pass)
            _eq_gate = _eq != "LOW"

            _ic_raw = self.df.get("int_coverage", pd.Series(["—"]*len(self.df)))
            _ic_num = pd.to_numeric(_ic_raw, errors="coerce")
            # Pass if Int Coverage ≥ 1.5 OR is missing
            _ic_gate = (_ic_num >= 1.5) | _ic_num.isna()

            mask = (
                (self.df["verdict"] == "BUY") &
                (self.df["composite_score"] >= 70) &
                (_mos >= 15) & (_mos <= 100) &
                (_storm >= 5) &
                (pd.to_numeric(_rsi, errors="coerce").fillna(50) <= 70) &
                (~_bs.str.contains("ALERT", na=False)) &
                (_pledge <= 10) &
                (self.df["spike_suppressed"] == False) &
                _alt_gate &    # v10.11: not distressed
                _eq_gate &     # v10.11: not accounting concern
                _ic_gate       # v10.11: can service interest
            )
            return self.df[mask].copy().reset_index(drop=True)
        except Exception: return pd.DataFrame()