"""
excel_generator.py  —  NSE/BSE Stock Analyser v6 Dashboard Generator
Matches the reference template exactly: 6 sheets, correct colours,
group headers, column order, row heights, and column widths.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime

NAVY  = "1E293B";  WHITE = "FFFFFF";  LG = "F8FAFC"

VERDICT_STYLES = {
    "DEEP VALUE EARLY MOVER": {"bg":"FAC775","text":"412402"},
    "EARLY MOVER":            {"bg":"FAC775","text":"412402"},
    "DEEP VALUE":             {"bg":"D1FAE5","text":"065F46"},
    "BUY":                    {"bg":"DBEAFE","text":"1E3A5F"},
    "BUY / EARLY MOVER":      {"bg":"DBEAFE","text":"1E3A5F"},
    "WATCHLIST":              {"bg":"FEF3C7","text":"78350F"},
    "NEUTRAL":                {"bg":"FEF3C7","text":"78350F"},
    "AVOID":                  {"bg":"FEE2E2","text":"7F1D1D"},
    "AVOID / EXIT":           {"bg":"FEE2E2","text":"7F1D1D"},
    "EXIT":                   {"bg":"FEE2E2","text":"7F1D1D"},
}
_DS = {"bg":"F8FAFC","text":"1E293B"}

FULL_GROUPS = [
    (1,  "IDENTITY",        "1E293B",7),(8, "SCORES",         "7C3AED",4),
    (12, "PRICE & MARKET",  "0369A1",7),(19,"WEEKLY CHANGE %", "0F766E",4),
    (23, "FAIR VALUE",      "B45309",13),(36,"VALUATION",      "0891B2",7),
    (43, "PROFITABILITY",   "059669",10),(53,"GROWTH",         "047857",10),
    (63, "FIN HEALTH",      "DC2626",10),(73,"CAP ALLOC",      "6D28D9",3),
    (76, "SHAREHOLDING",    "EA580C",9),(85,"QUALITY SCORES",  "0D9488",4),
    (89, "PIPELINE / OB",   "1D4ED8",5),(94,"EARLY DETECTION", "B45309",3),
    (97, "TECHNICAL",       "6D28D9",14),(111,"BALANCE SHEET", "D97706",2),
    (113,"TRADE PLAN",      "059669",7),(120,"NEWS & RISK",    "475569",4),
    (124,"ANALYSIS SUMMARY","0F172A",1),
]

FULL_COLS = [
    ("Symbol",12,"symbol"),("Company Name",28,"company_name"),("Sector",22,"sector"),
    ("Exchange",13,"exchange_tag"),("BSE Code",10,"bse_code"),("Cap Category",13,"cap_category"),
    ("Verdict",26,"verdict"),("Score /100",10,"composite_score"),("Early Entry /100",14,"early_entry_score"),
    ("Spike Score /6",11,"spike_count"),("Storm Score /10",12,"storm_score"),
    ("CMP (₹)",11,"close"),("Day Chg %",10,"day_change"),("52W High (₹)",12,"high_52w"),
    ("52W Low (₹)",12,"low_52w"),("Vol Spike (×50D)",14,"vol_ratio"),("Delivery %",11,"delivery_pct"),
    ("Beta",8,"beta"),("Chg% [2-Weekly]",14,"2w_chg"),("Chg% [4-Weekly]",14,"4w_chg"),
    ("Chg% [6-Weekly]",14,"6w_chg"),("Chg% [8-Weekly]",14,"8w_chg"),
    ("CFV (₹)",11,"cfv"),("FV Low (₹)",11,"cfv_low"),("FV High (₹)",11,"cfv_high"),
    ("MoS %",10,"mos_pct"),("Upside to FV %",14,"upside"),("MoS Label",22,"mos_label"),
    ("M1: DCF FV (₹)",14,"M1_DCF"),("M2: Graham FV (₹)",16,"M2_Graham"),
    ("M3: PE FV (₹)",14,"M3_PE"),("M4: PB FV (₹)",14,"M4_PB"),
    ("M5: EV FV (₹)",14,"M5_EV"),("M6: DDM FV (₹)",14,"M6_DDM"),("M7: PEG FV (₹)",14,"M7_PEG"),
    ("P/E TTM",9,"pe"),("Earn Yield %",11,"earnings_yield"),("P/CF",9,"p_cf"),
    ("PEG Ratio",10,"peg"),("P/B",9,"pb"),("P/S",9,"ps"),("EV/EBITDA",11,"ev_ebitda"),
    ("ROE %",9,"roe"),("ROCE %",9,"roce"),("ROA %",9,"roa"),
    ("Gross Mgn %",11,"gross_margin"),("EBITDA Mgn %",12,"ebitda_margin"),("NPM %",9,"npm"),
    ("NPM Q1 %",9,"npm_q1"),("NPM Q2 %",9,"npm_q2"),("NPM Q3 %",9,"npm_q3"),
    ("Margin Expansion",17,"margin_expansion"),
    ("Rev CAGR 1Y %",12,"rev_cagr_1y"),("Rev CAGR 3Y %",12,"rev_cagr_3y"),
    ("PAT CAGR 1Y %",12,"pat_cagr_1y"),("PAT CAGR 3Y %",12,"pat_cagr_3y"),
    ("EBITDA CAGR 1Y %",15,"ebitda_cagr_1y"),("Rev YoY %",10,"rev_yoy"),
    ("PAT YoY %",10,"pat_yoy"),("Q3 Rev (₹Cr)",13,"q3_rev"),
    ("Q3 PAT (₹Cr)",13,"q3_pat"),("Q3 EBITDA (₹Cr)",15,"q3_ebitda"),
    ("D/E Ratio",10,"debt_equity"),("ND/EBITDA",11,"nd_ebitda"),
    ("Int Coverage",12,"int_coverage"),("Current Ratio",13,"current_ratio"),
    ("Quick Ratio",11,"quick_ratio"),("Cash (₹Cr)",11,"cash"),
    ("Total Debt (₹Cr)",15,"total_debt"),("FCF (₹Cr)",11,"fcf"),
    ("FCF Yield %",12,"fcf_yield"),("CCC Days",10,"ccc_days"),
    ("Div Yield %",11,"div_yield"),("Payout Ratio %",13,"payout_ratio"),
    ("Capex / Rev %",12,"capex_rev"),
    ("Promoter %",11,"promoter_pct"),("Pro QoQ Δ",10,"promoter_qoq"),
    ("Pledge %",10,"pledge_pct"),("Pledge Direction",15,"pledge_direction"),
    ("FII %",9,"fii_pct"),("FII QoQ Δ",10,"fii_qoq"),
    ("DII %",9,"dii_pct"),("DII QoQ Δ",10,"dii_qoq"),("Public Float %",13,"public_float"),
    ("Piotroski F /9",13,"piotroski_f"),("Altman Z",10,"altman_z"),
    ("Beneish M",10,"beneish_m"),("Earn Quality",12,"earnings_quality"),
    ("OB/Bill Ratio",12,"ob_bill_ratio"),("Pipeline Vis",12,"pipeline_vis"),
    ("L1 Wins 90D",12,"l1_wins"),("L1 Est (₹Cr)",13,"l1_value"),
    ("New Mkt Entry",26,"new_market_entry"),
    ("Early Signals",42,"early_signals"),("Sector Stage",11,"rotation_stage"),
    ("Smart Money",28,"smart_money_signals"),
    ("SMA 200",10,"sma_200"),("Supertrend",12,"supertrend"),("ADX",8,"adx"),
    ("RSI (14)",9,"rsi"),("MACD Signal",18,"macd_signal"),("Stoch %K",9,"stoch_k"),
    ("MFI",8,"mfi"),("OBV Signal",14,"obv_signal"),("Above VWAP",11,"above_vwap"),
    ("Chart Pattern",22,"chart_pattern"),
    ("Support 1 (₹)",12,"support_1"),("Support 2 (₹)",12,"support_2"),
    ("Resist 1 (₹)",12,"resist_1"),("Resist 2 (₹)",12,"resist_2"),
    ("BS Health Flag",13,"bs_status"),("BS Health Note",40,"bs_flags"),
    ("Entry Range (₹)",15,"entry_range"),("Stop Loss (₹)",13,"stop_loss"),
    ("Target 1 (₹)",12,"t1"),("Target 2 (₹)",12,"t2"),("Target 3 (₹)",12,"t3"),
    ("Time Horizon",22,"horizon"),("Risk Level",11,"risk_level"),
    ("Key Catalyst",42,"key_catalyst"),("News Sentiment",15,"news_sentiment"),
    ("Primary Risk",42,"primary_risk"),("SEBI Flags",22,"sebi_flags"),
    ("View Analysis Summary",70,"Analysis_Summary_Block_H"),
]

GOLD_GROUPS = [
    (1,"IDENTITY","1E293B",6),(7,"SCORES","7C3AED",4),(11,"PRICE","0369A1",1),
    (12,"WEEKLY CHANGE %","0F766E",4),(16,"FAIR VALUE","B45309",3),
    (19,"KEY METRICS","0891B2",7),(26,"EARLY DETECTION","B45309",3),
    (29,"TECHNICAL","6D28D9",3),(32,"TRADE PLAN","059669",7),
    (39,"NEWS","475569",2),(41,"ANALYSIS SUMMARY","0F172A",1),
]

GOLD_COLS = [
    ("Symbol",12,"symbol"),("Company Name",28,"company_name"),("Sector",22,"sector"),
    ("Exchange",13,"exchange_tag"),("Cap Category",13,"cap_category"),("Verdict",26,"verdict"),
    ("Score /100",10,"composite_score"),("Early Entry /100",14,"early_entry_score"),
    ("Spike /6",9,"spike_count"),("Storm /10",9,"storm_score"),
    ("CMP (₹)",11,"close"),("Chg% [2-Wk]",13,"2w_chg"),("Chg% [4-Wk]",13,"4w_chg"),
    ("Chg% [6-Wk]",13,"6w_chg"),("Chg% [8-Wk]",13,"8w_chg"),
    ("CFV (₹)",11,"cfv"),("MoS %",10,"mos_pct"),("Upside %",11,"upside"),
    ("MoS Label",20,"mos_label"),("P/E",9,"pe"),("PEG",9,"peg"),
    ("ROE %",9,"roe"),("D/E",9,"debt_equity"),("PAT YoY %",10,"pat_yoy"),
    ("F-Score /9",11,"piotroski_f"),
    ("Early Signals",42,"early_signals"),("Smart Money",28,"smart_money_signals"),
    ("Sector Stage",11,"rotation_stage"),("Supertrend",12,"supertrend"),
    ("RSI (14)",9,"rsi"),("Pattern",24,"chart_pattern"),
    ("Entry Range (₹)",16,"entry_range"),("Stop Loss (₹)",12,"stop_loss"),
    ("Target 1 (₹)",12,"t1"),("Target 2 (₹)",12,"t2"),("Target 3 (₹)",12,"t3"),
    ("Horizon",22,"horizon"),("Risk Level",10,"risk_level"),
    ("Key Catalyst",42,"key_catalyst"),("Primary Risk",42,"primary_risk"),
    ("View Analysis Summary",70,"Analysis_Summary_Block_H"),
]

GLOSSARY_DATA = [
    ("IDENTITY","Symbol","NSE/BSE ticker symbol. Format: up to 20 chars, no spaces (e.g. TATAMOTORS, HDFCBANK)","All sheets"),
    ("IDENTITY","BSE Code","6-digit BSE scrip code (e.g. 500325 = Reliance). Use on BSE for orders.","Full Dashboard"),
    ("IDENTITY","Cap Category",
     "LARGE CAP = mcap > ₹20,000 Cr (top 100 stocks, lowest risk) | "
     "MID CAP = ₹5,000–20,000 Cr (101–250 stocks, moderate risk) | "
     "SMALL CAP = ₹500–5,000 Cr (251+ stocks, higher risk) | "
     "MICRO CAP = < ₹500 Cr (highest risk, highest potential)","All sheets"),
    ("IDENTITY","Verdict",
     "BUY = score above cap threshold AND MoS ≥ −10% (CMP at/below fair value, act now) | "
     "WATCHLIST = signal building, almost BUY or score qualifies but overvalued (monitor) | "
     "NEUTRAL = no clear signal yet, hold off | "
     "AVOID = weak fundamentals + bad technicals + overvalued (stay out) | "
     "DEEP VALUE = significantly undervalued (MoS>25%) with high score | "
     "EARLY MOVER = early signal detected before consensus (act ahead of crowd)","All sheets"),
    ("IDENTITY","Exchange",
     "NSE_ONLY = only on National Stock Exchange | "
     "BSE_ONLY = only on Bombay Stock Exchange | "
     "DUAL_LISTED = on both NSE and BSE (broader institutional access, preferred) | "
     "BSE_SME = BSE Small & Medium Enterprise platform (use limit orders, low liquidity)","All sheets"),
    ("SCORES","Score /100","Composite: Fundamental 35% + Technical 30% + Early 15% + News 10% + Risk 10%","All sheets"),
    ("SCORES","Early Entry /100","12-signal system measuring how early vs consensus. ≥70 = EARLY MOVER badge","All sheets"),
    ("SCORES","Spike Score /6","Count of active triggers from 6 IF-THEN spike conditions (Section 3H)","All sheets"),
    ("SCORES","Storm Score /10","Volatility resilience score. Higher = more defensive in downturns (VIX>18)","Full Dashboard"),
    ("PRICE & MARKET","CMP","Current Market Price in Indian Rupees (₹)","All sheets"),
    ("PRICE & MARKET","52W High","Highest closing price in past 52 weeks","Full Dashboard"),
    ("PRICE & MARKET","52W Low","Lowest closing price in past 52 weeks","Full Dashboard"),
    ("PRICE & MARKET","Vol Spike (×50D)","Today's volume ÷ 50-day average. >3× = institutional activity","Full Dashboard"),
    ("PRICE & MARKET","Delivery %","Share of traded volume with actual delivery. >60% = conviction","Full Dashboard"),
    ("PRICE & MARKET","Beta","Beta vs Nifty 50. <1 = defensive. >1.5 = volatile","Full Dashboard"),
    ("WEEKLY CHANGE %","Chg% [2-Weekly]","Price change % over past 2 weeks (10 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [4-Weekly]","Price change % over past 4 weeks (20 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [6-Weekly]","Price change % over past 6 weeks (30 trading days)","All sheets"),
    ("WEEKLY CHANGE %","Chg% [8-Weekly]","+25% on a quality stock over 8W = institutional accumulation","All sheets"),
    ("FAIR VALUE","CFV","Composite Fair Value — sector-weighted blend of 7 models (M1–M7)","All sheets"),
    ("FAIR VALUE","MoS %",
     "Margin of Safety = (CFV−CMP)/CMP×100. How much cheaper CMP is vs fair value. "
     "EXCEPTIONAL VALUE = MoS > 40% (deeply undervalued, strong BUY signal) | "
     "STRONG VALUE = MoS 25–40% (significantly undervalued) | "
     "GOOD VALUE = MoS 10–25% (moderately undervalued) | "
     "FAIR VALUE = MoS 0–10% (priced fairly) | "
     "SLIGHT PREMIUM = MoS −15–0% (slightly overvalued, ok for quality stocks) | "
     "OVERVALUED = MoS −30–−15% (CMP exceeds FV, caution) | "
     "SIGNIFICANTLY OVERVALUED = MoS < −30% (avoid, major downside risk)","All sheets"),
    ("FAIR VALUE","M1: DCF FV","3-Stage DCF. WACC = 10Y GSec + Beta×ERP. Terminal growth 4.5%","Full Dashboard"),
    ("FAIR VALUE","M2: Graham FV","Graham Number = √(22.5×EPS×BVPS). Skip if EPS negative","Full Dashboard"),
    ("FAIR VALUE","M3: PE FV","EPS × Sector 5yr median P/E (mean reversion)","Full Dashboard"),
    ("FAIR VALUE","M4: PB FV","BVPS × Sector median Price/Book. Good for asset-heavy sectors (banks, metals)","Full Dashboard"),
    ("FAIR VALUE","M5: EV FV","CMP × (Sector median EV/EBITDA ÷ Stock EV/EBITDA). Good for capital-intensive businesses","Full Dashboard"),
    ("FAIR VALUE","M6: DDM FV","Dividend Discount Model = DPS×(1+g)/(r−g). Only shown for dividend-paying stocks","Full Dashboard"),
    ("FAIR VALUE","M7: PEG FV","EPS × EPS Growth Rate. PEG=1 baseline","Full Dashboard"),
    ("VALUATION","P/E TTM","Price-to-Earnings trailing 12 months","All sheets"),
    ("VALUATION","Earn Yield %","EPS/CMP×100. >6% = beats 10Y GSec risk-free rate","All sheets"),
    ("VALUATION","P/E TTM",
     "Price ÷ EPS (trailing 12 months). How much you pay per ₹1 of earnings. "
     "< 10 = Very cheap (PSUs, cyclicals) | "
     "10–20 = Fair value for slow-growth / value stocks | "
     "20–35 = Growth premium (acceptable if ROE>20% and growth>15%) | "
     "35–60 = Expensive (only justified by very high growth) | "
     "> 60 = Very expensive (speculative, future growth already priced in)","All sheets"),
    ("VALUATION","PEG Ratio",
     "P/E ÷ EPS Growth Rate. Adjusts PE for growth. "
     "< 0.5 = Deeply undervalued vs growth (excellent BUY) | "
     "0.5–1.0 = Undervalued vs growth (good value) | "
     "1.0 = Fairly valued (Peter Lynch's neutral benchmark) | "
     "1.0–2.0 = Slight premium (acceptable for quality compounder) | "
     "> 2.0 = Expensive relative to growth (avoid unless market leader)","All sheets"),
    ("VALUATION","P/B",
     "Price ÷ Book Value per Share. "
     "< 1.0 = Trading below book (potential deep value or value trap) | "
     "1.0–2.0 = Reasonable (good for banks, metals, asset-heavy) | "
     "2.0–5.0 = Premium (justified if ROE>15%) | "
     "> 5.0 = High premium (only for asset-light compounders with high ROE)","All sheets"),
    ("VALUATION","EV/EBITDA",
     "Enterprise Value ÷ EBITDA. Better than PE as it ignores capital structure. "
     "< 8 = Cheap (cyclicals, turnarounds) | "
     "8–15 = Fair value range | "
     "15–25 = Premium (growth sectors like IT, pharma) | "
     "> 25 = Expensive (only for market leaders or high-growth)","Full Dashboard"),
    ("VALUATION","Earn Yield %",
     "EPS ÷ CMP × 100. Inverse of PE — shows what % return you earn per ₹ invested. "
     "< 4% = Very expensive (below risk-free rate) | "
     "4–6% = Below risk-free (10Y G-Sec ~7%) | "
     "> 6% = Beats G-Sec (attractive for value investors) | "
     "> 8% = High yield (deep value territory)","All sheets"),
    ("PROFITABILITY","ROE %","Return on Equity. >15% = efficient. 5yr avg preferred","All sheets"),
    ("PROFITABILITY","NPM Q1/Q2/Q3","Net Profit Margin last 3 quarters. 3 rising = Margin Expansion flag","Full Dashboard"),
    ("GROWTH","Rev CAGR 1Y %","Revenue CAGR over 1 year","Full Dashboard"),
    ("GROWTH","PAT CAGR 1Y %","Profit After Tax CAGR over 1 year","Full Dashboard"),
    ("GROWTH","PAT YoY %","PAT growth year-over-year","All sheets"),
    ("FIN HEALTH","D/E Ratio",
     "Debt ÷ Equity. "
     "0 = Zero debt (fortress balance sheet) | "
     "0–0.3 = Very low debt (conservative, safe) | "
     "0.3–1.0 = Moderate (acceptable for most sectors) | "
     "1.0–2.0 = High (acceptable only for banking/NBFC/infra sectors) | "
     "> 2.0 = Very high (risk of default, avoid unless sector-specific justification) | "
     "Note: Banks naturally run higher D/E (deposits = debt)","All sheets"),
    ("FIN HEALTH","Current Ratio",
     "Current Assets ÷ Current Liabilities — short-term liquidity. "
     "< 1.0 = Danger (cannot cover near-term obligations) | "
     "1.0–1.5 = Tight (manage carefully) | "
     "1.5–2.0 = Healthy (comfortable) | "
     "> 2.0 = Strong (ample short-term liquidity)","Full Dashboard"),
    ("FIN HEALTH","FCF (₹Cr)","Free Cash Flow = Operating CF − Capex","Full Dashboard"),
    ("SHAREHOLDING","Promoter %","Promoter holding. <25% = low conviction. 0% = hard drop","All sheets"),
    ("SHAREHOLDING","Pledge %","Pledged shares. >20% = anti-trigger fires. >40% = hard drop","All sheets"),
    ("SHAREHOLDING","FII %","Foreign Institutional Investor holding. Rising = smart money signal","Full Dashboard"),
    ("QUALITY SCORES","Piotroski F /9","9-point health score. ≥7 = strong. ≤3 = weak","All sheets"),
    ("QUALITY SCORES","Altman Z","Bankruptcy predictor. >2.99 safe. <1.81 = distress zone","Full Dashboard"),
    ("QUALITY SCORES","Beneish M","Manipulation score. >-2.22 = risk. Anti-trigger fires","Full Dashboard"),
    ("PIPELINE / OB","OB/Bill Ratio","Order Book ÷ Revenue. >1.5× = strong pipeline","All sheets"),
    ("EARLY DETECTION","Early Entry /100","12 signals: quiet accum, SME migration, analyst imminent, sector Stage 1","All sheets"),
    ("EARLY DETECTION","Sector Stage",
     "Stage 1 = Early Accumulation (sector just starting to move, best entry point, low risk/reward) | "
     "Stage 2 = Confirmed Uptrend (momentum building, institutional buying, good entry) | "
     "Stage 3 = Peak / Euphoria (sector fully priced, risk of reversal, tighten stops) | "
     "Stage 4 = Distribution / Decline (smart money exiting, avoid fresh entry)","All sheets"),
    ("TECHNICAL","SMA 200",
     "200-Day Simple Moving Average — long-term trend indicator. "
     "ABOVE = CMP above 200 SMA, stock in long-term uptrend (bullish) | "
     "BELOW = CMP below 200 SMA, stock in long-term downtrend (bearish) | "
     "Golden rule: only buy stocks ABOVE their 200 SMA","Full Dashboard"),
    ("TECHNICAL","Supertrend",
     "ATR-based trend-following indicator. "
     "BUY = price above supertrend line (uptrend confirmed, go long) | "
     "SELL = price below supertrend line (downtrend, exit or avoid) | "
     "NEUTRAL = indeterminate / sideways market","Full Dashboard"),
    ("TECHNICAL","MACD Signal",
     "Moving Average Convergence Divergence. "
     "BUY = MACD line crossed above signal line (bullish momentum) | "
     "SELL = MACD line crossed below signal line (bearish momentum) | "
     "NEUTRAL = no recent crossover","Full Dashboard"),
    ("TECHNICAL","RSI (14)",
     "Relative Strength Index, 0–100. "
     "< 30 = Oversold (potential reversal up, look for buy signal) | "
     "30–50 = Bearish zone (weak, avoid fresh entry) | "
     "50–60 = Neutral (no strong signal) | "
     "60–70 = Bullish zone (momentum building, good entry on dips) | "
     "> 70 = Overbought (potential reversal down, wait for pullback) | "
     "Best entry: RSI 50–65 with MACD BUY and Supertrend BUY","Full Dashboard"),
    ("TECHNICAL","ADX",
     "Average Directional Index — measures trend strength (0–100), not direction. "
     "< 20 = Weak/no trend (sideways, avoid trend strategies) | "
     "20–25 = Emerging trend (watch closely) | "
     "25–40 = Strong trend (good for momentum entry) | "
     "> 40 = Very strong trend (ride with trailing stop)","Full Dashboard"),
    ("TECHNICAL","OBV Signal",
     "On-Balance Volume — tracks smart money flow. "
     "ACCUMULATION = OBV rising even when price flat (institutions buying quietly, bullish) | "
     "DISTRIBUTION = OBV falling while price holds up (institutions selling, bearish) | "
     "NEUTRAL = no clear pattern","Full Dashboard"),
    ("TECHNICAL","Above VWAP",
     "Volume Weighted Average Price — intraday fair value benchmark. "
     "YES = CMP above VWAP (buyers in control, bullish intraday) | "
     "NO = CMP below VWAP (sellers in control, wait for VWAP reclaim)","Full Dashboard"),
    ("BALANCE SHEET","BS Health Flag",
     "HEALTHY = cash > debt, D/E < 1.0, no pledge risk, FCF positive (safe to invest) | "
     "WATCH = moderate debt or pledge 10–20%, monitor quarterly results | "
     "ALERT = high debt (D/E>2), pledge>20%, negative FCF, or interest coverage<1.5 (extra caution)","All sheets"),
    ("TRADE PLAN","R:R Ratio",
     "Risk:Reward = (Target1 − Entry) ÷ (Entry − Stop Loss). "
     "< 1:1 = Poor (risk more than you can gain, avoid) | "
     "1:1 to 2:1 = Acceptable (only for high-conviction BUY with strong MoS) | "
     "2:1 to 3:1 = Good (standard for positional trades) | "
     "> 3:1 = Excellent (ideal setup, asymmetric payoff) | "
     "Rule: never enter a trade with R:R below 1.5:1","Trade Summary"),
    ("TRADE PLAN","Time Horizon",
     "SWING = 5–15 trading days (short-term momentum play) | "
     "POSITIONAL = 1–3 months (medium-term trend follow) | "
     "INVESTMENT = 6–18 months (fundamental re-rating play)","Trade Summary"),
    ("TRADE PLAN","Risk Level",
     "LOW = large cap, low beta, positive MoS, strong balance sheet | "
     "MEDIUM = mid cap or slight premium or moderate debt | "
     "HIGH = small/micro cap or negative MoS or high D/E or high beta","Trade Summary"),
    ("ANALYSIS SUMMARY","View Analysis Summary","150–250 word AI note with exact ₹ figures, catalysts, risks","All sheets"),

    # ── IDENTITY ──────────────────────────────────────────────────────────────
    ("IDENTITY","Company Name","Full registered company name on NSE/BSE","All sheets"),
    ("IDENTITY","Sector",
     "Sector classified by NSE/BSE/yfinance. Used for sector-median PE and sector rotation stage. "
     "Examples: Information Technology, Financial Services, Automobiles, Pharmaceuticals, Energy","All sheets"),

    # ── PRICE & MARKET ────────────────────────────────────────────────────────
    ("PRICE & MARKET","CMP",
     "Current Market Price — closing price from today's bhav copy (NSE/BSE). "
     "All valuations and ratios are computed using this price.","All sheets"),
    ("PRICE & MARKET","Day Chg %",
     "Price change % today vs yesterday's close. "
     ">2% with high delivery = strong buying signal | <-2% = watch for panic or news","All sheets"),
    ("PRICE & MARKET","52W High",
     "Highest closing price in the past 52 weeks. "
     "CMP within 10% of 52W High = stock near peak, caution. CMP near 52W Low = potential value opportunity.","Full Dashboard"),
    ("PRICE & MARKET","52W Low",
     "Lowest closing price in the past 52 weeks. "
     "CMP at 52W Low with good fundamentals = potential deep value entry.","Full Dashboard"),

    # ── FAIR VALUE ────────────────────────────────────────────────────────────
    ("FAIR VALUE","CFV (₹)",
     "Composite Fair Value — sector-weighted average of all 7 models (M1–M7) that return a valid value. "
     "CFV > CMP = undervalued. CFV < CMP = overvalued.","All sheets"),
    ("FAIR VALUE","FV Low (₹)",
     "Bear-case fair value = CFV × 0.85 (15% margin of safety buffer). "
     "If CMP < FV Low, stock is deeply undervalued even in a pessimistic scenario.","Full Dashboard"),
    ("FAIR VALUE","FV High (₹)",
     "Bull-case fair value = CFV × 1.15. "
     "If CMP > FV High, stock is overvalued even optimistically.","Full Dashboard"),
    ("FAIR VALUE","Upside to FV %",
     "(CFV − CMP) / CMP × 100. % gain if CMP reaches fair value. "
     "Positive = upside potential | Negative = downside risk. Same formula as MoS%.","All sheets"),
    ("FAIR VALUE","MoS Label",
     "Text label for MoS %: "
     "EXCEPTIONAL VALUE (>40%) | STRONG VALUE (>25%) | GOOD VALUE (>10%) | "
     "FAIR VALUE (0–10%) | SLIGHT PREMIUM (0 to -15%) | OVERVALUED (-15% to -30%) | "
     "SIGNIFICANTLY OVERVALUED (<-30%)","All sheets"),
    ("FAIR VALUE","M1: DCF FV (₹)","3-Stage Discounted Cash Flow. WACC = 10Y GSec + Beta×5.5%. Terminal growth 4.5%. Best for steady compounders.","Full Dashboard"),
    ("FAIR VALUE","M2: Graham FV (₹)","Graham Number = √(22.5 × EPS × BVPS). Benjamin Graham's intrinsic value formula. Best for value stocks with positive EPS.","Full Dashboard"),
    ("FAIR VALUE","M3: PE FV (₹)","EPS × Sector 5yr median P/E. Mean-reversion model — assumes P/E reverts to sector norm.","Full Dashboard"),
    ("FAIR VALUE","M4: PB FV (₹)","BVPS × Sector median Price/Book. Best for asset-heavy sectors: banks, metals, real estate.","Full Dashboard"),
    ("FAIR VALUE","M5: EV FV (₹)","CMP × (Sector median EV/EBITDA ÷ Stock EV/EBITDA). Best for capital-intensive businesses.","Full Dashboard"),
    ("FAIR VALUE","M6: DDM FV (₹)",
     "Gordon Growth Model = DPS×(1+g)/(r−g). "
     "Only shown for dividend-paying stocks (yield 0.1%–15%). "
     "DPS = CMP × Div Yield. r = GSec + 4.5%. g = min(PAT growth/2, 6%).","Full Dashboard"),
    ("FAIR VALUE","M7: PEG FV (₹)","EPS × EPS Growth Rate. PEG=1 baseline. Best for high-growth stocks where PE alone overstates expensiveness.","Full Dashboard"),

    # ── VALUATION ─────────────────────────────────────────────────────────────
    ("VALUATION","P/B",
     "Price ÷ Book Value per Share. "
     "<1 = trading below book (possible deep value or value trap) | "
     "1–2 = reasonable (banks, metals) | 2–5 = premium (justified if ROE>15%) | "
     ">5 = high (asset-light compounders only)","All sheets"),
    ("VALUATION","P/S",
     "Price-to-Sales = MCap ÷ Annual Revenue. "
     "<1 = very cheap (cyclicals, PSUs) | 1–3 = reasonable | 3–8 = premium | "
     ">8 = expensive (only for high-margin businesses)","Full Dashboard"),
    ("VALUATION","P/CF",
     "Price-to-Cash Flow = MCap ÷ Operating Cash Flow. "
     "Better than P/E as cash flow is harder to manipulate. "
     "<10 = cheap | 10–20 = fair | 20–35 = premium | >35 = expensive. "
     "Derived as P/S ÷ EBITDA margin when direct cash flow data unavailable.","Full Dashboard"),
    ("VALUATION","EV/EBITDA",
     "Enterprise Value ÷ EBITDA. Ignores capital structure — better for comparing levered vs unlevered firms. "
     "<8 = cheap | 8–15 = fair | 15–25 = premium | >25 = expensive","Full Dashboard"),

    # ── PROFITABILITY ─────────────────────────────────────────────────────────
    ("PROFITABILITY","ROCE %",
     "Return on Capital Employed = EBIT ÷ Capital Employed × 100. "
     "Measures how efficiently the company uses ALL capital (debt + equity). "
     ">20% = excellent | 15–20% = good | 10–15% = average | <10% = poor. "
     "Derived from EBITDA margin × Revenue / Capital Employed when direct data unavailable.","Full Dashboard"),
    ("PROFITABILITY","ROA %",
     "Return on Assets = Net Income ÷ Total Assets × 100. "
     ">10% = excellent | 5–10% = good | <5% = poor. "
     "Derived as ROE ÷ (1 + D/E) when direct data unavailable.","Full Dashboard"),
    ("PROFITABILITY","Gross Mgn %",
     "Gross Profit ÷ Revenue × 100. Revenue minus direct costs (raw materials, COGS). "
     ">50% = high-margin business (software, pharma) | >30% = good | <15% = commodity/trading","Full Dashboard"),
    ("PROFITABILITY","EBITDA Mgn %",
     "EBITDA ÷ Revenue × 100. Operating profitability before interest, tax, depreciation. "
     ">25% = excellent | 15–25% = good | 8–15% = average | <8% = tight","Full Dashboard"),
    ("PROFITABILITY","NPM %",
     "Net Profit Margin = Net Income ÷ Revenue × 100. Bottom-line profitability after everything. "
     ">15% = excellent | 8–15% = good | 3–8% = average | <3% = thin (watch for debt servicing risk)","Full Dashboard"),
    ("PROFITABILITY","NPM Q1 %","Net Profit Margin in Q1 (Apr–Jun). Compare 3 quarters for Margin Expansion trend.","Full Dashboard"),
    ("PROFITABILITY","NPM Q2 %","Net Profit Margin in Q2 (Jul–Sep). Rising NPM Q1→Q2→Q3 = Margin Expansion flag.","Full Dashboard"),
    ("PROFITABILITY","NPM Q3 %","Net Profit Margin in Q3 (Oct–Dec). No free data source — requires BSE quarterly filings.","Full Dashboard"),
    ("PROFITABILITY","Margin Expansion",
     "YES = NPM has risen for 3 consecutive quarters (Q1→Q2→Q3). "
     "Strong signal of operational leverage or pricing power. No free data source.","Full Dashboard"),

    # ── GROWTH ────────────────────────────────────────────────────────────────
    ("GROWTH","Rev CAGR 3Y %","Revenue Compound Annual Growth Rate over 3 years. >15% = fast growing. No free source — requires multi-year financials.","Full Dashboard"),
    ("GROWTH","PAT CAGR 3Y %","Profit After Tax CAGR over 3 years. >15% = quality compounder. No free source.","Full Dashboard"),
    ("GROWTH","EBITDA CAGR 1Y %","EBITDA growth year-over-year. No free source — requires quarterly filings.","Full Dashboard"),
    ("GROWTH","Rev YoY %",
     "Revenue growth year-over-year (%) from yfinance revenueGrowth. "
     ">20% = fast growth | 10–20% = good | 0–10% = stable | <0% = shrinking","Full Dashboard"),
    ("GROWTH","Q3 Rev (₹Cr)","Revenue in Q3 (Oct–Dec) in ₹ Crore. No free data source.","Full Dashboard"),
    ("GROWTH","Q3 PAT (₹Cr)","Net Profit in Q3 in ₹ Crore. No free data source.","Full Dashboard"),
    ("GROWTH","Q3 EBITDA (₹Cr)","EBITDA in Q3 in ₹ Crore. No free data source.","Full Dashboard"),

    # ── FIN HEALTH ────────────────────────────────────────────────────────────
    ("FIN HEALTH","ND/EBITDA",
     "Net Debt ÷ EBITDA. How many years of earnings needed to repay net debt. "
     "<1× = very safe | 1–2× = manageable | 2–3× = stretched | >3× = high leverage. "
     "No free source — requires balance sheet detail.","Full Dashboard"),
    ("FIN HEALTH","Int Coverage",
     "EBIT ÷ Interest Expense. How many times earnings cover interest payments. "
     ">5× = safe | 3–5× = adequate | 1.5–3× = risky | <1.5× = danger. "
     "No free source — requires income statement detail.","Full Dashboard"),
    ("FIN HEALTH","Cash (₹Cr)",
     "Total cash and equivalents in ₹ Crore. "
     "High cash vs debt = fortress balance sheet. "
     "Cash > Total Debt = net cash company (very safe, often undervalued).","Full Dashboard"),
    ("FIN HEALTH","Total Debt (₹Cr)",
     "Total financial debt (short + long term) in ₹ Crore. "
     "0 = zero-debt company. Derived as D/E × Book Equity when direct data unavailable. "
     "Compare with Cash to get Net Debt.","Full Dashboard"),
    ("FIN HEALTH","FCF Yield %",
     "Free Cash Flow ÷ MCap × 100. How much FCF you get per ₹ invested. "
     ">5% = very good | 3–5% = good | 1–3% = modest | <1% = poor. "
     "Often proxy-computed from Operating Cash Flow when FCF not available.","Full Dashboard"),
    ("FIN HEALTH","CCC Days",
     "Cash Conversion Cycle = DIO + DSO − DPO. Days to convert inventory investment into cash. "
     "Lower = better. <30 days = excellent. No free source — requires AR/AP/inventory data.","Full Dashboard"),

    # ── CAP ALLOC ─────────────────────────────────────────────────────────────
    ("CAP ALLOC","Div Yield %",
     "Annual dividend ÷ CMP × 100. Income return on investment. "
     ">4% = high yield (income stock) | 1–4% = moderate | 0–1% = growth-focused | "
     "0% = no dividend (all profits reinvested). Normalised from yfinance fractions.","All sheets"),
    ("CAP ALLOC","Payout Ratio %",
     "Dividends ÷ Net Profit × 100. Percentage of earnings paid as dividends. "
     "<30% = growth company (retains earnings) | 30–60% = balanced | "
     ">70% = high payout (mature/PSU/FMCG). >100% = paying from reserves (unsustainable).","Full Dashboard"),
    ("CAP ALLOC","Capex / Rev %",
     "Capital Expenditure ÷ Revenue × 100. Reinvestment intensity. "
     ">15% = capital-heavy (infra, steel, telecom) | 3–15% = moderate | "
     "<3% = asset-light (IT, FMCG). No free source — requires cash flow statement.","Full Dashboard"),

    # ── SHAREHOLDING ──────────────────────────────────────────────────────────
    ("SHAREHOLDING","Pro QoQ Δ",
     "Promoter holding change quarter-over-quarter (%). "
     "Increasing = promoters buying → bullish signal. "
     "Decreasing = promoters selling → investigate reason. No free source (BSE filings only).","Full Dashboard"),
    ("SHAREHOLDING","Pledge Direction",
     "Direction of pledge change: INCREASING / DECREASING / STABLE. "
     "INCREASING pledge is a red flag — promoters may be under financial stress.","Full Dashboard"),
    ("SHAREHOLDING","DII %",
     "Domestic Institutional Investor holding %. "
     "DII (mutual funds, insurance) rising = domestic smart money accumulating. "
     "Cannot be separated from FII in yfinance — shown as combined institutional %.","Full Dashboard"),
    ("SHAREHOLDING","DII QoQ Δ","DII holding change quarter-over-quarter. No free source.","Full Dashboard"),
    ("SHAREHOLDING","FII QoQ Δ","FII holding change quarter-over-quarter. No free source.","Full Dashboard"),
    ("SHAREHOLDING","Public Float %",
     "% of shares held by retail/public (100% − Promoter% − Institutional%). "
     "Higher float = more liquid, lower impact cost for large orders.","Full Dashboard"),

    # ── QUALITY SCORES ────────────────────────────────────────────────────────
    ("QUALITY SCORES","Earn Quality",
     "Qualitative assessment: HIGH / MEDIUM / LOW. "
     "Checks if earnings are backed by cash flow (FCF/PAT ratio). "
     "LOW = earnings not converting to cash (red flag for manipulation).","Full Dashboard"),

    # ── PIPELINE / OB ─────────────────────────────────────────────────────────
    ("PIPELINE / OB","Pipeline Vis",
     "Pipeline Visibility: HIGH / MEDIUM / LOW / NONE. "
     "For defence, infra, capital goods stocks — is the revenue pipeline visible for next 12–24 months?","Full Dashboard"),
    ("PIPELINE / OB","L1 Wins 90D",
     "Number of L1 (lowest bid) wins in government tenders in the last 90 days. "
     "Only relevant for defence, EPC, infra companies. Source: CPP portal / BSE announcements.","Full Dashboard"),
    ("PIPELINE / OB","L1 Est (₹Cr)",
     "Estimated value of L1 wins in ₹ Crore over last 90 days. "
     "Compare with annual revenue for order-book visibility.","Full Dashboard"),
    ("PIPELINE / OB","New Mkt Entry",
     "Has the company announced entry into a new market/geography/product in the last quarter? "
     "YES = potential re-rating catalyst.","Full Dashboard"),

    # ── EARLY DETECTION ───────────────────────────────────────────────────────
    ("EARLY DETECTION","Early Signals",
     "Pipe-separated list of active early-mover triggers detected. "
     "Examples: INSTITUTIONAL FOOTPRINT | VOL SURGE + RSI ACCUMULATION | TREND CONFLUENCE | "
     "MOMENTUM BUILDING | DUAL-LISTED DISCOVERY","Full Dashboard"),
    ("EARLY DETECTION","Smart Money",
     "ACCUMULATION = FII+DII buying trend detected (3Q rising) | "
     "DISTRIBUTION = Institutional selling | NEUTRAL = no clear trend. "
     "Signals smart money positioning ahead of price move.","Full Dashboard"),

    # ── TECHNICAL ─────────────────────────────────────────────────────────────
    ("TECHNICAL","Stoch %K",
     "Stochastic Oscillator %K — momentum indicator comparing closing price to price range. "
     "<20 = Oversold (potential reversal up) | 20–40 = Bearish zone | "
     "40–60 = Neutral | 60–80 = Bullish zone | >80 = Overbought (potential reversal down). "
     "Best signal: %K crossing above 20 from below = buy; crossing below 80 from above = sell.","Full Dashboard"),
    ("TECHNICAL","MFI",
     "Money Flow Index — volume-weighted RSI. Uses both price AND volume. "
     "<20 = Oversold (strong reversal signal, especially with rising volume) | "
     ">80 = Overbought | 40–60 = Neutral. "
     "MFI divergence from price = early warning of trend reversal.","Full Dashboard"),
    ("TECHNICAL","Chart Pattern",
     "Most recent chart pattern detected: "
     "BULLISH CANDLE / BEARISH CANDLE / DOJI (indecision) / "
     "HAMMER (bullish reversal) / SHOOTING STAR (bearish reversal) / "
     "ENGULFING BULLISH / ENGULFING BEARISH. "
     "Candlestick patterns are more reliable when they occur near key support/resistance levels.","Full Dashboard"),
    ("TECHNICAL","Support 1 (₹)",
     "Nearest support level below CMP — price where buying is historically strong. "
     "Stop loss should be placed slightly below Support 1.","Full Dashboard"),
    ("TECHNICAL","Support 2 (₹)",
     "Second support level — deeper pullback zone. "
     "If price breaks Support 1 with volume, next target is Support 2.","Full Dashboard"),
    ("TECHNICAL","Resist 1 (₹)",
     "Nearest resistance level above CMP — price where selling pressure is historically strong. "
     "Target 1 is typically set at Resist 1.","Full Dashboard"),
    ("TECHNICAL","Resist 2 (₹)",
     "Second resistance level. Target 2 / Target 3 set at Resist 2.","Full Dashboard"),

    # ── BALANCE SHEET ─────────────────────────────────────────────────────────
    ("BALANCE SHEET","BS Health Note",
     "Detailed note explaining BS Health Flag. "
     "Examples: 'No red flags detected' | 'High D/E 2.1x — monitor debt serviceability' | "
     "'Pledge 23% — elevated risk'","Full Dashboard"),

    # ── TRADE PLAN ────────────────────────────────────────────────────────────
    ("TRADE PLAN","Entry Range (₹)",
     "Suggested entry price band: Low–High in ₹. "
     "Based on CMP ± 1.5% to account for intraday spread. "
     "Use limit orders within this range for best execution.","All sheets"),
    ("TRADE PLAN","Stop Loss (₹)",
     "Mandatory stop loss price. Exit the trade if CMP closes below this on a daily basis. "
     "Set at ~3–5% below entry, or just below key support level. "
     "ALWAYS place stop loss before entering any trade.","All sheets"),
    ("TRADE PLAN","Target 1 (₹)",
     "First price target = nearest resistance level. "
     "Book 40–50% of position at T1. Let rest run with trailing stop.","All sheets"),
    ("TRADE PLAN","Target 2 (₹)",
     "Second price target = next resistance zone. "
     "Book another 30% at T2. Trail stop to entry cost.","All sheets"),
    ("TRADE PLAN","Target 3 (₹)",
     "Final price target = CFV (Composite Fair Value). "
     "Hold remaining 20–30% position for full re-rating. "
     "Only applicable for BUY stocks with positive MoS.","All sheets"),

    # ── NEWS & RISK ───────────────────────────────────────────────────────────
    ("NEWS & RISK","Key Catalyst",
     "Primary upcoming event that could trigger price re-rating. "
     "Examples: order win, product launch, QIP, promoter buyback, index inclusion. "
     "Requires Anthropic API credits — populated by AI analyst.","Full Dashboard"),
    ("NEWS & RISK","News Sentiment",
     "AI-assessed recent news tone: BULLISH / NEUTRAL / BEARISH. "
     "Based on last 30 days of BSE announcements and news. "
     "Requires Anthropic API credits.","Full Dashboard"),
    ("NEWS & RISK","Primary Risk",
     "The single most important risk factor for this stock right now. "
     "Examples: regulatory overhang, promoter pledge, client concentration, commodity exposure. "
     "Requires Anthropic API credits.","Full Dashboard"),
    ("NEWS & RISK","SEBI Flags",
     "Any active SEBI actions, adjudication orders, or exchange surveillance flags. "
     "NONE = clean | Any other value = investigate before investing. "
     "Requires Anthropic API credits.","Full Dashboard"),

    # ── GOLD SHEET SPECIFIC ────────────────────────────────────────────────────
    ("SCORES","F-Score /9",
     "Proxy Financial Health Score (0–9) computed from available data. "
     "P1: ROA>0 | P2: FCF>0 | P3: PAT YoY>0 | P4: D/E<1 | P5: Current Ratio>1 | "
     "P6: Gross Margin>15% | P7: Rev YoY>0 | P8: ROE>10% | P9: Cash>0. "
     "≥7=Strong | 4–6=Average | ≤3=Weak | —=Insufficient data.","Gold Sheet"),
    ("SCORES","Spike /6",
     "Spike Score 0–6: count of institutional triggers fired. "
     "Triggers: Value Breakout | Perfect Storm | Institutional Accumulation | "
     "Technical Breakout | RSI Accumulation | Dual-Listed Discovery. "
     "≥3=Strong setup | 1–2=Watch | 0=No spike.","All sheets"),
    ("SCORES","Storm /10",
     "Storm Score 0–10: defensive quality in volatile markets. "
     "Rewards low beta, low D/E, positive FCF, high promoter holding, "
     "low pledge and strong cash cover. Higher = safer defensive pick.","All sheets"),
    ("TRADE PLAN","Horizon",
     "Abbreviated Time Horizon used in Gold sheet. "
     "SHORT TERM = 2–4 weeks (spike-driven entry) | "
     "POSITIONAL = 1–3 months (trend confirmed) | "
     "LONG TERM = 6–12 months (value accumulation).","Gold Sheet"),
    ("VALUATION","P/E",
     "Price-to-Earnings ratio (abbreviated in Gold sheet). "
     "CMP ÷ Trailing 12-month EPS. "
     "<15=Cheap | 15–25=Fair | 25–40=Premium | >40=Expensive.","Gold Sheet"),
    ("VALUATION","PEG",
     "PEG Ratio (abbreviated). P/E ÷ Earnings Growth Rate. "
     "<1=Undervalued vs growth | 1–2=Fair | >2=Expensive vs growth.","Gold Sheet"),
    ("VALUATION","D/E",
     "Debt-to-Equity ratio (abbreviated in Gold sheet). "
     "<0.5=Low debt | 0.5–1.5=Moderate | >2=High leverage risk.","Gold Sheet"),
    ("TECHNICAL","Pattern",
     "Latest candlestick pattern. Same as Chart Pattern in Full Dashboard. "
     "BULLISH/BEARISH CANDLE | HAMMER | SHOOTING STAR | DOJI | ENGULFING.","Gold Sheet"),
    ("FIN HEALTH","Quick Ratio",
     "Quick Ratio = (Current Assets − Inventory) ÷ Current Liabilities. "
     "More conservative than Current Ratio — excludes slow inventory. "
     ">1=Can meet short-term obligations | <1=Potential liquidity risk.","Full Dashboard"),

    # ── PRICE & MARKET ─────────────────────────────────────────────────────────
    ("PRICE & MARKET","CMP (₹)",
     "Current Market Price — closing price from today's bhav copy (NSE/BSE). "
     "All FV models and ratio calculations use this as the base price.","All sheets"),
    ("PRICE & MARKET","52W High (₹)",
     "Highest closing price in the past 52 weeks. "
     "CMP within 5% of 52W High = near resistance peak, caution on entry. "
     "Breakout above 52W High with high volume = very strong signal.","Full Dashboard"),
    ("PRICE & MARKET","52W Low (₹)",
     "Lowest closing price in the past 52 weeks. "
     "CMP near 52W Low + good fundamentals = potential deep value opportunity.","Full Dashboard"),

    # ── WEEKLY CHANGE ──────────────────────────────────────────────────────────
    ("WEEKLY CHANGE %","Chg% [2-Wk]",
     "Price return over last 2 weeks (10 trading days). "
     "2-Wk rising faster than 4-Wk = accelerating momentum — bullish signal.","All sheets"),
    ("WEEKLY CHANGE %","Chg% [4-Wk]",
     "Price return over last 4 weeks (20 trading days). "
     "Used in Sector Stage scoring and Early Entry detection. "
     ">5%=Strong uptrend | <-5%=Downtrend caution.","All sheets"),
    ("WEEKLY CHANGE %","Chg% [6-Wk]",
     "Price return over last 6 weeks. Medium-term trend confirmation.","Full Dashboard"),
    ("WEEKLY CHANGE %","Chg% [8-Wk]",
     "Price return over last 8 weeks. Confirms whether trend is sustained.","Full Dashboard"),

    # ── FAIR VALUE ─────────────────────────────────────────────────────────────
    ("FAIR VALUE","Upside %",
     "Upside to Fair Value = (CFV − CMP) ÷ CMP × 100. "
     "Same as MoS % in Full Dashboard. "
     "Positive=upside potential | Negative=currently priced above fair value.","Gold Sheet"),
]

GRP_COLORS = {
    "IDENTITY":"1E293B","SCORES":"7C3AED","PRICE & MARKET":"0369A1",
    "WEEKLY CHANGE %":"0F766E","FAIR VALUE":"B45309","VALUATION":"0891B2",
    "PROFITABILITY":"059669","GROWTH":"047857","FIN HEALTH":"DC2626",
    "CAP ALLOC":"6D28D9","SHAREHOLDING":"EA580C","QUALITY SCORES":"0D9488",
    "PIPELINE / OB":"1D4ED8","EARLY DETECTION":"B45309","TECHNICAL":"6D28D9",
    "BALANCE SHEET":"D97706","TRADE PLAN":"059669","NEWS & RISK":"475569",
    "ANALYSIS SUMMARY":"0F172A",
}

# Columns permanently blank — no free data source available
# These are highlighted with bold red headers so user knows at a glance
NO_FREE_SOURCE_COLS = {
    "Rev CAGR 1Y %","Rev CAGR 3Y %","PAT CAGR 1Y %","PAT CAGR 3Y %",
    "EBITDA CAGR 1Y %","Q3 Rev (₹Cr)","Q3 PAT (₹Cr)","Q3 EBITDA (₹Cr)",
    "ND/EBITDA","Int Coverage","CCC Days","Capex / Rev %",
    "Pro QoQ Δ","Pledge %","Pledge Direction","DII %","DII QoQ Δ",
    "FII QoQ Δ","Public Float %",
    "Piotroski F /9","Altman Z","Beneish M","Earn Quality",
    "OB/Bill Ratio","Pipeline Vis","L1 Wins 90D","L1 Est (₹Cr)","New Mkt Entry",
    "Key Catalyst","News Sentiment","Primary Risk","SEBI Flags",
    "NPM Q1 %","NPM Q2 %","NPM Q3 %","Margin Expansion",
}
# Needs Anthropic API credits — amber highlight
NEEDS_AI_CREDITS = {"View Analysis Summary"}

def _sf(val, default=0.0):
    if val is None or val == "" or str(val) in ("—", "--", "N/A"):
        return float(default)
    try:
        return float(val)
    except (ValueError, TypeError):
        return float(default)


def _f(h): return PatternFill("solid", fgColor=h)

# Thin border on all 4 sides — used for header rows and data cells
_THIN = Side(style="thin", color="FFFFFF")       # white border between same-section cols
_SECT = Side(style="medium", color="FFFFFF")      # medium white border between sections
_BDR  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BDR_SECT = Border(left=_SECT, right=_SECT, top=_THIN, bottom=_THIN)

def _border(is_section_edge=False):
    return _BDR_SECT if is_section_edge else _BDR
def _ft(bold=False,color=NAVY,size=9,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def _al(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _sv(v):
    if v is None: return None
    if isinstance(v,(str,int,float,bool)): return v
    if isinstance(v,(list,tuple)): return " | ".join(str(x) for x in v) if v else ""
    if isinstance(v,dict): return str(v)
    return str(v)
def _g(s,k,d="—"):
    v=s.get(k,d); return _sv(v) if v is not None else d

def _alt(bg):
    return {"FAC775":"FEF3C7","D1FAE5":"ECFDF5","DBEAFE":"EFF6FF",
            "FEE2E2":"FEF2F2","FEF3C7":"FFFBEB"}.get(bg,bg)

class ExcelGeneratorV6:
    REQUIRED_COLS = {
        "early_entry_score":0,"composite_score":0,"spike_count":0,"storm_score":0,
        "mos_pct":0.0,"upside":0.0,"verdict":"WATCHLIST","spike_suppressed":False,
        "symbol":"","close":0.0,"company_name":"","sector":"General",
        "exchange_tag":"NSE","Analysis_Summary_Block_H":"—",
        "2w_chg":0,"4w_chg":0,"6w_chg":0,"8w_chg":0,"cfv":0.0,
        "entry_range":"—","stop_loss":"—","t1":0,"t2":0,"t3":0,
        "horizon":"POSITIONAL","risk_level":"MEDIUM","mos_label":"—",
        "smart_money_signals":"","rotation_stage":"NEUTRAL","vol_ratio":1.0,
        "bs_status":"HEALTHY","bs_flags":"—",
    }

    def __init__(self,data,date_str,run_time=None,prev_scores=None):
        self.df=pd.DataFrame(data) if data else pd.DataFrame()
        self.date_str=date_str
        try: self.dlbl=datetime.strptime(date_str,"%Y%m%d").strftime("%-d %b %Y")
        except: self.dlbl=date_str
        # Actual pipeline run time in IST — used in Alert Log and Delivery Preview
        if run_time:
            self.run_time = run_time
        else:
            try:
                import pytz as _ptz2
                _ist2 = _ptz2.timezone("Asia/Kolkata")
                self.run_time = datetime.now(_ist2).strftime("%H:%M IST")
            except Exception:
                self.run_time = "—"
        # Previous day scores for Score Δ computation in Alert Log
        self.prev_scores = prev_scores or {}
        for col,dflt in self.REQUIRED_COLS.items():
            if col not in self.df.columns: self.df[col]=dflt

        # ── Filter NEUTRAL stocks: only keep if exceptionally good ──────────
        # NEUTRAL = score doesn't qualify for BUY/WATCHLIST
        # Keep NEUTRAL only if: ROE>20% AND PE<30 AND MoS>10% AND ts>65
        # i.e. strong fundamentals + undervalued + decent technicals
        if not self.df.empty and "verdict" in self.df.columns:
            def _is_exceptional_neutral(row):
                if str(row.get("verdict","")) != "NEUTRAL":
                    return True  # keep all non-neutral
                roe = float(row.get("roe_num", row.get("roe", 0)) or 0)
                pe  = float(row.get("pe_num",  row.get("pe",  99)) or 99)
                mos = float(row.get("mos_pct", 0) or 0)
                ts  = float(row.get("technical_score", 50) or 50)
                sc  = float(row.get("composite_score", 0) or 0)
                # Exceptional NEUTRAL: strong fundamentals AND undervalued AND good technicals
                exceptional = (roe > 20 and pe < 30 and mos > 10 and ts > 62)
                if exceptional:
                    return True
                return False

            self.df = self.df[self.df.apply(_is_exceptional_neutral, axis=1)].reset_index(drop=True)

    @staticmethod
    def _safe_val(v): return _sv(v)

    def generate_excel_reports(self):
        """Generates a single Excel file with all 6 sheets.
        Sheet 2 (Gold – Early Movers) is embedded inside the Full Dashboard file.
        No separate Gold file is produced — avoids duplication.
        Returns (master_file, None) — None is filtered out by master_funnel
        so the xlsx appears only once in the email attachment list.
        """
        master = self._full()
        return master, None

    def _full(self):
        wb=Workbook(); wb.active.title="📊 Full Dashboard"
        self._full_sheet(wb.active)
        self._gold_sheet(wb)
        self._trade_summary(wb)
        self._alert_log(wb)
        self._delivery_preview(wb)
        self._glossary(wb)
        for ws in wb.worksheets: ws.sheet_view.showGridLines=False
        fn=f"NSE_BSE_Full_Dashboard_{self.date_str}.xlsx"; wb.save(fn); return fn

    def _gold_file(self):
        """Kept for backward compatibility — returns master file path."""
        return f"NSE_BSE_Full_Dashboard_{self.date_str}.xlsx"

    def _full_sheet(self,ws):
        N=len(FULL_COLS)
        # R1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
        c=ws.cell(1,1,f"NSE / BSE STOCK ANALYSER  ·  FULL RESEARCH DASHBOARD  ·  v6.0  ·  {self.dlbl}")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,13); c.alignment=_al()
        ws.row_dimensions[1].height=34
        # R2
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=N)
        c2=ws.cell(2,1,"AutoFilter (row 4): Exchange · Cap Category · Sector · Verdict · MoS Label · BS Flag · Risk · Storm · Sector Stage · Weekly Change   |  Last column = 'View Analysis Summary' — scroll right to see full AI reasoning with recent company facts   |  GOLD=Early Mover · GREEN=Deep Value · BLUE=Buy · AMBER=Watch · RED=Avoid   |  RED column header = No free data source available (requires paid API / BSE filings). AMBER header = Needs Anthropic API credits.")
        c2.fill=_f(LG); c2.font=_ft(False,"475569",8,True); c2.alignment=_al("left","center")
        ws.row_dimensions[2].height=16
        # R3 groups
        ws.row_dimensions[3].height=20
        for sc,nm,col,sp in FULL_GROUPS:
            ec=sc+sp-1
            if sp>1: ws.merge_cells(start_row=3,start_column=sc,end_row=3,end_column=ec)
            c=ws.cell(3,sc,nm); c.fill=_f(col); c.font=_ft(True,WHITE,8); c.alignment=_al()
            c.border=_border(is_section_edge=True)
        # R4 headers
        # Build col→group_color map so each header cell matches its section
        _col_color = {}
        for _sc,_nm,_col,_sp in FULL_GROUPS:
            for _ci in range(_sc, _sc+_sp):
                _col_color[_ci] = _col

        ws.row_dimensions[4].height=40
        for i,(h,w,_) in enumerate(FULL_COLS,1):
            ws.column_dimensions[get_column_letter(i)].width=w
            hdr_bg = _col_color.get(i, NAVY)
            if h in NO_FREE_SOURCE_COLS:
                # Bold red — permanently blank, no free data source
                c=ws.cell(4,i,h); c.fill=_f("991B1B"); c.font=_ft(True,"FEE2E2",8)
                c.alignment=_al("center","center",True); c.border=_border()
            elif h in NEEDS_AI_CREDITS:
                # Amber — populated only when Anthropic API credits are loaded
                c=ws.cell(4,i,h); c.fill=_f("92400E"); c.font=_ft(True,"FEF3C7",8)
                c.alignment=_al("center","center",True); c.border=_border()
            else:
                c=ws.cell(4,i,h); c.fill=_f(hdr_bg); c.font=_ft(True,WHITE,8)
                c.alignment=_al("center","center",True); c.border=_border()
        ws.freeze_panes="A5"
        ws.auto_filter.ref=f"A4:{get_column_letter(N)}4"
        # Data
        # FV model keys that show "—" when value is 0 (model not applicable)
        FV_MODEL_KEYS = {"M1_DCF","M2_Graham","M3_PE","M4_PB","M5_EV","M6_DDM","M7_PEG",
                         "cfv","cfv_low","cfv_high"}

        stks=self.df.to_dict("records")
        for ri,stk in enumerate(stks):
            rn=ri+5; ws.row_dimensions[rn].height=20
            verd=str(stk.get("verdict","WATCHLIST"))
            st=VERDICT_STYLES.get(verd,_DS)
            bg,tx=st["bg"],st["text"]
            ubg=bg if ri%2==0 else _alt(bg)
            for ci,(_,_,key) in enumerate(FULL_COLS,1):
                val=_g(stk,key)
                # FV models: 0 means "not applicable" → show "—" not 0
                if key in FV_MODEL_KEYS and (val == 0 or val == 0.0):
                    val = "—"
                cell=ws.cell(rn,ci,val); cell.fill=_f(ubg); cell.font=_ft(False,tx,9)
                wrap_cols={2,3,7,28,94,96,112,120,122,124}
                cell.alignment=_al("left" if ci in wrap_cols else "center","center",wrap=(ci==N))
                cell.border=Border(
                    left=Side(style="thin",color="E2E8F0"),
                    right=Side(style="thin",color="E2E8F0"),
                    top=Side(style="thin",color="E2E8F0"),
                    bottom=Side(style="thin",color="E2E8F0")
                )
        # Cond format weekly changes
        lr=4+len(stks)
        if lr>4:
            for col_num in [19,20,21,22]:
                cl=get_column_letter(col_num)
                ws.conditional_formatting.add(f"{cl}5:{cl}{lr}",
                    ColorScaleRule(start_type="min",start_color="FEE2E2",
                    mid_type="num",mid_value=0,mid_color="FFFFFF",
                    end_type="max",end_color="D1FAE5"))

    def _gold_sheet(self,wb):
        ws=wb.create_sheet("⭐ Gold – Early Movers"); self._gold_ws(ws)

    def _gold_ws(self,ws):
        gdf=self._get_gold(); N=len(GOLD_COLS)
        gc=len(gdf)
        ae=gdf["early_entry_score"].mean() if not gdf.empty else 0
        au=gdf["upside"].mean()            if not gdf.empty else 0
        asp=gdf["spike_count"].mean()      if not gdf.empty else 0
        # R1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
        c=ws.cell(1,1,f"⭐  NSE / BSE  —  GOLD CATEGORY  ·  EARLY MOVER STOCKS  ·  Highest Upside Before Consensus  ·  {self.dlbl}")
        c.fill=_f("B45309"); c.font=_ft(True,WHITE,13); c.alignment=_al()
        ws.row_dimensions[1].height=30
        # R2
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=N)
        c2=ws.cell(2,1,f"Criteria: Early Entry Score ≥70 OR (MoS ≥25% + Score ≥70)  ·  {gc} stocks qualify  ·  No AVOID stocks  ·  Spike Suppressed = FALSE")
        c2.fill=_f("FEF3C7"); c2.font=_ft(False,"92400E",8,True); c2.alignment=_al("left","center")
        ws.row_dimensions[2].height=14
        # R3 summary strip
        strips=[(1,5,f"⭐ GOLD STOCKS: {gc}"),(6,11,f"AVG EARLY SCORE: {ae:.0f}/100"),
                (12,17,f"AVG UPSIDE TO FV: +{au:.1f}%"),(18,23,f"AVG SPIKE SCORE: {asp:.1f}/6"),
                (24,N,f"DELIVERED: 20:30 IST · WhatsApp + Email")]
        ws.row_dimensions[3].height=22
        for s,e,t in strips:
            if s!=e: ws.merge_cells(start_row=3,start_column=s,end_row=3,end_column=e)
            c=ws.cell(3,s,t); c.fill=_f("B45309"); c.font=_ft(True,WHITE,9); c.alignment=_al()
        # R4 groups
        ws.row_dimensions[4].height=16
        for sc,nm,col,sp in GOLD_GROUPS:
            ec=sc+sp-1
            if sp>1: ws.merge_cells(start_row=4,start_column=sc,end_row=4,end_column=ec)
            c=ws.cell(4,sc,nm); c.fill=_f(col); c.font=_ft(True,WHITE,8); c.alignment=_al()
        # R5 headers
        # Build col→group_color map for gold headers
        _gcol_color = {}
        for _sc,_nm,_col,_sp in GOLD_GROUPS:
            for _ci in range(_sc, _sc+_sp):
                _gcol_color[_ci] = _col

        ws.row_dimensions[5].height=38
        for i,(h,w,_) in enumerate(GOLD_COLS,1):
            ws.column_dimensions[get_column_letter(i)].width=w
            hdr_bg = _gcol_color.get(i, "92400E")
            if h in NO_FREE_SOURCE_COLS:
                c=ws.cell(5,i,h); c.fill=_f("991B1B"); c.font=_ft(True,"FEE2E2",8)
            else:
                c=ws.cell(5,i,h); c.fill=_f(hdr_bg); c.font=_ft(True,WHITE,8)
            c.border=_border()
            c.alignment=_al("center","center",True)
        ws.freeze_panes="A6"
        ws.auto_filter.ref=f"A5:{get_column_letter(N)}5"
        # Data
        for ri,stk in enumerate(gdf.to_dict("records")):
            rn=ri+6; ws.row_dimensions[rn].height=20
            bg="92400E" if ri%2==0 else "FAC775"
            tx=WHITE   if ri%2==0 else "412402"
            for ci,(_,_,key) in enumerate(GOLD_COLS,1):
                val=_g(stk,key); cell=ws.cell(rn,ci,val)
                cell.fill=_f(bg); cell.font=_ft(False,tx,9)
                cell.alignment=_al("left" if ci in {2,3,N} else "center","center",wrap=(ci==N))
                cell.border=Border(
                    left=Side(style="thin",color="E2E8F0"),
                    right=Side(style="thin",color="E2E8F0"),
                    top=Side(style="thin",color="E2E8F0"),
                    bottom=Side(style="thin",color="E2E8F0")
                )

    def _trade_summary(self,wb):
        ws=wb.create_sheet("📊 Trade Summary"); ws.sheet_properties.tabColor="059669"
        gdf=self._get_gold()
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=17)
        c=ws.cell(1,1,f"GOLD STOCKS — TRADE PLAN SUMMARY  ·  Entry / SL / Targets / R:R Ratio  ·  {self.dlbl}")
        c.fill=_f("059669"); c.font=_ft(True,WHITE,11); c.alignment=_al()
        ws.row_dimensions[1].height=26
        hdrs=[("Symbol",12),("Company",25),("CMP (₹)",11),("CFV (₹)",12),
              ("MoS %",10),("Upside %",10),("Chg% [2-Wk]",13),("Chg% [4-Wk]",13),
              ("Chg% [8-Wk]",13),("Entry Range (₹)",16),("Stop Loss (₹)",13),
              ("Target 1 (₹)",12),("Target 2 (₹)",12),("Target 3 (₹)",12),
              ("R:R Ratio",11),("Time Horizon",22),("Risk Level",11)]
        ws.row_dimensions[2].height=32
        for ci,(h,w) in enumerate(hdrs,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f("059669"); c.font=_ft(True,WHITE,9)
            c.alignment=_al("center","center",True)
        ws.freeze_panes="A3"
        for ri,stk in enumerate(gdf.to_dict("records")):
            rn=ri+3; ws.row_dimensions[rn].height=22
            bg="D1FAE5" if ri%2==0 else "ECFDF5"; tx="065F46"
            cmp=_g(stk,"close",0); cfv=_g(stk,"cfv",0)
            mos=_g(stk,"mos_pct",0); up=_g(stk,"upside",0)
            vals=[_g(stk,"symbol"),_g(stk,"company_name",""),
                  f"₹{cmp:,}" if isinstance(cmp,(int,float)) and cmp else cmp,
                  f"₹{cfv:,}" if isinstance(cfv,(int,float)) and cfv else cfv,
                  f"+{mos:.1f}%" if isinstance(mos,(int,float)) else mos,
                  f"+{up:.1f}%"  if isinstance(up,(int,float))  else up,
                  _g(stk,"2w_chg"),_g(stk,"4w_chg"),_g(stk,"8w_chg"),
                  _g(stk,"entry_range"),_g(stk,"stop_loss"),
                  _g(stk,"t1"),_g(stk,"t2"),_g(stk,"t3"),
                  None,_g(stk,"horizon"),_g(stk,"risk_level")]
            for ci,val in enumerate(vals,1):
                cell=ws.cell(rn,ci,val); cell.fill=_f(bg); cell.font=_ft(False,tx,9)
                cell.alignment=_al("center","center")
                if ci==11: cell.fill=_f("FEE2E2"); cell.font=_ft(False,"7F1D1D",9)
            # Compute R:R numerically — entry_range is text "353.8–364.7",
            # parse midpoint so Excel formula doesn't fail on text input
            try:
                _er = str(stk.get("entry_range","") or "")
                # Parse "low–high" or "low-high" range text
                for _sep in ["–","-","—","to"]:
                    if _sep in _er:
                        _parts = _er.split(_sep)
                        _entry_mid = (float(_parts[0].replace("₹","").strip()) +
                                      float(_parts[-1].replace("₹","").strip())) / 2
                        break
                else:
                    _entry_mid = float(_er.replace("₹","").strip() or 0)
                _sl_v  = float(str(stk.get("stop_loss","0") or 0).replace("₹","").replace(",",""))
                _t1_v  = float(str(stk.get("t1","0") or 0))
                if _entry_mid > 0 and _sl_v > 0 and _entry_mid > _sl_v and _t1_v > _entry_mid:
                    _rr_val = round((_t1_v - _entry_mid) / (_entry_mid - _sl_v), 2)
                else:
                    _rr_val = "—"
            except Exception:
                _rr_val = "—"
            rr=ws.cell(rn,15,_rr_val); rr.fill=_f(bg); rr.font=_ft(True,"065F46",9)
            rr.alignment=_al()
            if isinstance(_rr_val,(int,float)):
                rr.number_format="0.00"
                # Colour code R:R: green>2, amber 1-2, red<1
                if _rr_val >= 3:   rr.fill=_f("D1FAE5"); rr.font=_ft(True,"065F46",9)
                elif _rr_val >= 2: rr.fill=_f("FEF3C7"); rr.font=_ft(True,"92400E",9)
                else:              rr.fill=_f("FEE2E2"); rr.font=_ft(True,"7F1D1D",9)

    def _alert_log(self,wb):
        ws=wb.create_sheet("🔔 Alert Log"); ws.sheet_properties.tabColor="7C3AED"
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
        c=ws.cell(1,1,f"ALERT LOG  ·  Triggered Signals & Score Changes  ·  Generated {self.run_time}")
        c.fill=_f("7C3AED"); c.font=_ft(True,WHITE,11); c.alignment=_al()
        ws.row_dimensions[1].height=24
        hdrs=[("Date",12),("Time (IST)",11),("Symbol",12),("Alert Type",26),
              ("Trigger Detail",46),("Prev Score",11),("New Score",11),
              ("Score Δ",10),("Action Required",26),("Exchange",12)]
        ws.row_dimensions[2].height=28
        for ci,(h,w) in enumerate(hdrs,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f("7C3AED"); c.font=_ft(True,WHITE,9)
            c.alignment=_al("center","center",True)
        ws.freeze_panes="A3"
        ACOLS={"⭐ EARLY MOVER DETECTED":"FAC775","🔔 SPIKE FIRED":"D1FAE5",
               "💰 SMART MONEY ENTRY":"D1FAE5","📈 SECTOR STAGE CHANGE":"FAC775",
               "⬇ SCORE DEGRADED":"FEE2E2","🛑 EXIT ALERT":"FEE2E2"}
        ri=3
        for stk in self.df.to_dict("records"):
            sym  = stk.get("symbol","")
            spk  = int(stk.get("spike_count",0) or 0)
            early= _sf(stk.get("early_entry_score",0))
            comp = _sf(stk.get("composite_score",0))
            verd = str(stk.get("verdict","WATCHLIST"))
            mos  = _sf(stk.get("mos_pct", stk.get("upside", 0)))
            vol  = _sf(stk.get("vol_ratio", 1.0))

            if spk>=1 or early>=70 or comp<30:
                # ── Prev Score + Delta ─────────────────────────────────────
                prev = self.prev_scores.get(sym, None)
                if prev is not None and prev > 0:
                    prev_disp = f"{prev:.0f}"
                    delta     = comp - prev
                    delta_disp= f"+{delta:.0f}" if delta > 0 else f"{delta:.0f}"
                else:
                    prev_disp = "—"
                    delta_disp= "—"

                # ── Alert Type ─────────────────────────────────────────────
                if comp < 30:
                    at  = "⬇ SCORE DEGRADED"
                    det = f"Score {comp:.0f}/100 below threshold — review position"
                elif spk >= 1:
                    at  = "🔔 SPIKE FIRED"
                    det = f"Spike {spk}/6 | Score: {comp:.0f} | Early: {early:.0f}/100"
                else:
                    at  = "⭐ EARLY MOVER DETECTED"
                    det = f"Early Entry {early:.0f}/100 | Score: {comp:.0f}"

                # ── Action Required — logic based on verdict/score/MoS/Δ ──
                if comp < 30:
                    act = "REVIEW FOR EXIT"
                elif verd == "BUY" and mos > 10 and comp >= 65:
                    act = "CONSIDER ENTRY"
                elif verd == "BUY" and mos <= 0:
                    act = "BUY BUT OVERVALUED — WAIT"
                elif verd == "BUY":
                    act = "MONITOR FOR ENTRY"
                elif verd == "WATCHLIST" and delta_disp != "—" and float(delta_disp) >= 3:
                    act = "SCORE IMPROVING — WATCH"
                elif verd == "WATCHLIST" and delta_disp != "—" and float(delta_disp) <= -3:
                    act = "SCORE DECLINING — CAUTION"
                elif vol >= 3.0:
                    act = "VOLUME ALERT — INVESTIGATE"
                elif early >= 70:
                    act = "EARLY MOVER — ACCUMULATE"
                elif verd == "WATCHLIST":
                    act = "MONITOR CLOSELY"
                else:
                    act = "MONITOR CLOSELY"

                row_data=[self.dlbl, self.run_time, sym, at, det,
                          prev_disp, f"{comp:.0f}", delta_disp, act,
                          stk.get("exchange_tag","NSE")]
                ac=ACOLS.get(at,"FFFFFF"); ws.row_dimensions[ri].height=18
                for ci,val in enumerate(row_data,1):
                    cell=ws.cell(ri,ci,_sv(val))
                    if ci==4: cell.fill=_f(ac); cell.font=_ft(True,NAVY,9)
                    else: cell.fill=_f(LG); cell.font=_ft(False,NAVY,9)
                    cell.alignment=_al("left" if ci==5 else "center","center")
                ri+=1

    def _delivery_preview(self,wb):
        ws=wb.create_sheet("📱 Delivery Preview"); ws.sheet_properties.tabColor="0D9488"
        ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=72
        gdf=self._get_gold(); gc=len(gdf)
        top=gdf.iloc[0] if not gdf.empty else {}
        ws.row_dimensions[1].height=24
        c=ws.cell(1,1,"DELIVERY PREVIEW  ·  WhatsApp & Email Format  ·  Sent daily 20:30 IST")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,11)
        def R(r,a,b,bf=None,bb=False,h=16):
            ws.row_dimensions[r].height=h
            if a: ca=ws.cell(r,1,a); ca.font=_ft(True,NAVY,9)
            if b is not None:
                cb=ws.cell(r,2,b); cb.fill=_f(bf or LG)
                cb.font=_ft(bb,NAVY,10); cb.alignment=_al("left","center")
            return r+1
        r=3
        r=R(r,None,"WHATSAPP MESSAGE PREVIEW","0D9488",True,18)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"⭐  NSE/BSE GOLD STOCKS  ·  {self.dlbl}  ·  {self.run_time}","B45309",True)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"⭐  EARLY MOVERS TODAY: {gc} stocks identified","D1FAE5",False)
        if not gdf.empty:
            sym=top.get("symbol","—"); verd=top.get("verdict","—")
            cfv=top.get("cfv",0); cmp_=top.get("close",0)
            mos=top.get("mos_pct",0); up=top.get("upside",0)
            w8=top.get("8w_chg",0); w4=top.get("4w_chg",0); w2=top.get("2w_chg",0)
            spk=int(top.get("spike_count",0) or 0); rot=top.get("sector","—")
            r=R(r,None,f"📈  TOP PICK: {sym} — {verd}","FAC775",True)
            r=R(r,None,f"     CFV: ₹{cfv}  |  CMP: ₹{cmp_}  |  MoS: +{mos:.1f}%  |  Upside: +{up:.1f}%  |  Spike: {spk}/6","FEF3C7",False)
            r=R(r,None,f"     8-Week Chg: {w8}%  |  4-Week Chg: {w4}%  |  2-Week Chg: {w2}%","FEF3C7",False)
            r=R(r,None,f"🔔  SPIKE ALERTS: {spk}  |  🔁  SECTOR: {rot} — Stage {top.get('rotation_stage','—')}","D1FAE5",False)
        r=R(r,None,"━"*51,"0D9488",False,10)
        r=R(r,None,f"📎  NSE_BSE_Gold_EarlyMovers_{self.date_str}.xlsx  (attached)","DBEAFE",False)
        r=R(r,None,f"📎  NSE_BSE_Full_Dashboard_{self.date_str}.xlsx  (subscribers)","DBEAFE",False)
        r+=1; r=R(r,"EMAIL SUBJECT LINE PREVIEW",None,h=18)
        ts=top.get("symbol","—") if not gdf.empty else "—"
        tu=top.get("upside",0)   if not gdf.empty else 0
        r=R(r,None,f"NSE/BSE Research | {self.dlbl} | {gc} Early Movers | Top: {ts} +{tu:.1f}% Upside","EFF6FF",True)
        r+=1; r=R(r,"EXCEL TRIGGER COMMANDS",None,h=18)
        for cmd,desc in [('  "generate excel"',"→  Both files today"),
                         ('  "excel gold only"',"→  Gold file only"),
                         ('  "send to whatsapp"',"→  WhatsApp delivery"),
                         ('  "excel sector: Pharma"',"→  Sector-filtered export"),
                         ('  "refresh excel"',"→  Regenerate with latest data")]:
            ws.row_dimensions[r].height=15
            ws.cell(r,2,f"{cmd:<28}{desc}").font=_ft(False,NAVY,9); r+=1

    def _glossary(self,wb):
        ws=wb.create_sheet("📖 Glossary"); ws.sheet_properties.tabColor="475569"
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)
        c=ws.cell(1,1,"GLOSSARY  ·  Full Forms & Descriptions of All Column Abbreviations  ·  NSE/BSE Stock Analyser v6.0")
        c.fill=_f(NAVY); c.font=_ft(True,WHITE,12); c.alignment=_al()
        ws.row_dimensions[1].height=28
        ws.column_dimensions["A"].width=3
        for ci,(h,w) in enumerate([("Group",18),("Short Name / Abbreviation",28),
                                    ("Full Form & Description",70),("Where Used",16)],2):
            ws.column_dimensions[get_column_letter(ci)].width=w
            c=ws.cell(2,ci,h); c.fill=_f(NAVY); c.font=_ft(True,WHITE,9); c.alignment=_al()
            ws.row_dimensions[2].height=28
        ws.freeze_panes="B3"
        for ri,(grp,short,desc,where) in enumerate(GLOSSARY_DATA,3):
            ws.row_dimensions[ri].height=16
            bg=LG if ri%2==0 else WHITE
            gc=GRP_COLORS.get(grp,"475569")
            c=ws.cell(ri,2,grp); c.fill=_f(gc); c.font=_ft(True,WHITE,9); c.alignment=_al()
            c=ws.cell(ri,3,str(short) if short else ""); c.fill=_f(bg); c.font=_ft(True,NAVY,9); c.alignment=_al("left","center"); c.data_type="s"
            c=ws.cell(ri,4,str(desc) if desc else "");  c.fill=_f(bg); c.font=_ft(False,"475569",9); c.alignment=_al("left","center",True); c.data_type="s"
            c=ws.cell(ri,5,where); c.fill=_f(bg); c.font=_ft(False,NAVY,9);    c.alignment=_al()

    def _get_gold(self):
        if self.df.empty: return pd.DataFrame()
        try:
            mask=((self.df["early_entry_score"]>=70)|
                  ((self.df["mos_pct"]>=25)&(self.df["composite_score"]>=70)))&\
                 (~self.df["verdict"].isin(["AVOID","AVOID / EXIT","EXIT"]))&\
                 (self.df["spike_suppressed"]==False)
            return self.df[mask].copy().reset_index(drop=True)
        except: return pd.DataFrame()